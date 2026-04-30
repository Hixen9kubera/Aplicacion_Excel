"""
Agente Parser — Fase A de la arquitectura multi-agente FERRAFORME.

Responsabilidades:
- Leer y mapear columnas del Excel (analizar_encabezados)
- Leer productos (leer_productos)
- Corregir CBM (corregir_cbm, corregir_cbm_inner)
- Extraer y guardar imágenes (extraer_imagenes_excel, guardar_imagenes_temp)
- Extraer número de contenedor

Entrada : file_bytes (bytes del Excel), filename, respuestas_dudas (opcional)
Salida  : ParserResult (TypedDict con productos, imagenes, columnas, etc.)
"""
from __future__ import annotations

import io
import json
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import TypedDict

import openpyxl
from langchain_anthropic import ChatAnthropic
from langchain_core.messages import HumanMessage

from agents.utils import IMAGENES_TEMP_PATH, _safe_float, _sanitizar_nombre_archivo


# ── TypedDict de salida ────────────────────────────────────────────────────────

class ParserResult(TypedDict):
    analisis: dict
    contenedor: str
    dudas_relevantes: list[dict]
    dudas_menores: list[str]
    productos: list[dict]
    imagenes: dict          # {row_0idx: {data, ext}}
    columnas: dict
    fila_encabezado: int
    advertencias: list[str]


# ── Funciones individuales ─────────────────────────────────────────────────────

def extraer_contenedor(filename: str) -> str:
    """
    Extrae el número de contenedor del nombre del archivo.
    Busca el patrón ISO 6346: 4 letras seguidas de 6-7 dígitos (ej. TCKU1234567).
    Si no lo encuentra, devuelve el stem del archivo sin extensión.
    """
    nombre = Path(filename).stem
    match = re.search(r'([A-Z]{3,4})[\s_-]?(\d{5,8})', nombre.upper())
    if match:
        return match.group(1) + match.group(2)
    return nombre


def analizar_encabezados(file_bytes: bytes) -> dict:
    """Claude analiza los encabezados y devuelve mapeo + dudas estructuradas.
    Detecta automáticamente en qué fila comienza la tabla de productos,
    independientemente del formato (algunos packing lists tienen cabeceras
    de empresa en las primeras filas antes de la tabla real).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    filas_raw = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 35:
            break
        filas_raw.append([str(v) if v is not None else "" for v in row])

    filas_numeradas = [
        {"fila": i + 1, "valores": fila}
        for i, fila in enumerate(filas_raw)
    ]

    prompt = f"""Analiza este packing list de importación.
IMPORTANTE: Algunos packing lists tienen información de empresa/embarque en las primeras filas
antes de la tabla de productos. Identifica exactamente en qué fila (1-indexed) están los
encabezados de columna de la tabla de productos (la fila que dice cosas como "Description",
"Qty", "CBM", "Price", 数量, 品名, etc.).
Los encabezados pueden estar en cualquier idioma (chino, inglés, español, etc.).

Campos a identificar (índice base 0):
nombre_producto, nombre_producto_alt, id_guia, cajas_master, piezas_x_caja, piezas_total, cbm_por_pieza, cbm_master_carton, cbm_total_sku, precio_usd, largo_cm, ancho_cm, alto_cm, peso_kg, volumen_por_pieza, material, uso

IMPORTANTE para columnas CBM — usa los valores numéricos de las filas de muestra para inferir correctamente:

Definiciones:
- cbm_por_pieza:     CBM de UNA pieza individual         → valor típico 0.0001–0.05
- cbm_master_carton: CBM de UNA caja/master carton        → valor típico 0.02–0.5
- cbm_total_sku:     CBM TOTAL de todas las unidades SKU  → valor típico 0.1–20

Relaciones matemáticas para verificar con los datos de muestra:
  cbm_por_pieza × piezas_total  ≈  cbm_total_sku   (si cbm_por_pieza es por pieza)
  cbm_master_carton × cajas     ≈  cbm_total_sku   (si cbm_master_carton es por caja)
  cbm_master_carton / piezas_x_caja  ≈  cbm_por_pieza

Regla clave para columnas ambiguas (header solo dice "CBM" o "M³" sin especificar):
  1. Toma el valor de muestra de esa columna (ej: 0.085) y los valores de cajas y piezas_total.
  2. Calcula: valor × cajas  vs  valor × piezas_total
  3. El que se acerque al cbm_total (si existe) indica el tipo correcto.
  4. Si no hay cbm_total, usa la magnitud: si valor × piezas_total > 50 m³ es improbable
     para un contenedor estándar (máx ~76 m³) → probablemente es por caja, no por pieza.

Si hay más de una columna CBM, mapéalas todas. Nunca mapees la misma columna en dos campos distintos.

IMPORTANTE para nombre_producto y nombre_producto_alt:
- Si hay DOS columnas de nombre (ej: una en inglés y otra en chino), mapea la principal en nombre_producto y la alternativa en nombre_producto_alt.
- Si solo hay una columna de nombre, deja nombre_producto_alt con confianza "no_encontrado".

DATOS (filas del Excel con su número de fila real):
{json.dumps(filas_numeradas, ensure_ascii=False, indent=2)}

Responde ÚNICAMENTE con este JSON (sin texto adicional):
{{
  "idioma_detectado": "...",
  "fila_encabezado": 1,
  "numero_contenedor": "",
  "columnas": {{
    "nombre_producto":     {{"indice": 0,    "encabezado_original": "...", "valor_muestra": "...", "confianza": "alta",          "nota": ""}},
    "nombre_producto_alt": {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "nombre en otro idioma, si existe"}},
    "id_guia":             {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "número de guía o código de carga, opcional"}},
    "cajas_master":        {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "número de cajas/master cartons para este SKU"}},
    "piezas_x_caja":       {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "unidades por caja"}},
    "piezas_total":        {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "total piezas en el contenedor para ese SKU"}},
    "largo_cm":            {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": ""}},
    "ancho_cm":            {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": ""}},
    "alto_cm":             {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": ""}},
    "peso_kg":             {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "peso en kg por pieza o por caja (especificar en nota)"}},
    "volumen_por_pieza":   {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "volumen por pieza en m³, si viene separado del CBM"}},
    "cbm_por_pieza":       {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "CBM de una pieza individual"}},
    "cbm_master_carton":   {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "CBM de una caja/master carton completa"}},
    "cbm_total_sku":       {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": "CBM total de todas las unidades de este SKU"}},
    "precio_usd":     {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": ""}},
    "material":       {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": ""}},
    "uso":            {{"indice": null,  "encabezado_original": "",    "valor_muestra": "",    "confianza": "no_encontrado", "nota": ""}}
  }},
  "dudas_relevantes": [
    {{
      "id": 0,
      "tipo": "eleccion",
      "campo_afectado": "material",
      "descripcion": "Se encontraron dos columnas para 'material': col 11 en inglés y col 12 en chino. Parecen ser el mismo dato.",
      "pregunta": "¿Cuál columna de material deseas usar?",
      "opciones": ["Inglés — col 11 (MDF)", "Chino — col 12 (高密度板/MDF)"],
      "indices_opciones": [11, 12],
      "default": "Inglés — col 11 (MDF)"
    }}
  ],
  "dudas_menores": [
    "3 productos tienen precio USD vacío — quedarán en blanco en el Excel"
  ]
}}

NOTA sobre fila_encabezado:
- Pon el número de fila (1-indexed) donde están los encabezados de la tabla de productos.
- Si la fila 1 ya es el encabezado, pon 1.
- Si hay filas de información de empresa/embarque antes de la tabla, identifica la fila correcta.
- Los índices de columnas en "columnas" son SIEMPRE base 0 contando desde la izquierda del Excel,
  independientemente de en qué fila estén los encabezados.

NOTA sobre numero_contenedor:
- Busca en las filas previas a fila_encabezado (metadata del embarque) un número de contenedor.
- Patrones comunes: "Container No:", "CONT#", "Container:", "Contenedor:", "CTR", o un código
  alfanumérico que empiece con letras y tenga 11 caracteres (ej: TCKU1234567, MSCU9876543).
- Si encuentras más de uno, ponlos separados por coma.
- Si no encuentras ninguno, deja el campo vacío ("").

TIPOS DE DUDA RELEVANTE:
- "eleccion": el usuario debe elegir entre varias opciones (ej: qué columna usar cuando hay duplicados).
  Siempre pon como default la opción en INGLÉS si aplica.
- "confirmar": el usuario confirma o rechaza algo con Sí/No.
- "texto": el usuario necesita escribir un valor.

CRITERIO para dudas_relevantes: solo las que puedan causar un error o dato incorrecto en el Excel
(columna no encontrada, ambigüedad entre columnas, unidades dudosas, etc.).
CRITERIO para dudas_menores: observaciones informativas que no bloquean la generación
(celdas vacías ocasionales, campos opcionales ausentes, etc.).

Niveles de confianza: alta / media / baja / no_encontrado"""

    llm  = ChatAnthropic(model="claude-sonnet-4-6", max_tokens=4000)
    resp = llm.invoke([HumanMessage(content=prompt)])
    texto = resp.content.strip()
    if texto.startswith("```"):
        partes = texto.split("```")
        texto = partes[1].lstrip("json").strip() if len(partes) > 1 else texto
    if texto.endswith("```"):
        texto = texto[:-3].strip()
    return json.loads(texto)


def aplicar_respuestas(columnas: dict, dudas_relevantes: list, respuestas: dict) -> dict:
    """Actualiza el mapeo de columnas según las respuestas del usuario a las dudas."""
    columnas = dict(columnas)
    for duda in dudas_relevantes:
        duda_id   = str(duda["id"])
        respuesta = respuestas.get(duda_id)
        if duda["tipo"] == "eleccion" and respuesta and duda.get("indices_opciones"):
            try:
                idx_opcion   = duda["opciones"].index(respuesta)
                indice_nuevo = duda["indices_opciones"][idx_opcion]
                campo        = duda["campo_afectado"]
                if campo in columnas:
                    columnas[campo] = {**columnas[campo], "indice": indice_nuevo, "confianza": "alta"}
            except (ValueError, IndexError):
                pass
    return columnas


def leer_productos(file_bytes: bytes, columnas: dict, fila_encabezado: int = 1) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    def idx(campo):
        c = columnas.get(campo, {})
        if c.get("confianza") == "no_encontrado":
            return None
        i = c.get("indice")
        return int(i) if i is not None else None

    nombre_idx = idx("nombre_producto")
    if nombre_idx is None:
        return [], ["No se pudo identificar la columna de nombre de producto."]

    # ── Mapa de celdas combinadas ──────────────────────────────────────────────
    merged_values: dict[tuple[int, int], object] = {}
    secondary_merged_rows: set[int] = set()
    for rng in ws.merged_cells.ranges:
        top_val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_values[(r, c)] = top_val
        for r in range(rng.min_row + 1, rng.max_row + 1):
            secondary_merged_rows.add(r)

    _cbm_cols_1idx = {
        f: (idx(f) + 1)
        for f in ("cbm_master_carton", "cbm_total_sku", "cbm_por_pieza", "id_guia")
        if idx(f) is not None
    }

    def cell_val(row_num: int, col_0idx: int, row_vals: tuple):
        v = row_vals[col_0idx] if col_0idx < len(row_vals) else None
        if v is None:
            v = merged_values.get((row_num, col_0idx + 1))
        return v

    productos, advertencias = [], []
    primera_fila_datos = fila_encabezado + 1

    for row_num, row in enumerate(
        ws.iter_rows(min_row=primera_fila_datos, values_only=True),
        start=primera_fila_datos,
    ):
        if len(row) <= nombre_idx or cell_val(row_num, nombre_idx, row) is None:
            continue

        def val(campo):
            i = idx(campo)
            return cell_val(row_num, i, row) if i is not None else None

        prod = {
            "nombre":            str(val("nombre_producto") or "").strip() or None,
            "nombre_alt":        str(val("nombre_producto_alt") or "").strip() or None,
            "id_guia":           val("id_guia"),
            "cajas_master":      val("cajas_master"),
            "piezas_x_caja":     val("piezas_x_caja"),
            "piezas_total":      val("piezas_total"),
            "largo_cm":          val("largo_cm"),
            "ancho_cm":          val("ancho_cm"),
            "alto_cm":           val("alto_cm"),
            "cbm_por_pieza":     val("cbm_por_pieza"),
            "cbm_master_carton": val("cbm_master_carton"),
            "cbm_total_sku":     val("cbm_total_sku"),
            "precio_usd":        val("precio_usd"),
            "peso_kg":           val("peso_kg"),
            "volumen_por_pieza": val("volumen_por_pieza"),
            "material":          val("material") or "",
            "uso":               val("uso") or "",
            "fila_excel_0idx":   row_num - 1,
        }

        _CAMPOS_NUM = (
            "cajas_master", "piezas_x_caja", "piezas_total",
            "cbm_por_pieza", "cbm_master_carton", "cbm_total_sku",
            "precio_usd", "largo_cm", "ancho_cm", "alto_cm",
            "peso_kg", "volumen_por_pieza",
        )
        for _cn in _CAMPOS_NUM:
            _v = prod.get(_cn)
            if isinstance(_v, str):
                try:
                    prod[_cn] = float(_v.replace(",", ".").strip())
                except (ValueError, TypeError):
                    prod[_cn] = None

        nombre = prod["nombre"]
        if row_num not in secondary_merged_rows:
            if prod["precio_usd"] is None:
                advertencias.append(f"Fila {row_num} ({nombre}): precio USD vacío")
            if not any(prod.get(c) for c in ("cbm_por_pieza", "cbm_master_carton", "cbm_total_sku")):
                advertencias.append(f"Fila {row_num} ({nombre}): CBM vacío")

        productos.append(prod)

    # Filtrar filas de totales/notas
    _PREFIJOS_NO_PROD = (
        "total", "country of origin", "declaration", "declare",
        "subtotal", "grand total", "注意", "备注", "合计",
    )

    def _es_fila_footer(prod: dict) -> bool:
        nom = str(prod.get("nombre") or "").strip().lower()
        return any(nom.startswith(p) for p in _PREFIJOS_NO_PROD)

    productos = [p for p in productos if not _es_fila_footer(p)]

    _CAMPOS_CLAVE = ("cajas_master", "piezas_total", "piezas_x_caja",
                     "cbm_por_pieza", "cbm_master_carton", "cbm_total_sku")
    ultimo_completo = -1
    for i, p in enumerate(productos):
        if any(p.get(c) for c in _CAMPOS_CLAVE):
            ultimo_completo = i
    if ultimo_completo >= 0:
        productos = productos[: ultimo_completo + 1]

    return productos, advertencias


def corregir_cbm(productos: list[dict], advertencias: list[str]) -> list[dict]:
    """
    Detecta si se mapeó cbm_master_carton como cbm_por_pieza y corrige.
    """
    _TOL = 0.10

    for prod in productos:
        cbm_pz  = _safe_float(prod.get("cbm_por_pieza"))
        cbm_mc  = _safe_float(prod.get("cbm_master_carton"))
        cbm_tot = _safe_float(prod.get("cbm_total_sku"))
        cajas   = _safe_float(prod.get("cajas_master"))
        pzs_tot = _safe_float(prod.get("piezas_total"))
        pzs_cja = _safe_float(prod.get("piezas_x_caja"))

        if cbm_pz <= 0 or cbm_tot <= 0:
            continue

        ratio_pieza = (cbm_pz * pzs_tot) / cbm_tot if pzs_tot > 0 else None
        ratio_caja  = (cbm_pz * cajas)   / cbm_tot if cajas   > 0 else None

        err_pieza = abs(ratio_pieza - 1) if ratio_pieza is not None else 999
        err_caja  = abs(ratio_caja  - 1) if ratio_caja  is not None else 999

        if err_caja < _TOL and err_pieza > _TOL:
            nombre = prod.get("nombre", "?")
            cbm_pz_corr = round(cbm_pz / pzs_cja, 6) if pzs_cja > 0 else None
            advertencias.append(
                f"CBM corregido en '{nombre}': "
                f"la columna CBM ({cbm_pz}) era por carton, no por pieza. "
                f"cbm_por_pieza = {cbm_pz_corr}"
            )
            prod["cbm_master_carton"] = cbm_pz
            if not prod.get("cbm_inner_carton"):
                prod["cbm_inner_carton"] = cbm_pz
            prod["cbm_por_pieza"] = cbm_pz_corr

    return productos


def corregir_cbm_inner(productos: list[dict], advertencias: list[str],
                       file_bytes: bytes, columnas: dict) -> list[dict]:
    """
    Detecta productos que comparten un master carton (celdas combinadas o mismo id_guia)
    y calcula cbm_inner_carton y cbm_por_pieza correctamente.
    """
    if not productos:
        return productos

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    def _idx(campo):
        c = columnas.get(campo, {})
        if c.get("confianza") == "no_encontrado":
            return None
        i = c.get("indice")
        return (int(i) + 1) if i is not None else None  # 1-indexed

    cbm_mc_col = _idx("cbm_master_carton") or _idx("cbm_total_sku") or _idx("cbm_por_pieza")

    # ── Método 1: celdas combinadas en la columna CBM ─────────────────────────
    grupos_merged: dict[int, list[int]] = {}
    if cbm_mc_col:
        for rng in ws.merged_cells.ranges:
            if rng.min_col <= cbm_mc_col <= rng.max_col:
                top = rng.min_row
                for r in range(rng.min_row, rng.max_row + 1):
                    grupos_merged.setdefault(top, [])

        fila_a_prod = {}
        for i, p in enumerate(productos):
            fila = p.get("fila_excel_0idx")
            if fila is not None:
                fila_a_prod[fila + 1] = i

        for rng in ws.merged_cells.ranges:
            if cbm_mc_col and rng.min_col <= cbm_mc_col <= rng.max_col:
                top = rng.min_row
                for r in range(rng.min_row, rng.max_row + 1):
                    if r in fila_a_prod:
                        grupos_merged.setdefault(top, []).append(fila_a_prod[r])

    # ── Método 2: mismo id_guia no vacío ─────────────────────────────────────
    grupos_id: dict[str, list[int]] = {}
    for i, p in enumerate(productos):
        gid = str(p.get("id_guia") or "").strip()
        if gid:
            grupos_id.setdefault(gid, []).append(i)

    grupos_finales: list[list[int]] = []
    usados: set[int] = set()

    for indices in list(grupos_merged.values()) + list(grupos_id.values()):
        indices = [i for i in indices if i not in usados]
        if len(indices) > 1:
            grupos_finales.append(indices)
            usados.update(indices)

    for grupo in grupos_finales:
        n = len(grupo)
        cbm_master = next(
            (float(productos[i].get("cbm_master_carton") or
                   productos[i].get("cbm_total_sku") or
                   productos[i].get("cbm_por_pieza") or 0)
             for i in grupo), 0
        )
        if cbm_master <= 0:
            continue

        cbm_inner = round(cbm_master / n, 6)
        nombres = []
        for i in grupo:
            p = productos[i]
            pzs_caja = float(p.get("piezas_x_caja") or 0)
            cbm_pz   = round(cbm_inner / pzs_caja, 6) if pzs_caja > 0 else None
            p["cbm_master_carton"] = cbm_master
            p["cbm_inner_carton"]  = cbm_inner
            p["cbm_por_pieza"]     = cbm_pz
            nombres.append(p.get("nombre", "?"))

        advertencias.append(
            f"Master carton compartido ({n} productos): {', '.join(nombres)}. "
            f"CBM master={cbm_master} → CBM inner={cbm_inner} por producto"
        )

    return productos


def extraer_imagenes_excel(file_bytes: bytes) -> dict[int, dict]:
    """
    Extrae imágenes flotantes del Excel leyendo el XML de posicionamiento.
    Devuelve dict {row_0indexed: {data, ext}}.
    """
    NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    NS_A   = "http://schemas.openxmlformats.org/drawingml/2006/main"
    NS_R   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    resultado: dict[int, dict] = {}

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            nombres = set(z.namelist())

            drawing_paths = sorted(
                n for n in nombres
                if re.match(r"xl/drawings/drawing\d+\.xml$", n)
            )

            for drawing_path in drawing_paths:
                parts      = drawing_path.rsplit("/", 1)
                rels_path  = f"{parts[0]}/_rels/{parts[1]}.rels"
                if rels_path not in nombres:
                    continue

                rels_tree  = ET.fromstring(z.read(rels_path))
                rid_a_ruta = {}
                for rel in rels_tree:
                    rid    = rel.get("Id")
                    target = rel.get("Target", "")
                    if "media" in target:
                        ruta = target.replace("../", "xl/")
                        rid_a_ruta[rid] = ruta

                drawing_tree = ET.fromstring(z.read(drawing_path))
                for anchor in drawing_tree:
                    row_el = anchor.find(f"{{{NS_XDR}}}from/{{{NS_XDR}}}row")
                    if row_el is None:
                        continue
                    row_num = int(row_el.text)

                    blip = anchor.find(
                        f".//{{{NS_XDR}}}pic/{{{NS_XDR}}}blipFill/{{{NS_A}}}blip"
                    )
                    if blip is None:
                        continue
                    rid = blip.get(f"{{{NS_R}}}embed")
                    if not rid or rid not in rid_a_ruta:
                        continue

                    ruta = rid_a_ruta[rid]
                    if ruta not in nombres:
                        continue
                    ext = Path(ruta).suffix.lower().lstrip(".")
                    if ext not in ("png", "jpg", "jpeg", "gif", "bmp", "webp"):
                        continue

                    if row_num not in resultado:
                        resultado[row_num] = {"data": z.read(ruta), "ext": ext}

    except Exception:
        pass

    return resultado


def guardar_imagenes_temp(imagenes: dict[int, dict], productos: list[dict]) -> tuple[int, list[str]]:
    """
    Guarda las imágenes extraídas del Excel en IMAGENES_TEMP_PATH.
    Almacena en prod["imagen_temp_stem"] y prod["imagen_temp_path"].
    Devuelve (n_guardadas, errores).
    """
    errores: list[str] = []
    try:
        IMAGENES_TEMP_PATH.mkdir(exist_ok=True)
    except Exception as e:
        return 0, [f"No se pudo crear carpeta temporal: {e}"]

    fila_a_idx = {p.get("fila_excel_0idx"): i for i, p in enumerate(productos)
                  if p.get("fila_excel_0idx") is not None}

    guardadas = 0
    for row_num, img in imagenes.items():
        prod_idx = fila_a_idx.get(row_num)
        if prod_idx is None:
            continue
        nombre = productos[prod_idx].get("nombre", f"producto_{prod_idx + 1}")
        nombre_archivo = f"{row_num}_{_sanitizar_nombre_archivo(nombre)}"
        ruta = IMAGENES_TEMP_PATH / f"{nombre_archivo}.{img['ext']}"
        try:
            ruta.write_bytes(img["data"])
            productos[prod_idx]["imagen_temp_stem"] = nombre_archivo
            productos[prod_idx]["imagen_temp_path"] = str(ruta)
            guardadas += 1
        except Exception as e:
            errores.append(f"No se pudo guardar imagen de '{nombre}': {e}")
    return guardadas, errores


def renombrar_imagenes_con_sku(productos: list[dict]) -> tuple[int, list[str]]:
    """
    Renombra las imágenes en IMAGENES_TEMP_PATH usando el SKU definitivo del producto.
    Devuelve (n_renombradas, errores).
    """
    if not IMAGENES_TEMP_PATH.exists():
        return 0, ["La carpeta de imágenes temporales no existe — no se extrajeron imágenes del Excel."]

    errores: list[str] = []
    renombradas = 0
    for prod in productos:
        stem_buscado = prod.get("imagen_temp_stem")
        sku          = prod.get("sku", "")
        if not stem_buscado or not sku:
            continue
        encontrado = False
        for archivo in IMAGENES_TEMP_PATH.iterdir():
            if archivo.stem == stem_buscado:
                nueva_ruta = archivo.with_name(f"{sku}{archivo.suffix}")
                if nueva_ruta.exists():
                    try:
                        archivo.unlink()
                    except Exception:
                        pass
                    prod["imagen_temp_stem"] = sku
                    encontrado = True
                    break
                try:
                    archivo.rename(nueva_ruta)
                    prod["imagen_temp_stem"] = sku
                    renombradas += 1
                except Exception as e:
                    errores.append(f"No se pudo renombrar imagen '{stem_buscado}' → '{sku}': {e}")
                encontrado = True
                break
        if not encontrado:
            errores.append(f"No se encontró imagen para SKU '{sku}' (buscada: {stem_buscado})")
    return renombradas, errores


# ── Nodo LangGraph ─────────────────────────────────────────────────────────────

def agente_parser(
    file_bytes: bytes,
    filename: str,
    respuestas_dudas: dict | None = None,
) -> ParserResult:
    """
    Nodo LangGraph — orquesta el pipeline completo de parsing.

    Primera llamada (respuestas_dudas=None):
      - Ejecuta analizar_encabezados y devuelve con dudas_relevantes llenas.
      - Si hay dudas, productos=[] y el caller debe recoger las respuestas del usuario.

    Segunda llamada (respuestas_dudas=dict):
      - Aplica las respuestas, lee productos, corrige CBM, extrae imágenes.
      - Devuelve ParserResult completo.
    """
    # Paso 1: analizar encabezados (siempre)
    analisis    = analizar_encabezados(file_bytes)
    contenedor  = extraer_contenedor(filename)
    _cont_excel = analisis.get("numero_contenedor", "")
    if _cont_excel and not contenedor:
        contenedor = _cont_excel

    dudas  = [d for d in analisis.get("dudas_relevantes", []) if isinstance(d, dict)]
    menores = [d for d in analisis.get("dudas_menores", [])    if isinstance(d, str)]

    if dudas and respuestas_dudas is None:
        # Pausar: devolver resultado parcial para que el caller pida respuestas al usuario
        return ParserResult(
            analisis=analisis,
            contenedor=contenedor,
            dudas_relevantes=dudas,
            dudas_menores=menores,
            productos=[],
            imagenes={},
            columnas=analisis.get("columnas", {}),
            fila_encabezado=analisis.get("fila_encabezado", 1),
            advertencias=[],
        )

    # Paso 2: aplicar respuestas del usuario (si las hay)
    columnas = analisis.get("columnas", {})
    if respuestas_dudas:
        columnas = aplicar_respuestas(columnas, dudas, respuestas_dudas)
    fila_enc = analisis.get("fila_encabezado", 1)

    # Paso 3: leer y corregir productos
    advertencias: list[str] = []
    productos, advs = leer_productos(file_bytes, columnas, fila_encabezado=fila_enc)
    advertencias.extend(advs)
    productos = corregir_cbm(productos, advertencias)
    productos = corregir_cbm_inner(productos, advertencias, file_bytes, columnas)

    # Paso 4: extraer y guardar imágenes
    imagenes = extraer_imagenes_excel(file_bytes)
    guardar_imagenes_temp(imagenes, productos)

    return ParserResult(
        analisis=analisis,
        contenedor=contenedor,
        dudas_relevantes=dudas,
        dudas_menores=menores,
        productos=productos,
        imagenes=imagenes,
        columnas=columnas,
        fila_encabezado=fila_enc,
        advertencias=advertencias,
    )
