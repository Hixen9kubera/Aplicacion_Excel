"""
app_v2.py — Agente conversacional FERRAFORME v2
Uso: streamlit run app_v2.py

Mejoras vs v1:
- Dudas relevantes se resuelven con botones/radio ANTES de generar el Excel
- Dudas menores se muestran como avisos informativos
- Prioridad automática a columnas en inglés cuando hay duplicados
- Migrado a LangChain + LangSmith (desactivado hasta producto final)
"""

import io
import os
import re
import json
import math
import pickle
import zipfile
import base64
import xml.etree.ElementTree as ET
import xmlrpc.client
from datetime import datetime, timedelta
from pathlib import Path
from typing import List

import difflib
import concurrent.futures
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
try:
    from PIL import Image as PILImage
    _PILLOW_OK = True
except ImportError:
    _PILLOW_OK = False

from langchain_anthropic import ChatAnthropic
from langchain_core.tools import tool
from langchain_core.messages import HumanMessage, AIMessage, SystemMessage, ToolMessage

# ── Agentes (arquitectura multi-agente) ───────────────────────────────────────
from agents.utils import (
    _safe_float, _sanitizar_nombre_archivo,
    IMAGENES_TEMP_PATH, CACHE_PATH, CACHE_MAX_HORAS,
)
from agents.parser_agent import (
    extraer_contenedor, analizar_encabezados, aplicar_respuestas,
    leer_productos, corregir_cbm, corregir_cbm_inner,
    extraer_imagenes_excel, guardar_imagenes_temp, renombrar_imagenes_con_sku,
    agente_parser,
)
from agents.vision_agent import (
    analizar_imagen_claude, _inferir_categoria_manual,
)
from agents.nombres_agent import (
    _tiene_chino, normalizar_nombres_productos,
    _det_iguales, _nombre_norm, _sim_chino,
    _PALABRAS_VERDE, _PALABRAS_AMARILLO, _PALABRAS_ROJO, _PALABRAS_VARIANTE,
    _es_tecnico, _nombre_base_verde, _nombre_base_amarillo,
    _tiene_diferencia_roja, _nombres_son_similares,
    _diffs_display, _CAMPOS_DETERMINANTES, _CAMPOS_DIFF_DISPLAY, _LABELS_DIFF,
    detectar_productos_duplicados, aplicar_resolucion_duplicados,
    agente_nombres,
)
from agents.utils import _phash_imagen, _similitud_nombres
from agents.graph import ejecutar_vision_y_nombres, ejecutar_solo_nombres, ejecutar_vision_bloque_desde_disco
from agents.excel_agent import (
    generar_excel, generar_excel_master, generar_excel_purchase,
    generar_reporte_clasificacion,
    _cbm_por_pieza, _cbm_total_fila,
)
from agents.odoo_agent import (
    guardar_cache_odoo, cargar_cache_odoo, info_cache_odoo,
    _cargar_desde_supabase, _sincronizar_a_supabase,
    buscar_similares_odoo, _buscar_padre_en_odoo_por_nombre,
    cargar_skus_odoo, cargar_todos_productos_odoo, cargar_detalle_productos_odoo,
    aplicar_resoluciones_conflictos, validar_sku_vs_odoo,
    _subir_productos_a_odoo, _reintentar_skus_fallidos,
    cargar_productos_por_contenedor,
)

# ── Cargar .env ────────────────────────────────────────────────────────────────
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

# LangSmith activado — cambiar a "false" para desactivar
os.environ["LANGCHAIN_TRACING_V2"] = "true"

# ── Configuración fija ─────────────────────────────────────────────────────────
EMPRESA       = "FERRAFORME MS"

# ── Códigos válidos de SKU (base_sku_fixed) ────────────────────────────────────
SUBCATEGORIAS = {
    # MUEBLES - HOGAR
    "MUE": "Muebles Hogar",   "MES": "Mesa",          "SIL": "Silla",
    "CAM": "Cama",            "EST": "Estantería",     "ORG": "Organizador",
    "COM": "Comedor",         "ESCR": "Escritorio",    "BAN": "Mueble Baño",
    "JAR": "Mueble Jardín",   "TV": "Mueble TV",       "COC": "Mueble Cocina",
    "DEC": "Decoración Hogar","ILUM": "Iluminación",   "TEX": "Textiles Hogar",
    # BEBÉS - INFANTIL
    "BEB": "Artículos Bebé",  "CUNA": "Cuna y Catre",  "PAS": "Paseo Bebé",
    "CORR": "Corral y Andadera","ALIM": "Alimentación Bebé","HIG": "Higiene Bebé",
    "ROBB": "Ropa Bebé",      "SEG": "Seguridad Bebé", "JUG": "Juguetes Bebé",
    # JUGUETES - JUEGOS
    "JUGU": "Juguetes",       "MUN": "Muñeca y Figura","PEL": "Peluche",
    "VEH": "Vehículo Juguete","CONS": "Juego Construcción","JUEG": "Juego de Mesa",
    "CART": "Cartas y Mazo",  "ELEC": "Juguete Electrónico","CAS": "Casa de Juguete",
    "MONT": "Juguete Montable","DEPO": "Juguete Deportivo","EDU": "Juguete Educativo",
    # MODA Y ROPA
    "ROP": "Ropa",            "CALZ": "Calzado",       "ACC": "Accesorio Moda",
    # TECNOLOGÍA
    "TEC": "Electrónica",     "CEL": "Celular y Accesorios",
    # HERRAMIENTAS Y OFICINA
    "HERR": "Herramienta",    "OFI": "Oficina",
    # MASCOTAS
    "MASC": "Mascota",
    # VARIOS
    "LIB": "Libro y Papelería","ART": "Arte y Manualidad","DEP": "Deporte y Fitness",
    "VIA": "Viaje y Equipaje", "VAR": "Varios",
}
ATRIBUTOS = {
    # COLORES
    "NEG": "Negro",   "BLN": "Blanco",  "GRI": "Gris",    "ROJ": "Rojo",
    "AZL": "Azul",    "VER": "Verde",   "AMA": "Amarillo","ROS": "Rosa",
    "NAR": "Naranja", "MOR": "Morado",  "CAF": "Café",    "BEI": "Beige",
    "MUL": "Multicolor","PLA": "Plateado","DOR": "Dorado",
    # TALLAS
    "XS": "Extra Small","S": "Small",  "M": "Medium",    "L": "Large",
    "XL": "Extra Large","XXL": "Extra Extra Large","3XL": "3XL","UNI": "Talla Única",
    # MATERIALES
    "MAD": "Madera",  "MET": "Metal",   "TEL": "Tela",    "CUE": "Cuero",
    # OTROS
    "EST": "Estándar","PRE": "Premium", "ECO": "Económico","INF": "Infantil",
    "ADU": "Adulto",
}
_SUBCAT_DEFAULT = "VAR"
_ATTR_DEFAULT   = "EST"



# ── Herramientas del agente ────────────────────────────────────────────────────
# ── System prompt ──────────────────────────────────────────────────────────────
def build_system_prompt(analisis, productos, tipo_cambio, contenedor, respuestas_dudas=None):
    n_prods = len(productos) if productos else 0
    prompt = f"""Eres el agente FERRAFORME, especializado en analizar packing lists de importación \
y generar plantillas de productos Excel para la empresa FERRAFORME MS.

CONFIGURACIÓN ACTUAL:
- Tipo de cambio: ${tipo_cambio:.2f} MXN/USD
- Contenedor: {contenedor}
- Empresa: FERRAFORME MS

ARCHIVO YA PROCESADO:
El packing list ya fue subido y procesado por completo. Los {n_prods} productos están disponibles \
en la sección PRODUCTOS de este mismo mensaje. NUNCA le pidas al usuario que vuelva a subir el \
archivo ni digas que no tienes acceso a él. Todo el contexto necesario está aquí.

HERRAMIENTA DISPONIBLE — generar_excel:
Llámala cuando el usuario confirme proceder o pida generar el Excel.
NO necesitas pasar ningún argumento — los productos ya están almacenados en sesión.

REGLAS DE COMPORTAMIENTO:
1. Las dudas relevantes ya fueron resueltas por el usuario antes de llegar aquí.
2. Cuando el usuario confirme que puede proceder, llama a generar_excel.
3. Si el usuario pide una corrección, modifica los datos y vuelve a llamar generar_excel.
4. Aprende de las correcciones del usuario en esta sesión.
5. Responde siempre en español, de forma concisa y profesional.
6. En tus respuestas y resúmenes NO menciones el tipo de cambio aplicado.
7. Si el usuario pide datos de un producto específico, búscalos en la lista PRODUCTOS de abajo."""

    if respuestas_dudas:
        prompt += f"\n\n--- RESPUESTAS DEL USUARIO A LAS DUDAS ---\n{json.dumps(respuestas_dudas, ensure_ascii=False, indent=2)}"

    if analisis:
        prompt += f"\n\n--- ANÁLISIS DEL PACKING LIST ---\n{json.dumps(analisis, ensure_ascii=False, indent=2)}"

    if productos:
        prompt += f"\n\n--- PRODUCTOS ({len(productos)}) ---\n{json.dumps(productos, ensure_ascii=False, indent=2)}"

    return prompt


def _build_system_fase(fase: str, productos: list, analisis: dict | None = None,
                        extra: str = "") -> str:
    """
    System prompt para las fases intermedias (dudas, duplicados, conflictos).
    Incluye datos completos del packing list para que el agente pueda responder
    cualquier pregunta sobre los productos.
    Al final de cada respuesta debe preguntar si el usuario tiene otra duda
    o si ya puede continuar con el proceso.
    """
    n = len(productos) if productos else 0
    prompt = (
        f"Eres el agente FERRAFORME. En este momento el usuario está en la fase: **{fase}**.\n\n"
        "INSTRUCCIONES:\n"
        "1. Responde la pregunta o aclaración del usuario de forma concisa y en español.\n"
        "2. Si el usuario menciona una corrección, confírmale que la tomará en cuenta.\n"
        "3. Al final de CADA respuesta tuya, pregunta siempre:\n"
        '   "¿Tienes alguna otra duda o ya puedo continuar con el proceso?"\n'
        "4. Tienes acceso completo a todos los datos del packing list — NUNCA digas que no "
        "puedes ver el archivo o los productos.\n"
        "5. Si el usuario dice que ya puede continuar, responde brevemente con un mensaje "
        "de confirmación positivo (ej. 'Perfecto, ¡continuamos!'). "
        "El botón para avanzar sigue visible en la pantalla.\n"
    )
    if extra:
        prompt += f"\nCONTEXTO ADICIONAL:\n{extra}\n"
    if analisis:
        prompt += f"\nANÁLISIS DEL PACKING LIST:\n{json.dumps(analisis, ensure_ascii=False, indent=2)}\n"
    if productos:
        # Limitar a 30 productos para no saturar el contexto
        muestra = productos[:30]
        prompt += f"\nPRODUCTOS ({n} en total, mostrando primeros {len(muestra)}):\n"
        prompt += json.dumps(muestra, ensure_ascii=False, indent=2)
    return prompt


def sincronizar_contadores_con_odoo(skus_odoo: list[str]) -> None:
    """
    Actualiza sku_contadores en session_state para que los nuevos SKUs
    partan del máximo que ya existe en ODOO por subcategoría.
    """
    contadores = st.session_state.setdefault("sku_contadores", {})
    patron = re.compile(r"^([A-Z]{2,4})-(\d{4})-")
    for sku in skus_odoo:
        m = patron.match(sku)
        if m:
            sub, num = m.group(1), int(m.group(2))
            if contadores.get(sub, 0) < num:
                contadores[sub] = num




# ══════════════════════════════════════════════════════════════════════════════
# AGENTE SKU — Extracción de imágenes + Claude Vision + generación de SKU
# ══════════════════════════════════════════════════════════════════════════════




def generar_sku(subcategoria_cod: str, atributo_cod: str) -> str:
    """Genera un SKU único con formato SUBCAT-####-ATRIBUTO usando contadores en session_state.
    Valida que los códigos estén en las listas definidas; usa VAR/EST como fallback.
    """
    sub = subcategoria_cod.strip().upper() if subcategoria_cod else _SUBCAT_DEFAULT
    att = atributo_cod.strip().upper()     if atributo_cod     else _ATTR_DEFAULT
    if sub not in SUBCATEGORIAS:
        sub = _SUBCAT_DEFAULT
    if att not in ATRIBUTOS:
        att = _ATTR_DEFAULT
    contadores = st.session_state.setdefault("sku_contadores", {})
    contadores[sub] = contadores.get(sub, 0) + 1
    sku = f"{sub}-{contadores[sub]:04d}-{att}"
    if st.session_state.get("modo_prueba"):
        sku += "_test"
    return sku


def _sku_mismo_numero(sub: str, numero: int, atributo_cod: str) -> str:
    """Construye un SKU con el mismo número base que otro producto del grupo variante.
    NO incrementa el contador — solo cambia el atributo/color.
    """
    att = atributo_cod.strip().upper() if atributo_cod else _ATTR_DEFAULT
    if att not in ATRIBUTOS:
        att = _ATTR_DEFAULT
    sku = f"{sub}-{numero:04d}-{att}"
    if st.session_state.get("modo_prueba"):
        sku += "_test"
    return sku








def _reportar_subida_imagenes(subidas: list[str], errores: list[str]) -> None:
    """Agrega al chat un resumen de imágenes subidas a productos en ODOO."""
    lineas = []
    if subidas:
        lineas.append(f"**Imágenes subidas a Odoo ({len(subidas)}):**")
        for sku in subidas:
            lineas.append(f"- ✅ `{sku}`")
    if errores:
        lineas.append(f"\n**Problemas al subir imágenes ({len(errores)}):**")
        for e in errores:
            lineas.append(f"- ❌ {e}")
    if lineas:
        st.session_state.chat.append({"role": "assistant", "content": "\n".join(lineas)})


def _reportar_errores_imagenes(errores: list[str], etapa: str) -> None:
    """Agrega errores de imagen al chat y a session_state para que sean visibles."""
    if not errores:
        return
    lineas = "\n".join(f"- {e}" for e in errores)
    msg = f"⚠️ **Problemas en {etapa}:**\n{lineas}"
    st.session_state.setdefault("imagenes_errores", []).extend(errores)
    st.session_state.chat.append({"role": "assistant", "content": msg})




def enriquecer_con_imagenes(file_bytes: bytes, productos: list[dict]) -> tuple[list[dict], int, list[dict]]:
    """
    Extrae imágenes del Excel, analiza cada una con Claude Haiku Vision,
    genera SKU validado contra ODOO, busca similares por imagen/nombre y enriquece los productos.

    Devuelve (productos_enriquecidos, n_procesadas, conflictos).
    """
    imagenes       = extraer_imagenes_excel(file_bytes)
    skus_odoo      = st.session_state.get("odoo_skus", [])
    prods_odoo     = st.session_state.get("odoo_productos", [])
    phashes_odoo   = st.session_state.get("odoo_phashes", {})

    if skus_odoo:
        sincronizar_contadores_con_odoo(skus_odoo)

    if not imagenes:
        return productos, 0, []

    # Guardar imágenes en carpeta temporal con nombre del producto
    _, errores_guardado = guardar_imagenes_temp(imagenes, productos)
    if errores_guardado:
        st.session_state.setdefault("imagenes_errores", []).extend(errores_guardado)

    procesadas  = 0
    conflictos  = []
    errores_gemini: list[str] = []
    skus_ya_en_conflicto: set[int] = set()
    # Grupos de variantes: gv_id → (sub, numero) del primer producto procesado
    grupos_variante_sku: dict[str, tuple[str, int]] = {}

    for i, prod in enumerate(productos):
        # Usar fila_excel_0idx para buscar la imagen — coincide con las claves de
        # extraer_imagenes_excel (0-indexed desde el XML del Excel).
        # NO usar i+1 porque si el header no está en fila 1 los índices desfasan.
        row_num = prod.get("fila_excel_0idx")
        if row_num is None:
            row_num = i + 1  # fallback para productos sin fila conocida (ej. Tab 2)
        img = imagenes.get(row_num)
        if img is None:
            prod.setdefault("sin_imagen", True)
            prod.setdefault("fila_sin_imagen", row_num + 1)
            continue

        contexto = {
            "nombre":   prod.get("nombre"),
            "material": prod.get("material"),
            "uso":      prod.get("uso"),
            "largo_cm": prod.get("largo_cm"),
            "ancho_cm": prod.get("ancho_cm"),
            "alto_cm":  prod.get("alto_cm"),
        }
        datos = analizar_imagen_claude(img["data"], img["ext"], contexto)
        if datos.get("_error"):
            errores_gemini.append(f"Fila {row_num + 1} ({prod.get('nombre', '?')}): {datos['_error']}")

        sub_cod = datos.get("subcategoria_cod", "VAR")
        att_cod = datos.get("atributo_cod", "EST")
        gv = prod.get("_grupo_variante")
        if gv and gv in grupos_variante_sku:
            # Variante de un grupo ya numerado → mismo sub y número, solo cambia atributo
            base_sub, base_num = grupos_variante_sku[gv]
            sku_inicial = _sku_mismo_numero(base_sub, base_num, att_cod)
        else:
            sku_inicial = generar_sku(sub_cod, att_cod)
            if gv:
                # Primer producto del grupo → guardar su sub y número para los demás
                partes = sku_inicial.replace("_test", "").split("-")
                if len(partes) >= 2:
                    try:
                        grupos_variante_sku[gv] = (partes[0], int(partes[1]))
                    except (ValueError, IndexError):
                        pass

        conflicto_entry = None

        if skus_odoo:
            # ── 1. Conflicto por prefijo SKU ──────────────────────────────────
            validacion = validar_sku_vs_odoo(sku_inicial, skus_odoo)
            sku_final  = validacion["sku_ajustado"]
            if validacion.get("nuevo_num"):
                _m_sub = re.match(r"^([A-Z]{2,4})-", sku_inicial)
                if _m_sub:
                    st.session_state.setdefault("sku_contadores", {})[_m_sub.group(1)] = validacion["nuevo_num"]
            if validacion["conflicto"]:
                conflicto_entry = {
                    "idx":              i,
                    "nombre":           prod.get("nombre", f"Producto {i+1}"),
                    "sku_propuesto":    sku_inicial,
                    "sku_ajustado":     sku_final,
                    "datos_gemini":     datos,
                    "skus_odoo_match":  validacion["skus_odoo_match"],
                    "productos_odoo":   [],
                    "razon":            "sku",
                    "similares_odoo":   [],
                }

            # ── 2. Similitud por imagen, nombre y RAG ────────────────────────
            if prods_odoo:
                similares = buscar_similares_odoo(
                    img["data"], prod.get("nombre", ""),
                    prods_odoo, phashes_odoo,
                )
            else:
                similares = []

            if similares and conflicto_entry is None:
                razones = []
                if any(s.get("por_imagen") for s in similares):
                    razones.append("imagen similar")
                if any(s.get("por_nombre") for s in similares):
                    razones.append("nombre similar")
                conflicto_entry = {
                    "idx":              i,
                    "nombre":           prod.get("nombre", f"Producto {i+1}"),
                    "sku_propuesto":    sku_inicial,
                    "sku_ajustado":     sku_inicial,
                    "datos_gemini":     datos,
                    "skus_odoo_match":  [s["sku"] for s in similares],
                    "productos_odoo":   [],
                    "razon":            ", ".join(razones) if razones else "similitud detectada",
                    "similares_odoo":   similares,
                }
            elif similares and conflicto_entry is not None:
                conflicto_entry["similares_odoo"] = similares

            if conflicto_entry:
                conflictos.append(conflicto_entry)
                skus_ya_en_conflicto.add(i)
        else:
            sku_final = sku_inicial

        prod["sku"]           = sku_final if skus_odoo else sku_inicial
        prod["titulo_imagen"] = datos.get("titulo")       or prod.get("nombre", "")
        prod["descripcion"]   = datos.get("descripcion")  or prod.get("descripcion", "")
        prod["categoria"]     = datos.get("categoria")    or prod.get("categoria", "Varios")
        prod["atributo"]      = datos.get("atributo_desc") or prod.get("atributo", "Estándar")
        prod["atributo_cod"]  = datos.get("atributo_cod")  or "EST"
        procesadas += 1

    # Enriquecer conflictos con detalles completos de ODOO (1 sola llamada en batch)
    if conflictos:
        todos_skus  = list({s for c in conflictos for s in c["skus_odoo_match"]})
        detalles    = cargar_detalle_productos_odoo(todos_skus)
        detalle_map = {p["default_code"]: p for p in detalles}
        for c in conflictos:
            c["productos_odoo"] = [detalle_map[s] for s in c["skus_odoo_match"] if s in detalle_map]

    # Guardar errores de Gemini en session_state para mostrarlos en UI
    if errores_gemini:
        st.session_state["gemini_errores"] = errores_gemini

    return productos, procesadas, conflictos


# ══════════════════════════════════════════════════════════════════════════════
# Clasificación padre / variante — flujo unificado
# ══════════════════════════════════════════════════════════════════════════════

def _generar_sku_padre(subcategoria_cod: str) -> str:
    """Genera SKU de producto padre: SUBCAT-NNNN (sin atributo).
    Solo usa códigos de SUBCATEGORIAS; cualquier otro cae a VAR.
    """
    sub = subcategoria_cod.strip().upper() if subcategoria_cod else _SUBCAT_DEFAULT
    if sub not in SUBCATEGORIAS:
        sub = _SUBCAT_DEFAULT
    contadores = st.session_state.setdefault("sku_contadores", {})
    contadores[sub] = contadores.get(sub, 0) + 1
    sku = f"{sub}-{contadores[sub]:04d}"
    if st.session_state.get("modo_prueba"):
        sku += "_test"
    return sku


def _atributo_cod_desde_valor(atributo_valor: str | None) -> str | None:
    """Mapea atributo_valor al código de SKU del dict ATRIBUTOS. Fallback: 3 letras."""
    if not atributo_valor:
        return None
    val = atributo_valor.strip()
    val_lower = val.lower()
    # Quitar prefijo "talla " si Claude lo incluyó (ej: "Talla M" → "M")
    if val_lower.startswith("talla "):
        val = val[6:].strip()
        val_lower = val.lower()
    # Normalizar alias de tallas grandes
    _alias = {"xxxl": "3XL", "3xl": "3XL", "xxl": "XXL"}
    if val_lower in _alias:
        return _alias[val_lower]
    for cod, desc in ATRIBUTOS.items():
        if desc.lower() == val_lower or cod.lower() == val_lower:
            return cod
    return val[:3].upper()


def analizar_clasificacion_packing(file_bytes: bytes, productos: list[dict], modo_fase1: bool = False) -> list[dict]:
    """
    Análisis unificado padre/variante para todos los productos del packing list.
    Reemplaza la detección de duplicados + la detección de conflictos SKU.

    Flujo:
    1. Extrae imágenes del Excel y las guarda en temp
    2. Claude Vision por producto (subcategoria_cod, atributo_cod, descripcion…)
    3. Extrae nombre_base + atributo en batch (un solo call a Claude Haiku)
    4. Busca padres en Odoo por nombre
    5. Agrupa por nombre_base (detecta variantes y duplicados locales)
    6. Genera propuesta de SKU y acción para cada producto

    Devuelve lista de propuestas (una por producto del packing list).
    """
    imagenes     = extraer_imagenes_excel(file_bytes)
    prods_odoo   = st.session_state.get("odoo_productos", [])
    phashes_odoo = st.session_state.get("odoo_phashes", {})
    skus_odoo    = st.session_state.get("odoo_skus", [])

    if skus_odoo:
        sincronizar_contadores_con_odoo(skus_odoo)
    if imagenes:
        guardar_imagenes_temp(imagenes, productos)

    # ── Pasos 1+2: Vision y Nombres en PARALELO (LangGraph fan-out) ─────────
    if modo_fase1:
        # Solo texto — rápido, sin llamadas Vision. Los títulos/descripciones quedan vacíos.
        datos_vision, clasificaciones = ejecutar_solo_nombres(productos)
    else:
        # Vision completa — lento para listas grandes.
        datos_vision, clasificaciones = ejecutar_vision_y_nombres(productos, imagenes)

    # Liberar bytes de imágenes — ya están en disco y ya se usaron para Vision
    del imagenes

    # ── Paso 3: Construir propuestas ──────────────────────────────────────────
    # Registro local: nombre_base → {sku_padre, subcod, numero, odoo_id, odoo_nombre,
    #                                 padre_idx, variantes: {att_cod: {sku, idx}}}
    registro: dict[str, dict] = {}
    propuestas: list[dict] = []

    # Pre-computar phashes intra-packing
    # _phash_override: imagen idéntica → forzar mismo nombre_base (ya existía)
    # _phash_by_idx:   phash por índice, para detectar productos visualmente distintos
    _UMBRAL_SPLIT = 15   # distancia phash > 15 → productos diferentes aunque mismo nombre_base
    _phashes_pl:    list[tuple[object, str]] = []  # [(phash_obj, nombre_base)]
    _phash_override: dict[int, str]  = {}          # idx → nombre_base forzado por imagen idéntica
    _phash_by_idx:   dict[int, object] = {}        # idx → phash_obj
    for _ii, _pp in enumerate(productos):
        _img_p = _pp.get("imagen_temp_path")
        if not _img_p:
            continue
        from pathlib import Path as _Path
        if not _Path(_img_p).exists():
            continue
        try:
            _ph = _phash_imagen(_Path(_img_p).read_bytes())
            if _ph is None:
                continue
            _phash_by_idx[_ii] = _ph
            _clas_ii = clasificaciones[_ii] if _ii < len(clasificaciones) else {}
            _nom_ii  = str(_pp.get("nombre") or f"Producto {_ii+1}").strip()
            _nb_ii   = str(_clas_ii.get("nombre_base") or _nom_ii).strip()
            for _ph_prev, _nb_prev in _phashes_pl:
                if _ph - _ph_prev <= 3:
                    _phash_override[_ii] = _nb_prev
                    break
            else:
                _phashes_pl.append((_ph, _nb_ii))
        except Exception:
            pass

    ACCION_LABELS = {
        "crear_padre_y_variante": "🆕 Padre nuevo + variante nueva",
        "crear_padre_solo":       "🆕 Padre nuevo (sin variante)",
        "crear_variante":         "➕ Variante nueva en padre existente",
        "reutilizar":             "♻️ Reutilizar producto existente",
        "duplicado":              "🔁 Duplicado — mismo producto ya registrado",
    }

    for i, prod in enumerate(productos):
        datos   = datos_vision[i]
        clas    = clasificaciones[i] if i < len(clasificaciones) else {}
        row_num = prod.get("fila_excel_0idx", i + 1)
        nombre_original = str(prod.get("nombre") or f"Producto {i+1}").strip()
        nombre      = nombre_original
        nombre_base = _phash_override.get(i) or str(clas.get("nombre_base") or nombre).strip()
        # Quitar sufijos #N que el parser añade a nombres duplicados (causarían grupos distintos)
        nombre_base = re.sub(r'\s*#\d+$', '', nombre_base).strip()
        # Quitar color al final del nombre para que variantes del mismo producto se agrupen
        # bajo el mismo padre aunque Haiku no haya quitado el color (ej. "Caja azul" y "Caja rosa")
        _COLORES_TRAIL = (
            r'\s+(negro|negra|blanco|blanca|gris|rojo|roja|azul|verde|amarillo|amarilla|'
            r'rosa|naranja|morado|morada|café|beige|plateado|plateada|dorado|dorada|'
            r'multicolor|lila|turquesa|celeste|crema|nude|camel|marino|oliva)\s*$'
        )
        nombre_base = re.sub(_COLORES_TRAIL, '', nombre_base, flags=re.IGNORECASE).strip()
        # Sincronizar nombre del producto con nombre_base (ya en español, sin color)
        if nombre_base and nombre_base != nombre:
            prod["nombre"] = nombre_base
            nombre = nombre_base
        _color_val  = clas.get("atributo_color") or None
        _talla_val  = clas.get("atributo_talla") or None
        _color_cod  = _atributo_cod_desde_valor(_color_val)
        _talla_cod  = _atributo_cod_desde_valor(_talla_val)
        # Fallback de talla: buscar en el nombre original con regex (gratis, sin API)
        if not _talla_cod:
            import re as _re
            _m_talla = _re.search(r'\b(3XL|XXL|XL|XS|[SML])\b', nombre_original)
            if _m_talla:
                _talla_cod = _m_talla.group(1).upper()
                _talla_val = _talla_cod
        if _color_cod and _talla_cod:
            att_cod   = f"{_color_cod}-{_talla_cod}"
            att_tipo  = "Color-Talla"
            att_valor = f"{_color_val} {_talla_val}"
        elif _talla_cod:
            att_cod   = _talla_cod
            att_tipo  = "Talla"
            att_valor = _talla_val
        elif _color_cod:
            att_cod   = _color_cod
            att_tipo  = "Color"
            att_valor = _color_val
        else:
            att_cod   = datos.get("atributo_cod") or None
            att_tipo  = None
            att_valor = None
        sub_cod     = datos.get("subcategoria_cod") or "VAR"

        prop: dict = {
            "idx":             i,
            "producto":        prod,
            "nombre":          nombre,
            "nombre_base":     nombre_base,
            "atributo_tipo":   att_tipo,
            "atributo_valor":  att_valor,
            "atributo_cod":    att_cod,
            "subcategoria_cod": sub_cod,
            "datos_vision":    datos,
            "imagen_temp_path": prod.get("imagen_temp_path"),
            "padre_fuente":    "nuevo",
            "padre_odoo_id":   None,
            "padre_odoo_nombre": "",
            "padre_sku":       "",
            "sku":             "",
            "accion":          "crear_padre_y_variante",
            "accion_display":  "",
            "confianza_padre": 0.0,
            "requiere_revision": False,
            "nota_revision":   "",
            "duplicado_de_idx": None,
            "_vision_pendiente": bool(datos.get("_vision_pendiente")),
        }

        # ── Verificar que imagen es similar al padre registrado ───────────────
        # Si mismo nombre_base pero imágenes visualmente distintas → padre separado
        if nombre_base in registro:
            _ph_actual = _phash_by_idx.get(i)
            _ph_padre  = registro[nombre_base].get("padre_phash")
            if _ph_actual is not None and _ph_padre is not None:
                if _ph_actual - _ph_padre > _UMBRAL_SPLIT:
                    # Imagen distinta al padre original — buscar si ya existe un split similar
                    _sfx_check = 1
                    _matched_split = None
                    while f"{nombre_base}#{_sfx_check}" in registro:
                        _ph_split = registro[f"{nombre_base}#{_sfx_check}"].get("padre_phash")
                        if _ph_split is not None and _ph_actual - _ph_split <= _UMBRAL_SPLIT:
                            _matched_split = f"{nombre_base}#{_sfx_check}"
                            break
                        _sfx_check += 1
                    nombre_base = _matched_split if _matched_split else f"{nombre_base}#{_sfx_check}"
                    prop["nombre_base"] = nombre_base

        if nombre_base in registro:
            # ── Padre ya en registro local ─────────────────────────────────
            entrada = registro[nombre_base]
            prop.update({
                "padre_fuente":      "local",
                "padre_sku":         entrada["sku_padre"],
                "padre_odoo_id":     entrada["odoo_id"],
                "padre_odoo_nombre": entrada["odoo_nombre"],
                "confianza_padre":   1.0,
            })
            if not att_cod:
                # Sin atributo → duplicado del padre
                prop["accion"]           = "duplicado"
                prop["duplicado_de_idx"] = entrada.get("padre_idx")
                prop["sku"]              = entrada["sku_padre"]
            elif att_cod in entrada["variantes"]:
                # Misma variante ya existe en este packing list
                prop["accion"]           = "duplicado"
                prop["duplicado_de_idx"] = entrada["variantes"][att_cod]["idx"]
                prop["sku"]              = entrada["variantes"][att_cod]["sku"]
            else:
                # Nueva variante del mismo padre — el número es fijo (hereda del padre)
                prop["accion"] = "crear_variante"
                sku_var = _sku_mismo_numero(entrada["subcod"], entrada["numero"], att_cod)
                # NO llamar validar_sku_vs_odoo: el número lo hereda del padre y no
                # debe ajustarse aunque haya SKUs con número mayor en Odoo.
                prop["sku"] = sku_var
                entrada["variantes"][att_cod] = {"sku": sku_var, "idx": i}
        else:
            # ── Buscar en Odoo ─────────────────────────────────────────────
            candidato = None
            if prods_odoo:
                _img_path = prod.get("imagen_temp_path")
                _img_bytes = Path(_img_path).read_bytes() if _img_path and Path(_img_path).exists() else None
                candidato = _buscar_padre_en_odoo_por_nombre(
                    nombre_base, prods_odoo, phashes_odoo, _img_bytes
                )
                del _img_bytes

            if candidato:
                prod_odoo  = candidato["prod"]
                sku_odoo   = prod_odoo.get("default_code", "")
                partes     = sku_odoo.replace("_test", "").split("-")
                try:
                    num_odoo = int(partes[1]) if len(partes) >= 2 else 1
                    sub_odoo = partes[0]
                except (ValueError, IndexError):
                    num_odoo, sub_odoo = 1, sub_cod
                # padre_sku siempre SUBCAT-NNNN (sin atributo), aunque default_code en Odoo lo traiga completo
                sku_padre_base = f"{sub_odoo}-{num_odoo:04d}"
                if st.session_state.get("modo_prueba") and "_test" in sku_odoo:
                    sku_padre_base += "_test"
                prop.update({
                    "padre_fuente":      "odoo",
                    "padre_odoo_id":     prod_odoo.get("id"),
                    "padre_odoo_nombre": prod_odoo.get("name", ""),
                    "padre_sku":         sku_padre_base,
                    "confianza_padre":   candidato["score"],
                })
                if att_cod:
                    sku_var = _sku_mismo_numero(sub_odoo, num_odoo, att_cod)
                    # NO llamar validar_sku_vs_odoo: padre encontrado en Odoo tiene
                    # número fijo (ej. ILUM-0011). La variante hereda ese número
                    # y solo cambia el atributo → ILUM-0011-EST, nunca ILUM-0017-EST.
                    prop["sku"]    = sku_var
                    prop["accion"] = "crear_variante"
                else:
                    prop["sku"]    = sku_odoo
                    prop["accion"] = "reutilizar"
                registro[nombre_base] = {
                    "sku_padre": sku_padre_base, "subcod": sub_odoo, "numero": num_odoo,
                    "odoo_id": prod_odoo.get("id"), "odoo_nombre": prod_odoo.get("name", ""),
                    "padre_idx": i,
                    "padre_phash": _phash_by_idx.get(i),
                    "variantes": {att_cod: {"sku": prop["sku"], "idx": i}} if att_cod else {},
                }
            else:
                # ── Crear padre nuevo ──────────────────────────────────────
                sku_padre = _generar_sku_padre(sub_cod)
                partes    = sku_padre.replace("_test", "").split("-")
                try:
                    # sub_real puede diferir de sub_cod si sub_cod no era válido
                    # (ej. sub_cod="LUZ" → sub_real="VAR")
                    sub_real  = partes[0]
                    num_nuevo = int(partes[1]) if len(partes) >= 2 else 1
                except (ValueError, IndexError):
                    sub_real  = sub_cod if sub_cod in SUBCATEGORIAS else _SUBCAT_DEFAULT
                    num_nuevo = 1
                prop["padre_sku"] = sku_padre
                if att_cod:
                    sku_var = _sku_mismo_numero(sub_real, num_nuevo, att_cod)
                    if skus_odoo:
                        _val = validar_sku_vs_odoo(sku_var, skus_odoo)
                        sku_var = _val["sku_ajustado"]
                        # Si validar_sku_vs_odoo ajustó el número (conflicto con Odoo),
                        # actualizar el contador de sesión para que el siguiente padre
                        # parta de ahí y no vuelva a generar el mismo número.
                        if _val["nuevo_num"] is not None:
                            num_nuevo = _val["nuevo_num"]
                            st.session_state.setdefault("sku_contadores", {})[sub_real] = num_nuevo
                            sku_padre = f"{sub_real}-{num_nuevo:04d}"
                            if st.session_state.get("modo_prueba"):
                                sku_padre += "_test"
                            prop["padre_sku"] = sku_padre
                    prop["sku"]    = sku_var
                    prop["accion"] = "crear_padre_y_variante"
                else:
                    prop["sku"]    = sku_padre
                    prop["accion"] = "crear_padre_solo"
                registro[nombre_base] = {
                    "sku_padre": sku_padre, "subcod": sub_real, "numero": num_nuevo,
                    "odoo_id": None, "odoo_nombre": "",
                    "padre_idx": i,
                    "padre_phash": _phash_by_idx.get(i),
                    "variantes": {att_cod: {"sku": prop["sku"], "idx": i}} if att_cod else {},
                }

        prop["accion_display"] = ACCION_LABELS.get(prop["accion"], prop["accion"])

        # Auto-marcar revisión: duplicados y variantes/reutilizaciones de padre en Odoo
        accion_ = prop["accion"]
        if accion_ in ("duplicado", "reutilizar") or (
            accion_ == "crear_variante" and prop.get("padre_fuente") == "odoo"
        ):
            prop["requiere_revision"] = True

        # Enriquecer el producto con datos de visión para el Excel
        prod.update({
            "sku":             prop["sku"],
            "subcategoria_cod": sub_cod,
            "atributo_cod":    att_cod or datos.get("atributo_cod", "EST"),
            "atributo_desc":   att_valor or datos.get("atributo_desc", ""),
            "atributo":        att_valor or datos.get("atributo_desc", ""),
            "descripcion":     datos.get("descripcion", ""),
            "categoria":       datos.get("categoria", ""),
            "titulo":          datos.get("titulo") or nombre,
        })
        propuestas.append(prop)

    del datos_vision
    return propuestas


def enriquecer_vision_bloque(offset: int, block_size: int = 50) -> None:
    """
    Fase 2: delega Vision al agente y actualiza propuestas en session_state.
    Solo toca campos visuales (titulo, descripcion, categoria, atributo_desc);
    no regenera SKUs para no romper relaciones de Fase 1.
    """
    propuestas = st.session_state.clasificacion_propuestas
    productos  = st.session_state.productos

    datos_bloque = ejecutar_vision_bloque_desde_disco(productos, offset, block_size)

    for global_i, datos in datos_bloque.items():
        if global_i >= len(propuestas):
            continue
        prop = propuestas[global_i]
        prop["datos_vision"].update({
            "titulo":        datos.get("titulo", ""),
            "descripcion":   datos.get("descripcion", ""),
            "categoria":     datos.get("categoria", ""),
            "atributo_desc": datos.get("atributo_desc", ""),
        })
        prod = prop["producto"]
        prod.update({
            "descripcion": datos.get("descripcion") or prod.get("descripcion", ""),
            "categoria":   datos.get("categoria")   or prod.get("categoria", ""),
            "atributo":    datos.get("atributo_desc") or prod.get("atributo", ""),
            "titulo":      datos.get("titulo") or prod.get("nombre", ""),
        })
        prop["_vision_pendiente"] = False


def crear_clasificacion_en_odoo(propuestas: list[dict],
                                 tipo_cambio: float,
                                 costo_contenedor: float) -> tuple[list[str], list[str]]:
    """
    Crea templates padre y variantes en Odoo según las propuestas confirmadas.
    Llena: description_sale (material, uso, precio USD, dimensiones, piezas/caja),
           description/notas internas (CBM, guía, cajas, flete), volume (inventario).
    Respeta modo_prueba: nombre con sufijo _test y categoría PRUEBAS_AGENTE.
    Devuelve (resultados_ok, errores).
    """
    odoo_url  = os.environ.get("ODOO_URL", "")
    odoo_db   = os.environ.get("ODOO_DB", "")
    odoo_user = os.environ.get("ODOO_USER", "")
    odoo_pass = os.environ.get("ODOO_PASSWORD", "")
    if not all([odoo_url, odoo_db, odoo_user, odoo_pass]):
        return [], ["Faltan credenciales ODOO en .env"]
    try:
        common = xmlrpc.client.ServerProxy(f"{odoo_url}/xmlrpc/2/common", allow_none=True, encoding="utf-8")
        uid    = common.authenticate(odoo_db, odoo_user, odoo_pass, {})
        models = xmlrpc.client.ServerProxy(f"{odoo_url}/xmlrpc/2/object", allow_none=True, encoding="utf-8")
    except Exception as e:
        return [], [f"No se pudo conectar a ODOO: {e}"]

    modo_prueba  = st.session_state.get("modo_prueba", False)
    cbm_total    = sum(_cbm_total_fila(p["producto"])
                       for p in propuestas if p["accion"] != "duplicado")
    costo_por_m3 = costo_contenedor / cbm_total if cbm_total > 0 else 0.0

    # Categoría
    cat_cache: dict[str, int] = {}
    def _get_cat(nombre_cat: str) -> int | None:
        if nombre_cat in cat_cache:
            return cat_cache[nombre_cat]
        try:
            ids = models.execute_kw(odoo_db, uid, odoo_pass, "product.category", "search",
                                    [[["name", "=", nombre_cat]]])
            cid = ids[0] if ids else models.execute_kw(
                odoo_db, uid, odoo_pass, "product.category", "create", [{"name": nombre_cat}])
            cat_cache[nombre_cat] = cid
            return cid
        except Exception:
            return None

    templates_creados: dict[str, int] = {}  # sku_padre → tmpl_id
    attrs_cache:       dict[str, int] = {}
    attr_vals_cache:   dict[tuple, int] = {}
    tags_cache:        dict[str, int] = {}

    def _get_or_create_tag(nombre: str) -> int | None:
        if nombre in tags_cache:
            return tags_cache[nombre]
        try:
            ids = models.execute_kw(odoo_db, uid, odoo_pass, "product.tag", "search",
                                    [[["name", "=", nombre]]])
            tid = ids[0] if ids else models.execute_kw(
                odoo_db, uid, odoo_pass, "product.tag", "create", [{"name": nombre}])
            tags_cache[nombre] = tid
            return tid
        except Exception:
            return None

    def _get_or_create_attr(nombre: str) -> int:
        if nombre in attrs_cache:
            return attrs_cache[nombre]
        ids = models.execute_kw(odoo_db, uid, odoo_pass, "product.attribute", "search",
                                [[["name", "=", nombre]]])
        aid = ids[0] if ids else models.execute_kw(
            odoo_db, uid, odoo_pass, "product.attribute", "create", [{"name": nombre}])
        attrs_cache[nombre] = aid
        return aid

    def _get_or_create_val(attr_id: int, valor: str) -> int:
        key = (attr_id, valor)
        if key in attr_vals_cache:
            return attr_vals_cache[key]
        ids = models.execute_kw(odoo_db, uid, odoo_pass, "product.attribute.value", "search",
                                [[["attribute_id", "=", attr_id], ["name", "=", valor]]])
        vid = ids[0] if ids else models.execute_kw(
            odoo_db, uid, odoo_pass, "product.attribute.value", "create",
            [{"name": valor, "attribute_id": attr_id}])
        attr_vals_cache[key] = vid
        return vid

    def _build_vals(prod: dict, nombre_tmpl: str, precio_usd: float,
                    costo_unitario: float,
                    sku_direct: str | None = None) -> dict:
        """Construye el dict de valores para product.template con todos los campos."""
        cbm_pz       = _cbm_por_pieza(prod)
        costo_cbm_pz = round(cbm_pz * costo_por_m3, 2)
        largo  = prod.get("largo_cm")
        ancho  = prod.get("ancho_cm")
        alto   = prod.get("alto_cm")
        dims_str = (" × ".join(str(v) for v in [largo, ancho, alto] if v)
                    if any([largo, ancho, alto]) else None)

        # ── Descripción de venta (pestaña Ventas / visible al cliente) ─────────
        desc_partes = []
        if prod.get("descripcion"):
            desc_partes.append(prod["descripcion"])
        if prod.get("material"):
            desc_partes.append(f"Material: {prod['material']}")
        if prod.get("uso"):
            desc_partes.append(f"Uso: {prod['uso']}")
        if precio_usd > 0:
            desc_partes.append(f"Precio USD: ${precio_usd:.2f}")
        pzs = prod.get("piezas_x_caja")
        if pzs:
            try:
                desc_partes.append(f"Piezas por caja: {int(float(pzs))}")
            except (ValueError, TypeError):
                desc_partes.append(f"Piezas por caja: {pzs}")
        if dims_str:
            desc_partes.append(f"Dimensiones: {dims_str} cm")

        # ── Notas internas (pestaña Notas / internal notes) ────────────────────
        notas = []
        if prod.get("nombre_alt"):
            notas.append(f"Nombre alternativo: {prod['nombre_alt']}")
        if prod.get("id_guia"):
            notas.append(f"ID Guía / Referencia: {prod['id_guia']}")
        if prod.get("cajas_master"):
            try:
                notas.append(f"Cajas master: {int(float(prod['cajas_master']))}")
            except (ValueError, TypeError):
                notas.append(f"Cajas master: {prod['cajas_master']}")
        if prod.get("piezas_total"):
            try:
                notas.append(f"Piezas totales en contenedor: {int(float(prod['piezas_total']))}")
            except (ValueError, TypeError):
                notas.append(f"Piezas totales en contenedor: {prod['piezas_total']}")
        if cbm_pz > 0:
            notas.append(f"CBM por pieza: {cbm_pz:.6f} m³")
        if prod.get("cbm_master_carton"):
            notas.append(f"CBM master carton: {prod['cbm_master_carton']}")
        if prod.get("cbm_total_sku"):
            notas.append(f"CBM total SKU: {prod['cbm_total_sku']}")
        if prod.get("cbm_inner_carton"):
            notas.append(f"CBM inner carton: {prod['cbm_inner_carton']}")
        if dims_str:
            notas.append(f"Dimensiones caja: {dims_str} cm")
        if costo_cbm_pz > 0:
            notas.append(f"Costo flete CBM/pieza: ${costo_cbm_pz:.4f} MXN")

        vals: dict = {
            "name":             nombre_tmpl,
            "type":             "product",
            "sale_ok":          True,
            "purchase_ok":      True,
            "list_price":       costo_unitario,
        }
        if desc_partes:
            vals["description_sale"] = "\n".join(desc_partes)
        if notas:
            vals["description"] = "\n".join(notas)

        # ── Campos core (siempre disponibles) ────────────────────────────────
        vol_pz = _safe_float(prod.get("volumen_por_pieza")) or cbm_pz
        if vol_pz > 0:
            vals["volume"] = vol_pz
        peso = _safe_float(prod.get("peso_kg"))
        if peso > 0:
            vals["weight"] = peso

        if sku_direct:
            vals["default_code"] = sku_direct
        cat_nombre = "PRUEBAS_AGENTE" if modo_prueba else "Productos Agente"
        cid = _get_cat(cat_nombre)
        if cid:
            vals["categ_id"] = cid
        return vals

    def _write_campos_custom(tmpl_id: int, prod: dict, precio_usd: float, cbm_pz: float) -> None:
        """Escribe campos de módulos adicionales — si no existen en este Odoo, los ignora."""
        custom: dict = {}
        vol_pz = _safe_float(prod.get("volumen_por_pieza")) or cbm_pz
        if vol_pz > 0:
            custom["cbm_per_product"] = vol_pz
        if precio_usd > 0:
            custom["costo_usd"] = precio_usd
        cbm_mc = _safe_float(prod.get("cbm_master_carton"))
        if cbm_mc > 0:
            custom["cbm_master_box"] = cbm_mc
        pzs_caja = prod.get("piezas_x_caja")
        if pzs_caja:
            try:
                custom["units_per_master_box"] = int(float(pzs_caja))
            except (ValueError, TypeError):
                pass
        largo, ancho, alto = prod.get("largo_cm"), prod.get("ancho_cm"), prod.get("alto_cm")
        for campo, val in [("length", largo), ("width", ancho), ("height", alto)]:
            if val:
                try:
                    custom[campo] = float(val)
                except (ValueError, TypeError):
                    pass
        _contenedor_num = st.session_state.get("contenedor_val", "")
        if _contenedor_num:
            custom["container_numbers"] = _contenedor_num
        if custom:
            try:
                models.execute_kw(odoo_db, uid, odoo_pass,
                                  "product.template", "write", [[tmpl_id], custom])
            except Exception:
                pass  # Campos custom no instalados en este Odoo

    def _apply_rev_tag_and_note(tmpl_id: int, nota: str) -> None:
        existing = models.execute_kw(
            odoo_db, uid, odoo_pass, "product.template", "read",
            [[tmpl_id]], {"fields": ["description"]})
        prev = (existing[0].get("description") or "") if existing else ""
        prefijo = f"⚠️ REQUIERE REVISIÓN: {nota}" if nota else "⚠️ REQUIERE REVISIÓN"
        nueva_desc = f"{prefijo}\n{prev}".strip()
        models.execute_kw(odoo_db, uid, odoo_pass,
                          "product.template", "write", [[tmpl_id], {"description": nueva_desc}])

    def _img_b64(sku: str) -> str | None:
        """Lee la imagen del SKU desde la carpeta temporal y la devuelve en base64."""
        for ext in ("jpg", "jpeg", "png", "webp"):
            p = IMAGENES_TEMP_PATH / f"{sku}.{ext}"
            if p.exists():
                try:
                    return base64.b64encode(p.read_bytes()).decode("utf-8")
                except Exception:
                    return None
        return None

    resultados: list[str] = []
    errores:    list[str] = []

    # Piezas totales extra que aportan los duplicados a su padre
    piezas_extra_por_padre: dict[str, int] = {}
    for _p in propuestas:
        if "duplicado" in str(_p.get("accion", "")).lower():
            _padre = _p.get("padre_sku") or ""
            try:
                piezas_extra_por_padre[_padre] = (
                    piezas_extra_por_padre.get(_padre, 0)
                    + int(float(_p["producto"].get("piezas_total") or 0))
                )
            except (ValueError, TypeError):
                pass

    for prop in propuestas:
        if prop["accion"] == "duplicado":
            continue
        prod           = prop["producto"]
        accion         = prop["accion"]
        requiere_rev   = bool(prop.get("requiere_revision", False))
        nota_rev       = str(prop.get("nota_revision", "") or "")
        sku            = prop["sku"]
        sku_padre      = prop["padre_sku"]
        att_tipo       = prop["atributo_tipo"] or "Variante"
        att_valor      = prop["atributo_valor"] or "Estándar"
        nombre_base    = prop["nombre_base"]
        nombre         = prop["nombre"]
        precio_usd     = _safe_float(prod.get("precio_usd"))
        precio_mxn     = round(precio_usd * tipo_cambio, 2)
        cbm_pz         = _cbm_por_pieza(prod)
        costo_unitario = round(precio_mxn + cbm_pz * costo_por_m3, 2)

        # Nombre del template: sufijo _test en modo prueba
        nombre_tmpl = nombre_base + ("_test" if modo_prueba else "")

        try:
            if accion in ("crear_padre_y_variante", "crear_padre_solo"):
                # Para padre+variante: poner sku_padre en default_code del template
                # Para padre solo: poner el sku directamente
                sku_direct = sku_padre if accion == "crear_padre_y_variante" else sku
                # Sumar piezas de duplicados al padre
                _extra_pzs = piezas_extra_por_padre.get(sku_padre, 0)
                if _extra_pzs > 0:
                    try:
                        _base_pzs = int(float(prod.get("piezas_total") or 0))
                    except (ValueError, TypeError):
                        _base_pzs = 0
                    prod = {**prod, "piezas_total": _base_pzs + _extra_pzs}
                vals = _build_vals(prod, nombre_tmpl, precio_usd,
                                   costo_unitario, sku_direct=sku_direct)
                img = _img_b64(sku)
                if img:
                    vals["image_1920"] = img
                tmpl_id = models.execute_kw(
                    odoo_db, uid, odoo_pass, "product.template", "create", [vals])
                templates_creados[sku_padre] = tmpl_id

                _write_campos_custom(tmpl_id, prod, precio_usd, cbm_pz)
                if requiere_rev:
                    _apply_rev_tag_and_note(tmpl_id, nota_rev)

                if accion == "crear_padre_solo":
                    _rev_sfx = " ⚠️ rev" if requiere_rev else ""
                    resultados.append(f"✅ Padre creado: `{sku}` — {nombre_tmpl}{_rev_sfx}")
                else:
                    attr_id = _get_or_create_attr(att_tipo)
                    val_id  = _get_or_create_val(attr_id, att_valor)
                    models.execute_kw(odoo_db, uid, odoo_pass,
                                      "product.template.attribute.line", "create",
                                      [{"product_tmpl_id": tmpl_id, "attribute_id": attr_id,
                                        "value_ids": [(4, val_id)]}])
                    # Re-write volume/weight tras crear variante (Odoo puede resetearlos)
                    _rewrite: dict = {}
                    _vol_pz = _safe_float(prod.get("volumen_por_pieza")) or cbm_pz
                    if _vol_pz > 0:
                        _rewrite["volume"] = _vol_pz
                    _peso = _safe_float(prod.get("peso_kg"))
                    if _peso > 0:
                        _rewrite["weight"] = _peso
                    if _rewrite:
                        models.execute_kw(odoo_db, uid, odoo_pass,
                                          "product.template", "write", [[tmpl_id], _rewrite])
                    var_ids = models.execute_kw(odoo_db, uid, odoo_pass,
                                                "product.product", "search",
                                                [[["product_tmpl_id", "=", tmpl_id]]])
                    if var_ids:
                        models.execute_kw(odoo_db, uid, odoo_pass, "product.product", "write",
                                          [var_ids, {"default_code": sku}])
                    _rev_sfx = " ⚠️ rev" if requiere_rev else ""
                    resultados.append(f"✅ Padre + variante: `{sku_padre}` / `{sku}` — {nombre_tmpl}{_rev_sfx}")

            elif accion == "crear_variante":
                tmpl_id = (templates_creados.get(sku_padre)
                           or prop.get("padre_odoo_id"))
                if not tmpl_id:
                    errores.append(f"⚠️ No se encontró template padre para `{sku}` ({nombre})")
                    continue
                attr_id = _get_or_create_attr(att_tipo)
                val_id  = _get_or_create_val(attr_id, att_valor)
                lineas  = models.execute_kw(
                    odoo_db, uid, odoo_pass, "product.template.attribute.line", "search_read",
                    [[["product_tmpl_id", "=", tmpl_id], ["attribute_id", "=", attr_id]]],
                    {"fields": ["id"]})
                if lineas:
                    models.execute_kw(odoo_db, uid, odoo_pass,
                                      "product.template.attribute.line", "write",
                                      [[lineas[0]["id"]], {"value_ids": [(4, val_id)]}])
                else:
                    models.execute_kw(odoo_db, uid, odoo_pass,
                                      "product.template.attribute.line", "create",
                                      [{"product_tmpl_id": tmpl_id, "attribute_id": attr_id,
                                        "value_ids": [(4, val_id)]}])
                var_ids = models.execute_kw(
                    odoo_db, uid, odoo_pass, "product.product", "search",
                    [[["product_tmpl_id", "=", tmpl_id],
                      ["product_template_attribute_value_ids"
                       ".product_attribute_value_id", "=", val_id]]])
                if var_ids:
                    models.execute_kw(odoo_db, uid, odoo_pass, "product.product", "write",
                                      [var_ids, {"default_code": sku}])
                _write_campos_custom(tmpl_id, prod, precio_usd, cbm_pz)
                img = _img_b64(sku)
                if img:
                    models.execute_kw(odoo_db, uid, odoo_pass,
                                      "product.template", "write", [[tmpl_id], {"image_1920": img}])
                if requiere_rev:
                    _apply_rev_tag_and_note(tmpl_id, nota_rev)
                _rev_sfx = " ⚠️ rev" if requiere_rev else ""
                resultados.append(f"✅ Variante creada: `{sku}` — {nombre}{_rev_sfx}")

            elif accion == "reutilizar":
                tmpl_id = prop.get("padre_odoo_id")
                if tmpl_id:
                    _write_campos_custom(tmpl_id, prod, precio_usd, cbm_pz)
                resultados.append(f"♻️ Reutilizado: `{sku}` — {nombre}")

        except Exception as e:
            errores.append(f"❌ Error en `{sku}` ({nombre}): {e}")

    return resultados, errores


# ── Herramienta LangChain ──────────────────────────────────────────────────────
@tool
def generar_excel_tool(productos: List[dict] | None = None) -> str:
    """Genera o regenera el archivo Excel de FERRAFORME.
    Llama esta herramienta cuando el usuario confirme proceder o pida generar el Excel.
    No necesitas pasar productos — se usan los que ya están cargados en sesión."""
    tipo_cambio      = st.session_state.get("tipo_cambio", 19.0)
    contenedor       = st.session_state.get("contenedor_val", "")
    costo_contenedor = st.session_state.get("costo_contenedor", 525000.0)
    nombre_packing   = st.session_state.get("filename", "packing_list.xlsx")

    prods = productos if productos else st.session_state.get("productos", [])
    if not prods:
        return "Error: no hay productos cargados en sesión."

    # Recargar odoo_skus desde Odoo antes de subir para evitar duplicados
    # si productos fueron borrados/archivados externamente entre sesiones
    if st.session_state.get("odoo_conectado"):
        _skus_fresh, _err_fresh = cargar_skus_odoo(
            os.environ.get("ODOO_URL", ""), os.environ.get("ODOO_DB", ""),
            os.environ.get("ODOO_USER", ""), os.environ.get("ODOO_PASSWORD", ""),
        )
        if not _err_fresh and _skus_fresh:
            st.session_state.odoo_skus = _skus_fresh
            sincronizar_contadores_con_odoo(_skus_fresh)

    # costo_por_m3 se calcula una vez y lo usan tanto el master Excel como ODOO
    cbm_total    = sum(_cbm_total_fila(p) for p in prods)
    costo_por_m3 = costo_contenedor / cbm_total if cbm_total > 0 else 0.0

    # Generar Excels y subir a ODOO en paralelo
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
        fut_excel  = executor.submit(generar_excel,        prods, tipo_cambio, contenedor)
        fut_master = executor.submit(generar_excel_master, prods, tipo_cambio, costo_contenedor, nombre_packing)
        fut_odoo   = executor.submit(_subir_productos_a_odoo, prods, tipo_cambio, costo_por_m3)

        excel_bytes        = fut_excel.result()
        master_bytes       = fut_master.result()
        subidas, errs_odoo = fut_odoo.result()

    # Reintentar SKUs que fallaron antes de borrar imágenes
    if errs_odoo:
        _propuestas_rt = st.session_state.pop("propuestas_para_reintento", None) or []
        _sub2, _err2 = _reintentar_skus_fallidos(
            errs_odoo, prods, tipo_cambio, costo_por_m3, _propuestas_rt,
            modo_prueba=st.session_state.get("modo_prueba", False),
        )
        subidas   = subidas + _sub2
        errs_odoo = [e for e in errs_odoo if not any(s in e for s in _sub2)] + _err2
    else:
        st.session_state.pop("propuestas_para_reintento", None)

    # Limpiar imágenes temporales
    if IMAGENES_TEMP_PATH.exists():
        for _f in IMAGENES_TEMP_PATH.iterdir():
            try:
                _f.unlink()
            except Exception:
                pass

    st.session_state.excel_bytes  = excel_bytes
    st.session_state.master_bytes = master_bytes
    st.session_state.productos    = prods

    # Agregar SKUs recién subidos a odoo_skus para evitar duplicados en la misma sesión
    _nuevos_skus_tool = [p.get("sku", "") for p in prods if p.get("sku")]
    if _nuevos_skus_tool:
        _skus_set_tool = set(st.session_state.get("odoo_skus", []))
        _skus_set_tool.update(_nuevos_skus_tool)
        st.session_state.odoo_skus = sorted(_skus_set_tool)
        sincronizar_contadores_con_odoo(st.session_state.odoo_skus)

    _reportar_subida_imagenes(subidas, errs_odoo)

    resumen_odoo = (
        f" {len(subidas)} productos subidos a ODOO." if subidas
        else (f" Advertencia ODOO: {errs_odoo[0]}" if errs_odoo else "")
    )
    return f"Excel FERRAFORME y Excel master generados con {len(prods)} productos.{resumen_odoo}"


def llamar_agente(user_input: str, system: str, lc_messages: list) -> str:
    """Llama al agente con LangChain bind_tools. Devuelve el texto de respuesta."""
    llm = ChatAnthropic(model="claude-sonnet-4-6", max_tokens=8000)
    llm_con_tools = llm.bind_tools([generar_excel_tool])

    mensajes = [SystemMessage(content=system)] + lc_messages + [HumanMessage(content=user_input)]

    while True:
        respuesta = llm_con_tools.invoke(mensajes)
        mensajes.append(respuesta)

        if not respuesta.tool_calls:
            break

        for tc in respuesta.tool_calls:
            if tc["name"] == "generar_excel_tool":
                resultado = generar_excel_tool.invoke(tc["args"])
            else:
                resultado = "Herramienta no reconocida."
            mensajes.append(ToolMessage(content=resultado, tool_call_id=tc["id"]))

    # Normalizar: Claude puede devolver lista de bloques en lugar de string
    content = respuesta.content
    if isinstance(content, list):
        partes = []
        for bloque in content:
            if isinstance(bloque, dict) and bloque.get("type") == "text":
                partes.append(bloque.get("text", ""))
            elif isinstance(bloque, str):
                partes.append(bloque)
        return "\n".join(partes)
    return content


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAR DESDE ODOO — helpers y orquestador
# ══════════════════════════════════════════════════════════════════════════════

def _parsear_desc_sale(texto: str) -> dict:
    """Parsea description_sale de Odoo a campos individuales del producto."""
    r: dict = {}
    if not texto:
        return r
    for line in texto.splitlines():
        line = line.strip()
        if line.startswith("Material:"):
            r["material"] = line[9:].strip()
        elif line.startswith("Uso:"):
            r["uso"] = line[4:].strip()
        elif line.startswith("Precio USD:"):
            try:
                r["precio_usd"] = float(line[11:].strip().lstrip("$"))
            except Exception:
                pass
        elif line.startswith("Piezas por caja:"):
            try:
                r["piezas_x_caja"] = int(float(line[16:].strip()))
            except Exception:
                pass
        elif line.startswith("Dimensiones:"):
            dims = re.sub(r'\s*cm\s*$', '', line[12:].strip())
            parts = re.split(r'[×xX]', dims)
            if len(parts) >= 3:
                try:
                    r["largo_cm"] = float(parts[0].strip())
                    r["ancho_cm"] = float(parts[1].strip())
                    r["alto_cm"]  = float(parts[2].strip())
                except Exception:
                    pass
    return r


def _parsear_desc_interna(texto: str) -> dict:
    """Parsea la descripción interna (notas) de Odoo a campos de producto."""
    r: dict = {}
    if not texto:
        return r
    for line in texto.splitlines():
        line = line.strip()
        if line.startswith("Nombre alternativo:"):
            r["nombre_alt"] = line[19:].strip()
        elif line.lower().startswith("id guía"):
            r["id_guia"] = line.split(":", 1)[-1].strip()
        elif line.startswith("Cajas master:"):
            try:
                r["cajas_master"] = int(float(line[13:].strip()))
            except Exception:
                pass
        elif line.startswith("Piezas totales"):
            try:
                r["piezas_total"] = int(float(line.split(":", 1)[-1].strip()))
            except Exception:
                pass
        elif line.startswith("CBM por pieza:"):
            try:
                r["cbm_por_pieza"] = float(line[14:].strip().split()[0])
            except Exception:
                pass
        elif line.startswith("CBM master carton:"):
            try:
                r["cbm_master_carton"] = float(line[18:].strip())
            except Exception:
                pass
        elif line.startswith("CBM total SKU:"):
            try:
                r["cbm_total_sku"] = float(line[14:].strip())
            except Exception:
                pass
    return r


def _cod_atrib_a_tipo(cod: str) -> str:
    _COLORES    = {"NEG","BLN","GRI","ROJ","AZL","VER","AMA","ROS","NAR","MOR","CAF","BEI","MUL","PLA","DOR"}
    _TALLAS     = {"XS","S","M","L","XL","XXL","3XL","UNI"}
    _MATERIALES = {"MAD","MET","TEL","CUE"}
    if cod in _COLORES:
        return "Color"
    if cod in _TALLAS:
        return "Talla"
    if cod in _MATERIALES:
        return "Material"
    return "Tipo"


def generar_excels_desde_contenedor(contenedor: str) -> dict:
    """
    Fetches products from Odoo by container number, reconstructs product data,
    and generates all four Excel files.

    Returns dict with keys: reporte_bytes, ferraforme_bytes, master_bytes,
    purchase_bytes, n_prods, error.
    """
    odoo_url  = os.environ.get("ODOO_URL", "")
    odoo_db   = os.environ.get("ODOO_DB", "")
    odoo_user = os.environ.get("ODOO_USER", "")
    odoo_pass = os.environ.get("ODOO_PASSWORD", "")
    if not all([odoo_url, odoo_db, odoo_user, odoo_pass]):
        return {"error": "Faltan credenciales ODOO en el .env"}

    tmpls, err = cargar_productos_por_contenedor(odoo_url, odoo_db, odoo_user, odoo_pass, contenedor)
    if err:
        return {"error": err}

    tc = st.session_state.get("tipo_cambio", 19.0)
    cc = st.session_state.get("costo_contenedor", 525000.0)

    productos:  list[dict] = []
    propuestas: list[dict] = []

    for tmpl in tmpls:
        padre_sku  = (tmpl.get("default_code") or "").replace("_test", "")
        nombre_raw = tmpl.get("name") or ""
        nombre     = nombre_raw[:-5] if nombre_raw.endswith("_test") else nombre_raw

        ds = _parsear_desc_sale(tmpl.get("description_sale") or "")
        dn = _parsear_desc_interna(tmpl.get("description") or "")

        cbm_pz = dn.get("cbm_por_pieza") or _safe_float(tmpl.get("volume"))

        m_padre = re.match(r'^([A-Z]{2,4})-(\d{4})-([A-Z]{2,5})', padre_sku)
        subcat_padre = m_padre.group(1) if m_padre else "VAR"

        prod_base: dict = {
            "nombre":            nombre,
            "precio_usd":        ds.get("precio_usd", 0),
            "material":          ds.get("material", ""),
            "uso":               ds.get("uso", ""),
            "piezas_x_caja":     ds.get("piezas_x_caja"),
            "largo_cm":          ds.get("largo_cm"),
            "ancho_cm":          ds.get("ancho_cm"),
            "alto_cm":           ds.get("alto_cm"),
            "piezas_total":      dn.get("piezas_total"),
            "cajas_master":      dn.get("cajas_master"),
            "cbm_por_pieza":     cbm_pz or None,
            "cbm_master_carton": dn.get("cbm_master_carton"),
            "cbm_total_sku":     dn.get("cbm_total_sku"),
            "tipo_producto":     "Producto almacenable",
            "descripcion":       tmpl.get("description_sale") or "",
        }

        # Variants with a different SKU from the template = real color/size variants
        variants_con_sku = [
            v for v in (tmpl.get("variants") or [])
            if v.get("default_code")
            and v["default_code"].replace("_test", "") != padre_sku
        ]

        if not variants_con_sku:
            m = re.match(r'^([A-Z]{2,4})-(\d{4})-([A-Z]{2,5})', padre_sku)
            atrib_cod = m.group(3) if m else "EST"
            fila = len(propuestas)
            prod = {
                **prod_base,
                "sku":             padre_sku,
                "padre_sku":       padre_sku,
                "categoria":       SUBCATEGORIAS.get(subcat_padre, "Varios"),
                "atributo":        ATRIBUTOS.get(atrib_cod, "Estándar"),
                "fila_excel_0idx": fila,
            }
            productos.append(prod)
            propuestas.append({
                "nombre":            nombre,
                "nombre_base":       nombre,
                "sku":               padre_sku,
                "padre_sku":         padre_sku,
                "atributo_tipo":     _cod_atrib_a_tipo(atrib_cod),
                "atributo_valor":    ATRIBUTOS.get(atrib_cod, "Estándar"),
                "accion":            "nuevo_padre",
                "accion_display":    "Padre nuevo",
                "padre_odoo_nombre": nombre,
                "padre_fuente":      "odoo",
                "requiere_revision": False,
                "nota_revision":     "",
                "producto":          prod,
            })
        else:
            for var in variants_con_sku:
                var_sku = (var.get("default_code") or "").replace("_test", "")
                m = re.match(r'^([A-Z]{2,4})-(\d{4})-([A-Z]{2,5})', var_sku)
                atrib_cod = m.group(3) if m else "EST"
                fila = len(propuestas)
                prod = {
                    **prod_base,
                    "sku":             var_sku,
                    "padre_sku":       padre_sku,
                    "categoria":       SUBCATEGORIAS.get(subcat_padre, "Varios"),
                    "atributo":        ATRIBUTOS.get(atrib_cod, "Estándar"),
                    "fila_excel_0idx": fila,
                }
                productos.append(prod)
                propuestas.append({
                    "nombre":            nombre,
                    "nombre_base":       nombre,
                    "sku":               var_sku,
                    "padre_sku":         padre_sku,
                    "atributo_tipo":     _cod_atrib_a_tipo(atrib_cod),
                    "atributo_valor":    ATRIBUTOS.get(atrib_cod, "Estándar"),
                    "accion":            "nueva_variante",
                    "accion_display":    "Variante nueva",
                    "padre_odoo_nombre": nombre,
                    "padre_fuente":      "odoo",
                    "requiere_revision": False,
                    "nota_revision":     "",
                    "producto":          prod,
                })

    if not productos:
        return {"error": "No se pudieron reconstruir productos desde Odoo"}

    nombre_packing = f"{contenedor}_odoo"
    return {
        "reporte_bytes":    generar_reporte_clasificacion(propuestas, nombre_packing),
        "ferraforme_bytes": generar_excel(productos, tc, contenedor),
        "master_bytes":     generar_excel_master(productos, tc, cc, nombre_packing),
        "purchase_bytes":   generar_excel_purchase(productos, tc, cc),
        "n_prods":          len(productos),
        "error":            None,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="FERRAFORME — Agente v2", page_icon="📦", layout="wide")

defaults = {
    "chat":              [],
    "lc_messages":       [],   # historial LangChain (HumanMessage / AIMessage)
    "chat_dudas":        [],   # mensajes de aclaración durante la fase de dudas
    "chat_fase":         [],   # mensajes de aclaración durante duplicados/conflictos
    "analisis":          None,
    "productos":         [],
    "excel_bytes":       None,
    "master_bytes":      None,
    "purchase_bytes":    None,
    "archivo_id":        None,
    "dudas_relevantes":  [],
    "dudas_menores":     [],
    "respuestas_dudas":  {},
    "esperando_dudas":   False,
    "_quick_reply":      None,   # respuesta rápida desde botón
    "sku_contadores":          {},
    "odoo_skus":               [],
    "odoo_productos":          [],   # todos los productos ODOO con name (sin image_128)
    "odoo_phashes":            {},   # {sku: phash}
    "odoo_imagen_urls":        {},   # {sku: url_publica_supabase}
    "odoo_conectado":          False,
    "esperando_duplicados":    False,
    "dup_paso":                1,
    "dup_respuestas":          {},
    # ── Clasificación padre/variante ──────────────────────────────────────────
    "clasificacion_activa":    False,
    "clasificacion_propuestas": [],
    "clasificacion_reporte_bytes": None,
    "vision_bloque_offset":    0,
    "duplicados_pendientes":   [],
    "dup_grupos_segunda":      [],      # grupos encontrados en segunda vuelta de detección
    "imagenes_excel":          {},      # {row_0idx: {data, ext}} extraídas antes de duplicados
    "historial_duplicados":    [],      # grupos resueltos para consulta posterior
    "dup_snapshot":            None,   # snapshot visual del último análisis de duplicados
    "esperando_conflictos":    False,
    "conflictos_pendientes":   [],
    "resoluciones_conflictos": {},
    # ── Agregar productos en lote ──────────────────────────────────────────────
    "agregar_lote_activo": False,
    "agregar_lote_paso":   1,       # 1=subir imgs, 2=procesar, 3=revisar/confirmar
    "agregar_lote_n":      1,
    "agregar_lote_imgs":   {},      # {idx: {"data": bytes, "ext": str, "nombre": str}}
    "agregar_lote_prods":  [],      # productos procesados (con SKU)
    "agregar_lote_conf":   [],      # conflictos ODOO para estos productos
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Auto-cargar cache ODOO al inicio ───────────────────────────────────────────
if not st.session_state.odoo_conectado and not st.session_state.odoo_skus:
    _cache_auto = cargar_cache_odoo()
    if _cache_auto:
        st.session_state.odoo_skus      = _cache_auto["skus"]
        st.session_state.odoo_productos = _cache_auto["productos"]
        st.session_state.odoo_phashes   = _cache_auto["phashes"]
        st.session_state.odoo_conectado = True
        sincronizar_contadores_con_odoo(_cache_auto["skus"])

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Configuración")
    tipo_cambio = st.number_input("Tipo de cambio USD → MXN", value=19.00, step=0.50, format="%.2f", key="tipo_cambio")
    costo_contenedor = st.number_input("Costo contenedor (MXN)", value=525000.0, step=1000.0, format="%.0f", key="costo_contenedor")

    if "contenedor_val" not in st.session_state:
        st.session_state.contenedor_val = ""
    contenedor_input = st.text_input("Número de contenedor", value=st.session_state.contenedor_val)
    st.session_state.contenedor_val = contenedor_input
    contenedor = contenedor_input

    st.divider()
    modo_prueba = st.toggle(
        "🧪 Modo prueba",
        value=st.session_state.get("modo_prueba", False),
        help="Los SKUs generados llevarán el sufijo _test para no mezclar con productos reales",
    )
    st.session_state.modo_prueba = modo_prueba
    if modo_prueba:
        st.caption("⚠️ SKUs con sufijo `_test` — solo para pruebas")

    st.divider()
    st.subheader("🔗 ODOO")

    # Mostrar info del cache si existe
    info_cache = info_cache_odoo()
    if info_cache:
        st.caption(f"💾 Cache: {info_cache}")

    cargar_btn = st.button("🔄 Cargar / Actualizar SKUs", width="stretch",
                           help="Carga desde caché si está disponible; si no, descarga desde ODOO y sincroniza")

    def _aplicar_datos_odoo(skus, productos, phashes, imagen_urls=None, desde_cache=False):
        st.session_state.odoo_skus       = skus
        st.session_state.odoo_productos  = productos
        st.session_state.odoo_phashes    = phashes
        st.session_state.odoo_imagen_urls = imagen_urls or {}
        st.session_state.odoo_conectado  = True
        sin_img = len(skus) - len(phashes)
        origen  = "Supabase" if desde_cache else "ODOO"
        st.success(f"✅ {len(skus)} SKUs · {len(phashes)} imágenes · {sin_img} sin imagen  —  {origen}")

    def _cargar_desde_odoo(forzar=False):
        odoo_url  = os.environ.get("ODOO_URL", "")
        odoo_db   = os.environ.get("ODOO_DB", "")
        odoo_user = os.environ.get("ODOO_USER", "")
        odoo_pass = os.environ.get("ODOO_PASSWORD", "")
        if not all([odoo_url, odoo_db, odoo_user, odoo_pass]):
            st.warning("Faltan credenciales ODOO en el .env")
            return

        # 1. Intentar Supabase primero (salvo recarga forzada)
        if not forzar:
            # 1a. pkl local primero (instantáneo)
            cache = cargar_cache_odoo()
            if cache:
                _aplicar_datos_odoo(
                    cache["skus"], cache["productos"], cache["phashes"],
                    desde_cache=True,
                )
                return
            # 1b. Supabase como fallback (cuando no hay pkl — ej. Digital Ocean)
            with st.spinner("Cargando desde Supabase..."):
                sb_data = _cargar_desde_supabase()
            if sb_data:
                _aplicar_datos_odoo(
                    sb_data["skus"], sb_data["productos"], sb_data["phashes"],
                    imagen_urls=sb_data["imagen_urls"], desde_cache=True,
                )
                return

        # 2. Cargar desde ODOO
        with st.spinner("Conectando a ODOO..."):
            skus, error = cargar_skus_odoo(odoo_url, odoo_db, odoo_user, odoo_pass)
        if error:
            st.error(f"Error: {error}")
            return
        with st.spinner("Descargando productos e imágenes desde ODOO..."):
            prods_odoo = cargar_todos_productos_odoo(odoo_url, odoo_db, odoo_user, odoo_pass)

        # 3. Sincronizar a Supabase (calcula phashes, sube imágenes, guarda filas)
        with st.spinner("Sincronizando con Supabase..."):
            sb_result = _sincronizar_a_supabase(prods_odoo)
        phashes     = sb_result["phashes"]
        imagen_urls = sb_result["imagen_urls"]

        # Si Supabase no está configurado, calcular phashes localmente y guardar pkl
        if not phashes:
            with st.spinner("Calculando hashes de imágenes..."):
                for p in prods_odoo:
                    if p.get("image_128"):
                        try:
                            h = _phash_imagen(base64.b64decode(p["image_128"]))
                            if h is not None:
                                phashes[p["default_code"]] = h
                        except Exception:
                            pass
            guardar_cache_odoo(skus, prods_odoo, phashes)

        # Quitar image_128 de memoria antes de guardar en session_state
        prods_sin_img = [{k: v for k, v in p.items() if k != "image_128"} for p in prods_odoo]
        _aplicar_datos_odoo(skus, prods_sin_img, phashes, imagen_urls=imagen_urls, desde_cache=False)

    if cargar_btn:
        _cargar_desde_odoo(forzar=False)

    if st.session_state.odoo_conectado:
        st.success(f"✅ ODOO activo — {len(st.session_state.odoo_skus)} SKUs en sesión")
    else:
        st.warning("⚠️ ODOO no conectado — presiona **Cargar SKUs**")

    # ── Limpiar productos _test ────────────────────────────────────────────────
    if st.button("🗑️ Eliminar productos _test", width="stretch",
                 help="Elimina de ODOO todos los productos de la categoría PRUEBAS_AGENTE"):
        odoo_url  = os.environ.get("ODOO_URL", "")
        odoo_db   = os.environ.get("ODOO_DB", "")
        odoo_user = os.environ.get("ODOO_USER", "")
        odoo_pass = os.environ.get("ODOO_PASSWORD", "")
        if not all([odoo_url, odoo_db, odoo_user, odoo_pass]):
            st.warning("Faltan credenciales ODOO en el .env")
        else:
            try:
                common = xmlrpc.client.ServerProxy(f"{odoo_url}/xmlrpc/2/common", allow_none=True, encoding="utf-8")
                uid    = common.authenticate(odoo_db, odoo_user, odoo_pass, {})
                models = xmlrpc.client.ServerProxy(f"{odoo_url}/xmlrpc/2/object", allow_none=True, encoding="utf-8")

                # Buscar la categoría PRUEBAS_AGENTE
                cat_ids = models.execute_kw(odoo_db, uid, odoo_pass,
                    "product.category", "search",
                    [[["name", "=", "PRUEBAS_AGENTE"]]])

                if not cat_ids:
                    st.warning("No se encontró la categoría PRUEBAS_AGENTE en ODOO.")
                else:
                    ids_a_borrar = models.execute_kw(odoo_db, uid, odoo_pass,
                        "product.template", "search",
                        [[["categ_id", "in", cat_ids]]])

                    if not ids_a_borrar:
                        st.info("No hay productos en PRUEBAS_AGENTE para eliminar.")
                    else:
                        # Intentar borrar; si hay restricciones de FK, archivar en su lugar
                        borrados  = 0
                        archivados = 0
                        errores_ind = []
                        for pid in ids_a_borrar:
                            try:
                                models.execute_kw(odoo_db, uid, odoo_pass,
                                    "product.template", "unlink", [[pid]])
                                borrados += 1
                            except Exception:
                                try:
                                    models.execute_kw(odoo_db, uid, odoo_pass,
                                        "product.template", "write",
                                        [[pid], {"active": False}])
                                    archivados += 1
                                except Exception as e2:
                                    errores_ind.append(str(e2))
                        partes = []
                        if borrados:
                            partes.append(f"{borrados} eliminado(s)")
                        if archivados:
                            partes.append(f"{archivados} archivado(s)")
                        if partes:
                            st.success(f"✅ {', '.join(partes)} de PRUEBAS_AGENTE.")
                        if errores_ind:
                            st.warning(f"⚠️ {len(errores_ind)} producto(s) no pudieron procesarse: {errores_ind[0]}")
            except Exception as e:
                st.error(f"Error al eliminar: {e}")

    # ── Eliminar productos de un contenedor ──────────────────────────────────
    st.divider()
    st.caption("🗑️ Eliminar productos de contenedor")
    _del_cont = st.text_input(
        "Contenedor a eliminar",
        placeholder="Ej: EITU9122770",
        key="eliminar_cont_input",
    )
    _del_ok = st.checkbox("Confirmo la eliminación", key="eliminar_cont_confirm")
    if st.button(
        "🗑️ Eliminar productos de Odoo",
        use_container_width=True,
        key="eliminar_cont_btn",
        disabled=not (bool(_del_cont) and _del_ok),
    ):
        odoo_url  = os.environ.get("ODOO_URL", "")
        odoo_db   = os.environ.get("ODOO_DB", "")
        odoo_user = os.environ.get("ODOO_USER", "")
        odoo_pass = os.environ.get("ODOO_PASSWORD", "")
        if not all([odoo_url, odoo_db, odoo_user, odoo_pass]):
            st.warning("Faltan credenciales ODOO en el .env")
        else:
            try:
                _del_common = xmlrpc.client.ServerProxy(
                    f"{odoo_url}/xmlrpc/2/common", allow_none=True, encoding="utf-8")
                _del_uid = _del_common.authenticate(odoo_db, odoo_user, odoo_pass, {})
                _del_mdl = xmlrpc.client.ServerProxy(
                    f"{odoo_url}/xmlrpc/2/object", allow_none=True, encoding="utf-8")

                _del_cats = _del_mdl.execute_kw(odoo_db, _del_uid, odoo_pass,
                    "product.category", "search", [[["name", "=", "Productos Agente"]]])
                if not _del_cats:
                    st.warning("No se encontró la categoría 'Productos Agente' en Odoo.")
                else:
                    _del_ids = _del_mdl.execute_kw(odoo_db, _del_uid, odoo_pass,
                        "product.template", "search",
                        [[["categ_id", "in", _del_cats],
                          ["container_numbers", "ilike", _del_cont.strip()]]])
                    if not _del_ids:
                        st.info(f"No se encontraron productos del contenedor '{_del_cont}' en Productos Agente.")
                    else:
                        with st.spinner(f"Eliminando {len(_del_ids)} productos..."):
                            _del_borrados = _del_archivados = 0
                            _del_errs = []
                            for _pid in _del_ids:
                                # 1. Intentar eliminar registros product.measurement que bloquean el FK
                                try:
                                    _meas_ids = _del_mdl.execute_kw(odoo_db, _del_uid, odoo_pass,
                                        "product.measurement", "search",
                                        [[["product_tmpl_id", "=", _pid]]])
                                    if _meas_ids:
                                        _del_mdl.execute_kw(odoo_db, _del_uid, odoo_pass,
                                            "product.measurement", "unlink", [_meas_ids])
                                except Exception:
                                    pass  # sin permisos en measurement — se intentará igual el unlink

                                # 2. Intentar borrado completo
                                try:
                                    _del_mdl.execute_kw(odoo_db, _del_uid, odoo_pass,
                                        "product.template", "unlink", [[_pid]])
                                    _del_borrados += 1
                                except Exception:
                                    # 3. Fallback: archivar si el FK sigue bloqueando
                                    try:
                                        _del_mdl.execute_kw(odoo_db, _del_uid, odoo_pass,
                                            "product.template", "write",
                                            [[_pid], {"active": False}])
                                        _del_archivados += 1
                                    except Exception as _e2:
                                        _del_errs.append(str(_e2))

                        _del_partes = []
                        if _del_borrados:
                            _del_partes.append(f"**{_del_borrados} eliminado(s)**")
                        if _del_archivados:
                            _del_partes.append(f"{_del_archivados} archivado(s) (FK pendiente)")
                        if _del_partes:
                            st.success(f"✅ {', '.join(_del_partes)} del contenedor {_del_cont}.")
                        if _del_errs:
                            st.warning(f"⚠️ {len(_del_errs)} no pudieron procesarse: {_del_errs[0]}")
            except Exception as _e:
                st.error(f"Error: {_e}")

    # ── Generar Master Costos rápido (sin pasar por todo el flujo) ───────────
    st.divider()
    st.caption("🧪 Prueba rápida de costos")
    _prods_test = st.session_state.get("productos", [])
    if _prods_test:
        if st.button("📊 Generar Master Costos", width="stretch",
                     help="Genera el Excel de costos con los productos ya cargados"):
            try:
                _tc   = st.session_state.get("tipo_cambio", 19.0)
                _cc   = st.session_state.get("costo_contenedor", 525000.0)
                _npk  = st.session_state.get("filename", "packing_list.xlsx")
                _mbytes = generar_excel_master(_prods_test, _tc, _cc, _npk)
                st.session_state.master_bytes = _mbytes
                st.rerun()
            except Exception as _e:
                st.error(f"Error: {_e}")
    else:
        st.caption("_(carga un packing list primero)_")

    st.divider()
    archivo = st.file_uploader("📂 Packing list (.xlsx)", type=["xlsx"])

    st.divider()
    if st.session_state.excel_bytes:
        filename   = st.session_state.get("filename", "packing_list.xlsx")
        nombre_out = filename.replace(".xlsx", "") + "_FERRAFORME.xlsx"
        st.success("✅ Excels listos para descargar")
        st.download_button(
            label="⬇️ Descargar Excel FERRAFORME",
            data=st.session_state.excel_bytes,
            file_name=nombre_out,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    if st.session_state.master_bytes:
        filename     = st.session_state.get("filename", "packing_list.xlsx")
        nombre_master = filename.replace(".xlsx", "") + "_MASTER.xlsx"
        st.download_button(
            label="⬇️ Descargar Excel Master Costos",
            data=st.session_state.master_bytes,
            file_name=nombre_master,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
        if st.button("🔄 Nuevo packing list", width="stretch"):
            # Limpiar carpeta de imágenes temporales
            if IMAGENES_TEMP_PATH.exists():
                for _f in IMAGENES_TEMP_PATH.iterdir():
                    try:
                        _f.unlink()
                    except Exception:
                        pass
            _chat_hist   = st.session_state.chat[:]
            _lc_hist     = st.session_state.lc_messages[:]
            _odoo_s      = st.session_state.odoo_skus
            _odoo_p      = st.session_state.odoo_productos
            _odoo_ph     = st.session_state.odoo_phashes
            _odoo_c      = st.session_state.odoo_conectado
            for k in defaults:
                st.session_state[k] = defaults[k]
            st.session_state.chat           = _chat_hist
            st.session_state.lc_messages    = _lc_hist
            st.session_state.odoo_skus      = _odoo_s
            st.session_state.odoo_productos = _odoo_p
            st.session_state.odoo_phashes   = _odoo_ph
            st.session_state.odoo_conectado = _odoo_c
            st.session_state.pop("filename", None)
            st.rerun()

    # ── Exportar desde Odoo ──────────────────────────────────────────────────
    st.divider()
    st.subheader("📤 Exportar desde Odoo")
    st.caption("Genera los Excels desde un contenedor ya cargado en Odoo, sin necesidad de re-subir el packing list.")

    _exp_cont = st.text_input(
        "Número de contenedor",
        placeholder="Ej: EITU9122770",
        key="exportar_od_contenedor",
    )
    if st.button(
        "🔍 Generar Excels desde Odoo",
        use_container_width=True,
        key="exportar_od_btn",
        disabled=not bool(_exp_cont),
    ):
        with st.spinner(f"Cargando productos de {_exp_cont} desde Odoo..."):
            _exp_res = generar_excels_desde_contenedor(_exp_cont.strip())
        if _exp_res.get("error"):
            st.error(f"Error: {_exp_res['error']}")
        else:
            st.session_state["exportar_od_resultado"]   = _exp_res
            st.session_state["exportar_od_contenedor_ok"] = _exp_cont.strip()
            st.rerun()

    if st.session_state.get("exportar_od_resultado"):
        _exp_r  = st.session_state["exportar_od_resultado"]
        _exp_ec = st.session_state.get("exportar_od_contenedor_ok", "")
        st.success(f"✅ {_exp_r.get('n_prods', 0)} productos recuperados de **{_exp_ec}**")
        _exp_c1, _exp_c2 = st.columns(2)
        with _exp_c1:
            st.download_button(
                "⬇️ Reporte Bodega",
                data=_exp_r["reporte_bytes"],
                file_name=f"{_exp_ec}_reporte_bodega.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="exp_od_dl_bodega",
            )
            st.download_button(
                "⬇️ Master Costos",
                data=_exp_r["master_bytes"],
                file_name=f"{_exp_ec}_master_costos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="exp_od_dl_master",
            )
        with _exp_c2:
            st.download_button(
                "⬇️ FERRAFORME",
                data=_exp_r["ferraforme_bytes"],
                file_name=f"{_exp_ec}_FERRAFORME.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="exp_od_dl_ferraforme",
            )
            st.download_button(
                "⬇️ Purchase Order",
                data=_exp_r["purchase_bytes"],
                file_name=f"{_exp_ec}_purchase.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="exp_od_dl_purchase",
            )

    # ── Historial de duplicados/variantes ─────────────────────────────────────
    _hist_dup = st.session_state.get("historial_duplicados", [])
    if _hist_dup:
        st.divider()
        with st.expander(f"📋 Historial de similares/variantes ({sum(len(e['grupos']) for e in _hist_dup)} grupos)"):
            _TIPO_LABEL = {
                "exacto":         "🔁 Exacto",
                "probable":       "❓ Probable",
                "similar":        "📦 Datos iguales",
                "nombre_similar": "🔤 Nombre parecido",
            }
            for entrada in reversed(_hist_dup):
                st.caption(f"**{entrada['archivo']}** — {entrada['fecha']}")
                for g in entrada["grupos"]:
                    tipo_lbl = _TIPO_LABEL.get(g["tipo"], g["tipo"])
                    dec = entrada["respuestas"].get(str(g["id"]))
                    if dec == "diferente":
                        dec_lbl = "🔗 Independientes"
                    elif isinstance(dec, dict):
                        tipo_d  = dec.get("tipo", "variantes")
                        n_sel   = len(dec.get("sel", []))
                        n_indep = len(g["indices"]) - n_sel
                        if tipo_d == "mismo":
                            dec_lbl = f"🔀 Fusionados ({n_sel})" + (f" · 🔗 {n_indep} indep." if n_indep else "")
                        else:
                            dec_lbl = f"📂 Variantes ({n_sel})" + (f" · 🔗 {n_indep} indep." if n_indep else "")
                    else:
                        dec_lbl = "📂 Variantes"
                    nombres = [p.get("nombre", "—") for p in g["productos"]]
                    filas   = [str(p.get("fila_excel_0idx", g["indices"][ci]) + 1) for ci, p in enumerate(g["productos"])]
                    st.markdown(
                        f"- {tipo_lbl} · {dec_lbl} · Filas {', '.join(filas)}:  \n"
                        + "  ".join(f"`{n}`" for n in nombres)
                    )
                st.divider()


def _agregar_producto_manual(
    nombre: str,
    precio_usd, piezas_x_caja, piezas_total, cajas_master,
    cbm_por_pieza, cbm_master_carton, cbm_total_sku,
    material: str, uso: str,
    imagen_bytes: bytes | None,
    imagen_ext: str | None,
) -> str:
    """Crea el producto manualmente, genera su SKU y lo agrega a la sesión. Devuelve el SKU."""
    n_manuales = sum(1 for p in st.session_state.get("productos", []) if p.get("_manual"))
    fila_0idx  = -(n_manuales + 1)   # índice negativo para no colisionar con filas del Excel

    prod: dict = {
        "nombre":            nombre,
        "precio_usd":        precio_usd        or None,
        "piezas_x_caja":     piezas_x_caja     or None,
        "piezas_total":      piezas_total       or None,
        "cajas_master":      cajas_master       or None,
        "cbm_por_pieza":     cbm_por_pieza      or None,
        "cbm_master_carton": cbm_master_carton  or None,
        "cbm_total_sku":     cbm_total_sku      or None,
        "material":          material           or "",
        "uso":               uso                or "",
        "fila_excel_0idx":   fila_0idx,
        "_manual":           True,
    }

    # Obtener metadata para el SKU
    if imagen_bytes and imagen_ext:
        datos = analizar_imagen_claude(imagen_bytes, imagen_ext, contexto=prod)
    else:
        datos = _inferir_categoria_manual(nombre, material, uso)

    sub_cod     = datos.get("subcategoria_cod", "VAR")
    att_cod     = datos.get("atributo_cod",     "GEN")
    sku_inicial = generar_sku(sub_cod, att_cod)
    skus_odoo   = st.session_state.get("odoo_skus", [])
    if skus_odoo:
        prod["sku"] = validar_sku_vs_odoo(sku_inicial, skus_odoo)["sku_ajustado"]
    else:
        prod["sku"] = sku_inicial

    if st.session_state.get("modo_prueba"):
        prod["sku"] += "_test"

    # Guardar imagen en carpeta temporal si existe
    if imagen_bytes and imagen_ext:
        try:
            stem = f"manual_{abs(fila_0idx)}"
            IMAGENES_TEMP_PATH.mkdir(parents=True, exist_ok=True)
            (IMAGENES_TEMP_PATH / f"{stem}.{imagen_ext}").write_bytes(imagen_bytes)
            prod["imagen_temp_stem"] = stem
        except Exception:
            pass

    st.session_state.productos.append(prod)
    return prod["sku"]


def _iniciar_chat(analisis, productos, advertencias, tipo_cambio, contenedor, n_imgs: int = 0):
    """Lanza el primer turno del chat después de resolver las dudas."""
    # ── Calcular CBM total y añadirlo a la tabla de resumen ya mostrada ────────
    total_cbm = sum(_cbm_total_fila(p) for p in productos)
    cbm_status = "⚠️ **Supera los 70 CBM** — verifica la capacidad del contenedor." if total_cbm > 70 else "✅ Dentro del límite de 70 CBM."
    cbm_line = f"\n\n---\n📦 **CBM Total del cargamento: {total_cbm:.4f}** — {cbm_status}"
    if n_imgs > 0:
        n_con_sku      = sum(1 for p in productos if p.get("sku"))
        sin_imagen     = [p for p in productos if p.get("sin_imagen")]
        errores_gemini = st.session_state.pop("gemini_errores", [])
        cbm_line      += f"\n🏷️ **SKUs generados: {n_con_sku} / {len(productos)} productos**"
        if errores_gemini:
            cbm_line += f"\n⚠️ **{len(errores_gemini)} imágenes no pudieron analizarse** — esos productos usaron valores por defecto. Error: {errores_gemini[0]}"
        if sin_imagen:
            filas = ", ".join(str(p.get("fila_sin_imagen", "?")) for p in sin_imagen)
            cbm_line += f"\n⚠️ **Sin imagen en filas:** {filas}"

    # ── Estado de validación ODOO ─────────────────────────────────────────────
    if st.session_state.odoo_conectado:
        n_skus_odoo  = len(st.session_state.odoo_skus)
        n_conflictos = len(st.session_state.resoluciones_conflictos)
        if n_conflictos > 0:
            cbm_line += f"\n🔀 **Validación ODOO:** {n_conflictos} conflicto(s) resueltos por el usuario"
        else:
            cbm_line += f"\n✅ **Validación ODOO:** sin conflictos detectados ({n_skus_odoo} SKUs comparados)"
    else:
        cbm_line += f"\n⚠️ **ODOO no conectado** — SKUs generados sin validar contra la base de datos"
    # Buscar el último mensaje del asistente y anexar el resumen al final
    for i in range(len(st.session_state.chat) - 1, -1, -1):
        if st.session_state.chat[i]["role"] == "assistant":
            c = st.session_state.chat[i]["content"]
            if isinstance(c, list):
                # Contenido rico (bloques_res de conflictos) → agregar como bloque de texto
                st.session_state.chat[i]["content"].append({"type": "text", "value": cbm_line})
            else:
                st.session_state.chat[i]["content"] += cbm_line
            break

    sku_info = f"{n_imgs} imágenes procesadas, {sum(1 for p in productos if p.get('sku'))} SKUs generados. " if n_imgs > 0 else ""
    primer_msg = (
        f"El usuario subió '{st.session_state.get('filename', '')}'. "
        f"El análisis y los {len(productos)} productos ya están listos. "
        f"{sku_info}"
        f"CBM total del cargamento: {total_cbm:.4f} ({'SUPERA los 70 CBM' if total_cbm > 70 else 'dentro de los 70 CBM'}). "
        f"Advertencias de datos: {advertencias if advertencias else 'ninguna'}. "
        "Preséntate brevemente, confirma cuántos productos encontraste "
        "y pregunta si puede proceder a generar el Excel."
    )
    system = build_system_prompt(
        analisis, productos, tipo_cambio, contenedor,
        respuestas_dudas=st.session_state.respuestas_dudas or None,
    )
    texto = llamar_agente(primer_msg, system, [])

    # Guardar en historial LangChain
    st.session_state.lc_messages.append(HumanMessage(content=primer_msg))
    st.session_state.lc_messages.append(AIMessage(content=texto))

    st.session_state.chat.append({"role": "assistant", "content": texto})


# ── Detectar nuevo archivo ─────────────────────────────────────────────────────
if archivo is not None:
    archivo_id = f"{archivo.name}_{archivo.size}"

    if archivo_id != st.session_state.archivo_id:
        # Preservar historial de chat y conexión ODOO al cambiar de archivo
        _chat_previo     = st.session_state.chat[:]
        _lc_previo       = st.session_state.lc_messages[:]
        _odoo_skus       = st.session_state.odoo_skus
        _odoo_productos  = st.session_state.odoo_productos
        _odoo_phashes    = st.session_state.odoo_phashes
        _odoo_conectado  = st.session_state.odoo_conectado
        for k in defaults:
            st.session_state[k] = defaults[k]
        st.session_state.chat            = _chat_previo
        st.session_state.lc_messages     = _lc_previo
        st.session_state.odoo_skus       = _odoo_skus
        st.session_state.odoo_productos  = _odoo_productos
        st.session_state.odoo_phashes    = _odoo_phashes
        st.session_state.odoo_conectado  = _odoo_conectado
        st.session_state.archivo_id  = archivo_id
        st.session_state["filename"] = archivo.name
        if not st.session_state.get("contenedor_val", "").strip():
            st.session_state.contenedor_val = extraer_contenedor(archivo.name)

        file_bytes = archivo.read()
        st.session_state.file_bytes = file_bytes

        with st.spinner("Analizando el formato del packing list..."):
            try:
                analisis = analizar_encabezados(file_bytes)
                st.session_state.analisis         = analisis
                _dr = analisis.get("dudas_relevantes", [])
                _dm = analisis.get("dudas_menores", [])
                # Normalizar: asegurar que son listas de dicts / listas de strings
                st.session_state.dudas_relevantes = [d for d in _dr if isinstance(d, dict)] if isinstance(_dr, list) else []
                st.session_state.dudas_menores    = [d for d in _dm if isinstance(d, str)]  if isinstance(_dm, list) else []

                if st.session_state.dudas_relevantes:
                    st.session_state.esperando_dudas = True

                    # ── Construir mensaje detallado de análisis para el historial ──
                    columnas  = analisis.get("columnas", {})
                    n_dudas   = len(st.session_state.dudas_relevantes)
                    CONFIANZA = {"alta": "✅", "media": "⚠️", "baja": "🔴", "no_encontrado": "❌"}
                    LABELS = {
                        "nombre_producto": "Nombre del producto",
                        "piezas_x_caja":  "Piezas por caja",
                        "largo_cm":       "Largo (cm)",
                        "ancho_cm":       "Ancho (cm)",
                        "alto_cm":        "Alto (cm)",
                        "cbm_por_pieza":     "CBM por pieza",
                        "cbm_master_carton": "CBM master carton",
                        "cbm_total_sku":     "CBM total SKU",
                        "precio_usd":     "Precio USD",
                    }

                    lineas_mapeo = []
                    for campo, label in LABELS.items():
                        _raw = columnas.get(campo, {})
                        info = _raw if isinstance(_raw, dict) else {}
                        conf = info.get("confianza", "no_encontrado")
                        enc  = info.get("encabezado_original") or "—"
                        mues = info.get("valor_muestra") or "—"
                        nota = f" *({info.get('nota')})*" if info.get("nota") else ""
                        lineas_mapeo.append(
                            f"| {CONFIANZA.get(conf,'❓')} **{label}** | `{enc}` | {mues}{nota} |"
                        )

                    lineas_dudas = []
                    for i, d in enumerate(st.session_state.dudas_relevantes):
                        lineas_dudas.append(f"**{i+1}.** {d['descripcion']}\n   → *{d['pregunta']}*")

                    lineas_menores = [f"- {dm}" for dm in st.session_state.dudas_menores]

                    msg_analisis = (
                        f"Analicé el archivo **{archivo.name}**.\n\n"
                        "**Mapeo de columnas detectado:**\n\n"
                        "| | Campo | Encabezado original | Valor muestra |\n"
                        "|---|---|---|---|\n"
                        + "\n".join(lineas_mapeo)
                        + (
                            f"\n\n**{n_dudas} duda{'s' if n_dudas != 1 else ''} relevante{'s' if n_dudas != 1 else ''} — respóndelas abajo antes de continuar:**\n\n"
                            + "\n\n".join(lineas_dudas)
                            if lineas_dudas else ""
                        )
                        + (
                            "\n\n**Avisos informativos:**\n" + "\n".join(lineas_menores)
                            if lineas_menores else ""
                        )
                    )
                    st.session_state.chat.append({"role": "assistant", "content": msg_analisis})
                else:
                    # Sin dudas relevantes → leer productos y arrancar chat directamente
                    # Auto-rellenar número de contenedor si el sidebar está vacío
                    _cont_excel = analisis.get("numero_contenedor", "")
                    if _cont_excel and not st.session_state.get("contenedor_val", ""):
                        st.session_state.contenedor_val = _cont_excel
                    if not st.session_state.odoo_conectado:
                        st.session_state.chat.append({
                            "role": "assistant",
                            "content": "⚠️ **ODOO no conectado** — los SKUs se generarán sin validar contra la base de datos. Conecta ODOO desde el panel izquierdo antes de subir el archivo para evitar duplicados."
                        })
                    productos, advertencias = leer_productos(
                        file_bytes, analisis["columnas"],
                        fila_encabezado=analisis.get("fila_encabezado", 1),
                    )
                    productos = corregir_cbm(productos, advertencias)
                    productos = corregir_cbm_inner(productos, advertencias, file_bytes, analisis["columnas"])
                    with st.spinner("Traduciendo y normalizando nombres..."):
                        productos = normalizar_nombres_productos(productos)
                    with st.spinner("Clasificando productos (Fase 1 — texto)..."):
                        propuestas = analizar_clasificacion_packing(file_bytes, productos, modo_fase1=True)
                    st.session_state.clasificacion_propuestas  = propuestas
                    st.session_state.clasificacion_activa      = True
                    st.session_state.productos                 = productos
                    st.session_state.advertencias_productos    = advertencias
                    st.session_state.vision_bloque_offset      = 0

            except Exception as e:
                st.session_state.chat.append({"role": "assistant", "content": f"Error al analizar: {e}"})

        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ══════════════════════════════════════════════════════════════════════════════

st.title("📦 FERRAFORME — Agente de Productos v2")

tab_pl, tab_agregar = st.tabs(["📦 Packing List", "➕ Agregar Productos"])

with tab_pl:
    # ── Historial de chat — siempre visible ───────────────────────────────────────
    for msg in st.session_state.chat:
        with st.chat_message(msg["role"]):
            content = msg["content"]
            if isinstance(content, str):
                st.markdown(content)
            else:
                # Lista de bloques: {"type": "text"|"image_bytes", ...}
                for bloque in content:
                    if not isinstance(bloque, dict):
                        st.markdown(str(bloque))
                        continue
                    if bloque["type"] == "text":
                        st.markdown(bloque["value"])
                    elif bloque["type"] == "image_bytes":
                        try:
                            st.image(
                                bloque["data"],
                                width=bloque.get("width", 160),
                                caption=bloque.get("caption", ""),
                            )
                        except Exception:
                            pass
                    elif bloque["type"] == "columns":
                        # Bloque de columnas: {"type": "columns", "items": [{"img": bytes, "caption": "...", "text": "..."}]}
                        items = bloque.get("items", [])
                        if items:
                            cols = st.columns(min(len(items), 3))
                            for ci, item in enumerate(items):
                                with cols[ci % 3]:
                                    if item.get("img"):
                                        try:
                                            st.image(item["img"], width=120, caption=item.get("caption", ""))
                                        except Exception:
                                            pass
                                    if item.get("text"):
                                        st.markdown(item["text"])

    # ── Botones de descarga en área principal ────────────────────────────────────
    _eb = st.session_state.get("excel_bytes")
    _mb = st.session_state.get("master_bytes")
    _pb = st.session_state.get("purchase_bytes")
    if _eb or _mb or _pb:
        _fn = st.session_state.get("filename", "packing_list.xlsx").replace(".xlsx", "")
        st.divider()
        _dcols = st.columns(2)
        with _dcols[0]:
            if _eb:
                st.download_button(
                    label="⬇️ Descargar Excel FERRAFORME",
                    data=_eb,
                    file_name=f"{_fn}_FERRAFORME.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
        with _dcols[1]:
            if _mb:
                st.download_button(
                    label="⬇️ Descargar Excel Master Costos",
                    data=_mb,
                    file_name=f"{_fn}_MASTER.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        # ── Reporte clasificación ─────────────────────────────────────────────
        _rb = st.session_state.get("clasificacion_reporte_bytes")
        if _rb:
            st.download_button(
                label="⬇️ Descargar Reporte Clasificación (bodega)",
                data=_rb,
                file_name=f"{_fn}_CLASIFICACION.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # ── Purchase Order ────────────────────────────────────────────────────
        if _eb and _mb:
            _pcols = st.columns([1, 1])
            with _pcols[0]:
                if st.button("🛒 Crear Purchase", use_container_width=True,
                             help="Genera el Excel de Purchase Order listo para importar a Odoo"):
                    try:
                        _prods_po    = st.session_state.get("productos", [])
                        _tc_po       = st.session_state.get("tipo_cambio", 19.0)
                        _cc_po       = st.session_state.get("costo_contenedor", 525000.0)
                        _pb_new      = generar_excel_purchase(_prods_po, _tc_po, _cc_po)
                        st.session_state.purchase_bytes = _pb_new
                        st.rerun()
                    except Exception as _e:
                        st.error(f"Error al generar Purchase: {_e}")
            with _pcols[1]:
                if _pb:
                    st.download_button(
                        label="⬇️ Descargar Purchase Order",
                        data=_pb,
                        file_name=f"{_fn}_PURCHASE.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )

    # ── Snapshot de duplicados/variantes (editable, visible tras confirmar) ──
    _snap = st.session_state.get("dup_snapshot")
    if _snap and not st.session_state.get("esperando_duplicados"):
        _snap_grupos = _snap.get("grupos", [])
        _snap_resp   = _snap.get("respuestas", {})
        _snap_imgs   = _snap.get("imagenes", {})
        _n_grupos    = len(_snap_grupos)
        _TIPO_LABEL_S = {
            "exacto":          "🔁 Duplicado exacto",
            "probable":        "❓ Muy similar",
            "similar":         "🈁 Posible variante",
            "nombre_similar":  "🔤 Nombre parecido",
            "variante_imagen": "🖼️ Imagen similar",
        }
        _DEC_OPTS = ["Mismo producto", "Variante", "Productos diferentes"]
        with st.expander(
            f"📋 Similares/variantes detectados: {_n_grupos} grupo(s) — {_snap.get('fecha', '')}",
            expanded=False,
        ):
            st.caption("Puedes corregir las decisiones si algo quedó mal y luego aplicar los cambios.")
            _nuevas_resp: dict = {}
            for _sg in _snap_grupos:
                _sgid   = str(_sg["id"])
                _sprods = _sg["productos"]
                _sp0    = _sprods[0]
                _sfilas = [
                    str(p.get("fila_excel_0idx", _sg["indices"][ci]) + 1)
                    for ci, p in enumerate(_sprods)
                ]
                _sdec_orig = _snap_resp.get(_sgid)
                if _sdec_orig == "diferente":
                    _def_idx = 2
                elif isinstance(_sdec_orig, dict) and _sdec_orig.get("tipo") == "mismo":
                    _def_idx = 0
                else:
                    _def_idx = 1

                _sc1, _sc2, _sc3 = st.columns([3, 5, 3])
                with _sc1:
                    _stcols = st.columns(min(len(_sprods), 4))
                    for _ci, (_sp, _stc) in enumerate(zip(_sprods[:4], _stcols)):
                        with _stc:
                            _sf0 = _sp.get("fila_excel_0idx")
                            _si  = _snap_imgs.get(_sf0) if _sf0 is not None else None
                            if _si:
                                st.image(_si["data"], width=55)
                            st.caption(f"F{_sfilas[_ci]}")
                with _sc2:
                    st.markdown(f"**{_sp0.get('nombre', '—')}**")
                    _scn = _sp0.get("nombre_chino_orig", "")
                    if _scn and _tiene_chino(_scn):
                        st.caption(f"🈁 {_scn}")
                    st.caption(
                        f"{_TIPO_LABEL_S.get(_sg['tipo'], _sg['tipo'])}  \n"
                        f"Filas: {', '.join(_sfilas)}"
                    )
                    if len(_sprods) > 1:
                        _sotros = [p.get("nombre", "—") for p in _sprods[1:4]]
                        st.caption("Con: " + " · ".join(f"`{n}`" for n in _sotros))
                with _sc3:
                    _nueva = st.selectbox(
                        "Decisión",
                        _DEC_OPTS,
                        index=_def_idx,
                        key=f"snap_dec_{_sgid}",
                        label_visibility="collapsed",
                    )
                    _nuevas_resp[_sgid] = _nueva
                st.divider()

            if st.button("🔄 Aplicar correcciones", key="snap_aplicar"):
                # Convertir opciones de texto a formato interno
                _resp_corr: dict = {}
                for _sgid2, _dec_txt in _nuevas_resp.items():
                    _gidx = next((i for i, g in enumerate(_snap_grupos) if str(g["id"]) == _sgid2), None)
                    if _gidx is None:
                        continue
                    _gx = _snap_grupos[_gidx]
                    if _dec_txt == "Productos diferentes":
                        _resp_corr[_sgid2] = "diferente"
                    elif _dec_txt == "Mismo producto":
                        _resp_corr[_sgid2] = {"tipo": "mismo", "sel": list(range(len(_gx["productos"])))}
                    else:
                        _resp_corr[_sgid2] = {"tipo": "variantes", "sel": list(range(len(_gx["productos"])))}
                # Re-entrar a paso 2 con las correcciones
                _prods_orig = _snap.get("productos_orig")
                if _prods_orig:
                    st.session_state.productos             = _prods_orig
                    st.session_state.duplicados_pendientes = _snap_grupos
                    st.session_state.imagenes_excel        = _snap_imgs
                    st.session_state.dup_respuestas        = _resp_corr
                    st.session_state.dup_paso              = 2
                    st.session_state.esperando_duplicados  = True
                    st.session_state.dup_snapshot          = None
                    st.rerun()

    # Scroll automático al último mensaje del chat
    if st.session_state.chat:
        st.html(
            """<script>
            var main = window.parent.document.querySelector('section.main');
            if (main) main.scrollTo(0, main.scrollHeight);
            </script>"""
        )



# ── PASO: Resolver dudas relevantes ───────────────────────────────────────────
if st.session_state.esperando_dudas:
    with tab_pl:
        analisis         = st.session_state.analisis
        dudas_relevantes = st.session_state.dudas_relevantes
        dudas_menores    = st.session_state.dudas_menores

        st.subheader("El agente necesita aclarar algunas dudas antes de continuar")
        st.markdown(
            f"Se detectó el archivo en **{analisis.get('idioma_detectado', 'idioma desconocido')}**. "
            "Responde las siguientes preguntas para asegurar que el Excel se genere correctamente."
        )
        st.divider()

        respuestas_temp = {}

        for duda in dudas_relevantes:
            if not isinstance(duda, dict):
                continue
            duda_id = str(duda["id"])
            st.markdown(f"**{duda['id'] + 1}. {duda['descripcion']}**")

            tipo    = duda.get("tipo", "confirmar")
            opciones = duda.get("opciones", ["Sí, continuar", "No, revisar"])
            default  = duda.get("default", opciones[0])
            idx_default = opciones.index(default) if default in opciones else 0

            if tipo in ("eleccion", "confirmar"):
                respuesta = st.radio(
                    duda["pregunta"],
                    opciones,
                    index=idx_default,
                    key=f"duda_radio_{duda_id}",
                    horizontal=True,
                )
            elif tipo == "texto":
                respuesta = st.text_input(
                    duda["pregunta"],
                    value=duda.get("default", ""),
                    key=f"duda_texto_{duda_id}",
                )
            else:
                respuesta = default

            respuestas_temp[duda_id] = respuesta
            st.divider()

        # Dudas menores — solo informativas
        if dudas_menores:
            with st.expander(f"ℹ️ {len(dudas_menores)} avisos informativos (no bloquean la generación)"):
                for dm in dudas_menores:
                    st.markdown(f"- {dm}")

        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("✅ Continuar con estas respuestas", type="primary", width="stretch"):
                st.session_state.respuestas_dudas = respuestas_temp

                # Mover aclaraciones al historial principal antes de continuar
                for m in st.session_state.chat_fase:
                    st.session_state.chat.append(m)
                st.session_state.chat_fase = []

                # Agregar resumen de respuestas del usuario al historial
                resumen_resp = "**Mis respuestas:**\n" + "\n".join(
                    f"- **{dudas_relevantes[int(k)].get('descripcion', '')[:60]}...** → {v}"
                    for k, v in respuestas_temp.items()
                    if int(k) < len(dudas_relevantes) and isinstance(dudas_relevantes[int(k)], dict)
                )
                st.session_state.chat.append({"role": "user", "content": resumen_resp})

                # Aplicar respuestas al mapeo de columnas
                columnas_actualizadas = aplicar_respuestas(
                    analisis["columnas"], dudas_relevantes, respuestas_temp
                )
                analisis_actualizado = {**analisis, "columnas": columnas_actualizadas}
                st.session_state.analisis        = analisis_actualizado
                st.session_state.esperando_dudas = False

                # Auto-rellenar número de contenedor si el sidebar está vacío
                _cont_excel = analisis_actualizado.get("numero_contenedor", "")
                if _cont_excel and not st.session_state.get("contenedor_val", ""):
                    st.session_state.contenedor_val = _cont_excel

                # Leer productos, enriquecer con imágenes y arrancar chat
                productos, advertencias = leer_productos(
                    st.session_state.file_bytes, columnas_actualizadas,
                    fila_encabezado=st.session_state.analisis.get("fila_encabezado", 1),
                )
                productos = corregir_cbm(productos, advertencias)
                productos = corregir_cbm_inner(productos, advertencias, st.session_state.file_bytes, columnas_actualizadas)
                with st.spinner("Traduciendo y normalizando nombres..."):
                    productos = normalizar_nombres_productos(productos)
                with st.spinner("Clasificando productos (Fase 1 — texto)..."):
                    propuestas = analizar_clasificacion_packing(st.session_state.file_bytes, productos, modo_fase1=True)
                st.session_state.clasificacion_propuestas  = propuestas
                st.session_state.clasificacion_activa      = True
                st.session_state.productos                 = productos
                st.session_state.advertencias_productos    = advertencias
                st.session_state.vision_bloque_offset      = 0

                st.rerun()

        with col_b:
            if st.button("↩️ Cancelar y revisar el archivo", width="stretch"):
                _chat_c  = st.session_state.chat[:]
                _lc_c    = st.session_state.lc_messages[:]
                for k in defaults:
                    st.session_state[k] = defaults[k]
                st.session_state.chat        = _chat_c
                st.session_state.lc_messages = _lc_c
                st.rerun()

        # ── Chat de aclaraciones (dentro de la fase de dudas) ─────────────────────
        if st.session_state.chat_fase:
            st.divider()
            for msg in st.session_state.chat_fase:
                with st.chat_message(msg["role"]):
                    st.markdown(msg["content"])


# ── PASO: Resolver filas duplicadas ──────────────────────────────────────────
if st.session_state.esperando_duplicados:
    with tab_pl:
        grupos_dup    = st.session_state.duplicados_pendientes
        productos_dup = st.session_state.productos
        imagenes_dup  = st.session_state.get("imagenes_excel", {})
        dup_paso      = st.session_state.get("dup_paso", 1)

        # ════════════════════════════════════════════════════════════════════════
        # PASO 1 — Auto-resolver exactos; pedir decisión solo para el resto
        # ════════════════════════════════════════════════════════════════════════
        if dup_paso == 1:
            # Todos los grupos son editables — los "exacto" tienen Mismo producto pre-seleccionado
            todos_grupos = grupos_dup

            st.subheader(f"🔍 {len(todos_grupos)} grupo(s) detectados — revisa y corrige si es necesario")
            st.caption("Puedes cambiar la decisión de cualquier grupo, incluso los que detectamos con alta confianza.")

            # Encabezados
            _hA, _hB, _hC, _hD = st.columns([3, 4, 3, 3])
            with _hA: st.markdown("**Imágenes**")
            with _hB: st.markdown("**Producto / Filas**")
            with _hC: st.markdown("**Detectado como**")
            with _hD: st.markdown("**Tu decisión**")
            st.divider()

            for grupo in todos_grupos:
                gid    = str(grupo["id"])
                prods  = grupo["productos"]
                p0     = prods[0]
                tipo_g = grupo["tipo"]
                n_p    = len(prods)
                filas  = [str(p.get("fila_excel_0idx", grupo["indices"][ci]) + 1)
                          for ci, p in enumerate(prods)]
                diffs  = grupo.get("diffs", {})

                tipo_badge = {
                    "exacto":          "✅ Mismo producto (alta confianza)",
                    "probable":        "❓ Posible mismo producto",
                    "similar":         "🈁 Posible variante (nombre chino similar)",
                    "nombre_similar":  "🔤 Nombres parecidos",
                    "variante_imagen": "🖼️ Imágenes similares",
                }.get(tipo_g, tipo_g)

                default_idx = {
                    "exacto":          0,  # Mismo producto
                    "probable":        0,  # Mismo producto
                    "similar":         1,  # Variante
                    "variante_imagen": 1,  # Variante
                    "nombre_similar":  2,  # Productos diferentes
                }.get(tipo_g, 2)

                col_imgs, col_info, col_tipo, col_dec = st.columns([3, 4, 3, 3])

                # ── Thumbnails ──────────────────────────────────────────────────
                with col_imgs:
                    n_thumb = min(n_p, 5)
                    tcols   = st.columns(n_thumb)
                    for ci in range(n_thumb):
                        with tcols[ci]:
                            fila_0 = prods[ci].get("fila_excel_0idx")
                            img_t  = imagenes_dup.get(fila_0) if fila_0 is not None else None
                            if img_t:
                                st.image(img_t["data"], width=65)
                            else:
                                st.caption("—")
                    if n_p > 5:
                        st.caption(f"+{n_p - 5} más")

                # ── Info ────────────────────────────────────────────────────────
                with col_info:
                    st.markdown(f"**{p0.get('nombre', '—')}**")
                    _cn_orig = p0.get("nombre_chino_orig", "")
                    if _cn_orig and _tiene_chino(_cn_orig):
                        st.caption(f"🈁 {_cn_orig}")
                    st.caption(f"Filas: {', '.join(filas)}")
                    if diffs:
                        diff_parts = []
                        for campo, vals in list(diffs.items())[:3]:
                            lbl = _LABELS_DIFF.get(campo, campo)
                            diff_parts.append(f"{lbl}: {' / '.join(str(v) for v in vals)}")
                        st.caption("↕ " + " · ".join(diff_parts))

                # ── Tipo detectado ───────────────────────────────────────────────
                with col_tipo:
                    st.caption(tipo_badge)

                # ── Decisión ────────────────────────────────────────────────────
                with col_dec:
                    dec = st.selectbox(
                        "Decisión",
                        ["Mismo producto", "Variante", "Productos diferentes"],
                        index=default_idx,
                        key=f"dup1_{gid}",
                        label_visibility="collapsed",
                    )

                # ── Sub-grupos ───────────────────────────────────────────────────
                # Mostrar siempre que haya más de 1 producto y la decisión no sea "diferentes"
                if dec != "Productos diferentes" and n_p > 1:
                    with st.expander(
                        f"✂️ Separar en sub-grupos ({n_p} productos) — haz clic si alguno no pertenece aquí",
                        expanded=False,
                    ):
                        st.caption(
                            "Mismo número → van juntos en el mismo SKU base  ·  "
                            "Número distinto → se tratan como grupos separados  ·  "
                            "Independiente → SKU completamente propio"
                        )
                        sub_cols = st.columns(min(n_p, 6))
                        for ci, prod_s in enumerate(prods):
                            with sub_cols[ci % 6]:
                                fila_0s = prod_s.get("fila_excel_0idx")
                                img_s   = imagenes_dup.get(fila_0s) if fila_0s is not None else None
                                if img_s:
                                    st.image(img_s["data"], width=90)
                                _nom_s = prod_s.get("nombre", f"Prod {ci+1}")
                                st.caption(f"F{filas[ci]}: {_nom_s[:25]}")
                                st.selectbox(
                                    "Sub-grupo",
                                    ["1", "2", "3", "4", "5", "Independiente"],
                                    index=0,
                                    key=f"dup_sub_{gid}_{ci}",
                                    label_visibility="collapsed",
                                )

                st.divider()

            # ── Botón de confirmación ──────────────────────────────────────────
            if st.button("✅ Confirmar — generar SKUs →", type="primary", width="stretch"):
                respuestas_paso1: dict = {}
                # Leer decisión de TODOS los grupos (incluyendo los antes "auto")
                for g in todos_grupos:
                    gid_b   = str(g["id"])
                    dec_val = st.session_state.get(f"dup1_{gid_b}", "Mismo producto")
                    if "Productos diferentes" in dec_val:
                        respuestas_paso1[gid_b] = "diferente"
                    else:
                        tipo = "mismo" if dec_val == "Mismo producto" else "variantes"
                        sub_map: dict[str, list[int]] = {}
                        for ci in range(len(g["productos"])):
                            sg = st.session_state.get(f"dup_sub_{gid_b}_{ci}", "1")
                            if sg != "Independiente":
                                sub_map.setdefault(sg, []).append(ci)
                        subgrupos = list(sub_map.values())
                        if len(subgrupos) <= 1:
                            respuestas_paso1[gid_b] = {"tipo": tipo, "sel": subgrupos[0] if subgrupos else list(range(len(g["productos"])))}
                        else:
                            respuestas_paso1[gid_b] = {"tipo": tipo, "subgrupos": subgrupos}
                st.session_state.dup_respuestas = respuestas_paso1
                st.session_state.dup_paso = 2
                st.rerun()

            # ── Mini-chat de aclaraciones (dentro del panel de duplicados) ────
            if st.session_state.chat_fase:
                st.divider()
                for _fcm in st.session_state.chat_fase:
                    with st.chat_message(_fcm["role"]):
                        st.markdown(_fcm["content"])

        # ════════════════════════════════════════════════════════════════════════
        # PASO 2 — Ejecutar resolución y proceder directamente con SKUs
        # ════════════════════════════════════════════════════════════════════════
        elif dup_paso == 2:
            respuestas_finales  = dict(st.session_state.get("dup_respuestas", {}))
            productos_resueltos = aplicar_resolucion_duplicados(
                productos_dup, grupos_dup, respuestas_finales
            )

            # Guardar historial
            st.session_state.historial_duplicados = (
                st.session_state.get("historial_duplicados", []) + [{
                    "fecha":      datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "archivo":    st.session_state.get("filename", "—"),
                    "grupos":     grupos_dup,
                    "respuestas": respuestas_finales,
                }]
            )

            # Proceder directamente a generar SKUs
            st.session_state.esperando_duplicados  = False
            st.session_state.dup_paso              = 1
            st.session_state.dup_respuestas        = {}
            st.session_state.duplicados_pendientes = []
            st.session_state.dup_grupos_segunda    = []
            st.session_state.chat_fase             = []

            analisis_res     = st.session_state.analisis
            advertencias_res = st.session_state.get("advertencias_productos", [])
            tipo_cambio_res  = st.session_state.get("tipo_cambio", 17.5)
            contenedor_res   = st.session_state.get("contenedor_val", "")

            with st.spinner("Analizando imágenes y generando SKUs..."):
                prods_enr, n_imgs_res, conflictos_res = enriquecer_con_imagenes(
                    st.session_state.file_bytes, productos_resueltos
                )
            st.session_state.productos              = prods_enr
            st.session_state.advertencias_productos = advertencias_res
            st.session_state.n_imgs_procesadas      = n_imgs_res

            # Guardar snapshot visual antes de limpiar imágenes
            import copy as _copy
            st.session_state.dup_snapshot = {
                "grupos":         grupos_dup,
                "respuestas":     respuestas_finales,
                "imagenes":       dict(st.session_state.get("imagenes_excel", {})),
                "productos_orig": _copy.deepcopy(productos_dup),  # pre-resolución
                "fecha":          datetime.now().strftime("%Y-%m-%d %H:%M"),
                "archivo":        st.session_state.get("filename", "—"),
            }
            st.session_state.imagenes_excel         = {}

            if conflictos_res:
                st.session_state.conflictos_pendientes = conflictos_res
                st.session_state.esperando_conflictos  = True
            else:
                _, errs_ren = renombrar_imagenes_con_sku(prods_enr)
                _reportar_errores_imagenes(errs_ren, "renombrado de imágenes")
                with st.spinner("Iniciando el agente..."):
                    _iniciar_chat(analisis_res, prods_enr, advertencias_res,
                                  tipo_cambio_res, contenedor_res, n_imgs_res)
            st.rerun()



# ── PASO: Resolver conflictos de SKU con ODOO ────────────────────────────────
if st.session_state.esperando_conflictos:
    with tab_pl:
        conflictos   = st.session_state.conflictos_pendientes
        analisis_cf  = st.session_state.analisis
        productos_cf = st.session_state.productos
        advertencias_cf = st.session_state.get("advertencias_productos", [])
        n_imgs_cf       = st.session_state.get("n_imgs_procesadas", 0)

        st.subheader(f"⚠️ Posibles duplicados detectados ({len(conflictos)})")
        st.markdown(
            "Los siguientes productos coinciden con productos ya existentes en ODOO "
            "por **prefijo SKU**, **imagen similar** o **nombre similar**. "
            "Indica si es el mismo producto o uno diferente."
        )
        st.divider()

        resoluciones_temp = dict(st.session_state.resoluciones_conflictos)

        for conflicto in conflictos:
            c_idx        = conflicto["idx"]
            c_nombre     = conflicto["nombre"]
            c_sku_prop   = conflicto["sku_propuesto"]
            c_sku_aj     = conflicto["sku_ajustado"]
            c_datos_gem  = conflicto.get("datos_gemini", {})
            c_prods_odoo = conflicto.get("productos_odoo", [])
            c_razon      = conflicto.get("razon", "sku")
            c_similares  = conflicto.get("similares_odoo", [])

            # Badge de razón
            razon_badge = {
                "sku":               "🔑 Mismo prefijo SKU",
                "imagen similar":    "🖼️ Imagen similar",
                "nombre similar":    "🔤 Nombre similar",
                "semántica similar": "🧠 Semántica similar (RAG)",
            }.get(c_razon, f"⚠️ {c_razon}")

            st.markdown(f"### Producto: **{c_nombre}**")
            st.markdown(f"**Razón:** {razon_badge}")
            st.markdown(f"- SKU generado: `{c_sku_prop}`")
            if c_sku_aj != c_sku_prop:
                st.markdown(f"- SKU alternativo (max+1): `{c_sku_aj}`")

            # Imagen del producto extraída del Excel
            prod_excel = productos_cf[c_idx] if c_idx < len(productos_cf) else {}
            stem_excel = prod_excel.get("imagen_temp_stem")
            if stem_excel and IMAGENES_TEMP_PATH.exists():
                archivo_img = next(
                    (f for f in IMAGENES_TEMP_PATH.iterdir() if f.stem == stem_excel),
                    None,
                )
                if archivo_img:
                    st.markdown("**Tu producto (del Excel):**")
                    st.image(str(archivo_img), width=160)

            # Scores de similitud si vienen de búsqueda visual/nombre
            if c_similares:
                for sim in c_similares[:3]:
                    badges = []
                    if sim.get("por_imagen") and sim.get("similitud_imagen") is not None:
                        pct = int(sim["similitud_imagen"] * 100)
                        badges.append(f"🖼️ imagen {pct}%")
                    if sim.get("por_nombre"):
                        pct = int(sim["similitud_nombre"] * 100)
                        badges.append(f"🔤 nombre {pct}%")
                    if sim.get("por_rag"):
                        pct = int(sim["similitud"] * 100)
                        badges.append(f"🧠 RAG {pct}%")
                    st.caption(f"`{sim['sku']}` — {sim['nombre']} — {' · '.join(badges)}")

            if c_prods_odoo:
                st.markdown("**Productos existentes en ODOO con ese prefijo:**")
                n_cols = min(len(c_prods_odoo), 3)
                cols_odoo = st.columns(n_cols)
                for ci, prod_odoo in enumerate(c_prods_odoo):
                    with cols_odoo[ci % n_cols]:
                        if prod_odoo.get("image_128"):
                            try:
                                img_bytes_odoo = base64.b64decode(prod_odoo["image_128"])
                                st.image(img_bytes_odoo, width=120)
                            except Exception:
                                pass
                        categ_raw  = prod_odoo.get("categ_id", [None, "—"])
                        categ_name = categ_raw[1] if isinstance(categ_raw, (list, tuple)) and len(categ_raw) > 1 else str(categ_raw)
                        st.markdown(
                            f"**SKU:** `{prod_odoo.get('default_code', '—')}`  \n"
                            f"**Nombre:** {prod_odoo.get('name', '—')}  \n"
                            f"**Categoría:** {categ_name}  \n"
                            f"**Descripción:** {prod_odoo.get('description_sale') or '—'}  \n"
                            f"**Precio:** ${prod_odoo.get('list_price', 0):.2f}"
                        )

            decision = st.radio(
                "¿Este producto nuevo es el mismo que alguno de arriba?",
                ["Diferente — asignar SKU nuevo", "Mismo producto — reutilizar SKU de ODOO"],
                key=f"conflicto_tipo_{c_idx}",
                horizontal=True,
            )

            res = resoluciones_temp.setdefault(str(c_idx), {})

            if "Mismo" in decision and c_prods_odoo:
                opciones_skus_odoo = [p.get("default_code", "") for p in c_prods_odoo]
                sku_sel = st.selectbox(
                    "SKU de ODOO a reutilizar:",
                    opciones_skus_odoo,
                    key=f"conflicto_sku_sel_{c_idx}",
                )
                prod_sel = next((p for p in c_prods_odoo if p.get("default_code") == sku_sel), {})
                categ_sel_raw  = prod_sel.get("categ_id", [None, ""])
                categ_sel_name = categ_sel_raw[1] if isinstance(categ_sel_raw, (list, tuple)) and len(categ_sel_raw) > 1 else str(categ_sel_raw)

                res["tipo"]        = "mismo"
                res["sku_final"]   = sku_sel
                res["nombre"]      = st.text_input(
                    "Nombre del producto:", value=prod_sel.get("name", c_nombre),
                    key=f"conflicto_nom_{c_idx}"
                )
                res["descripcion"] = st.text_input(
                    "Descripción:", value=prod_sel.get("description_sale") or c_datos_gem.get("descripcion", ""),
                    key=f"conflicto_desc_{c_idx}"
                )
                res["categoria"]   = st.text_input(
                    "Categoría:", value=categ_sel_name or c_datos_gem.get("categoria", ""),
                    key=f"conflicto_cat_{c_idx}"
                )
                res["atributo"]    = st.text_input(
                    "Atributo:", value=c_datos_gem.get("atributo_desc", ""),
                    key=f"conflicto_attr_{c_idx}"
                )
            else:
                res["tipo"]      = "diferente"
                res["sku_final"] = c_sku_aj

            st.divider()

        # ── Mini-chat de aclaraciones (dentro del panel de conflictos) ───────
        if st.session_state.chat_fase:
            st.divider()
            for _fcm in st.session_state.chat_fase:
                with st.chat_message(_fcm["role"]):
                    st.markdown(_fcm["content"])

        if st.button("✅ Confirmar todas las resoluciones", type="primary", width="stretch"):
            st.session_state.resoluciones_conflictos = resoluciones_temp
            productos_finales = aplicar_resoluciones_conflictos(
                productos_cf, conflictos, resoluciones_temp
            )
            st.session_state.productos             = productos_finales
            st.session_state.esperando_conflictos  = False
            st.session_state.conflictos_pendientes = []
            st.session_state.chat_fase             = []

            # SKUs definitivos confirmados → renombrar imágenes temp
            _, errs_ren = renombrar_imagenes_con_sku(productos_finales)
            _reportar_errores_imagenes(errs_ren, "renombrado de imágenes")


            # Guardar resumen rico de resoluciones en el historial del chat
            bloques_res = [{"type": "text", "value": f"**Resolución de {len(conflictos)} conflicto(s) SKU:**"}]
            for c in conflictos:
                c_idx_r    = c["idx"]
                res        = resoluciones_temp.get(str(c_idx_r), {})
                nombre_r   = c.get("nombre", f"Producto {c_idx_r + 1}")
                razon_r    = {"sku": "🔑 Mismo prefijo SKU", "imagen similar": "🖼️ Imagen similar",
                              "nombre similar": "🔤 Nombre similar", "semántica similar": "🧠 Semántica similar"
                              }.get(c.get("razon", "sku"), f"⚠️ {c.get('razon', '')}")
                if res.get("tipo") == "mismo":
                    decision_txt = f"✅ Reutilizado SKU `{res.get('sku_final', '')}` de ODOO"
                else:
                    decision_txt = f"🆕 SKU nuevo asignado: `{c['sku_ajustado']}`"
                bloques_res.append({"type": "text", "value": f"---\n### {nombre_r}\n**Razón:** {razon_r}  \n{decision_txt}"})

                # Imagen del producto del Excel (aún en carpeta temp antes de borrar)
                prod_r    = productos_cf[c_idx_r] if c_idx_r < len(productos_cf) else {}
                stem_r    = prod_r.get("imagen_temp_stem")
                img_excel = None
                if stem_r and IMAGENES_TEMP_PATH.exists():
                    arch_r = next((f for f in IMAGENES_TEMP_PATH.iterdir() if f.stem == stem_r), None)
                    if arch_r:
                        try:
                            img_excel = arch_r.read_bytes()
                        except Exception:
                            pass

                # Imágenes de ODOO similares
                odoo_items = []
                for prod_odoo in c.get("productos_odoo", [])[:3]:
                    item = {
                        "text": (
                            f"**SKU:** `{prod_odoo.get('default_code', '—')}`  \n"
                            f"**Nombre:** {prod_odoo.get('name', '—')}  \n"
                            f"**Precio:** ${prod_odoo.get('list_price', 0):.2f}"
                        ),
                        "img": None,
                        "caption": prod_odoo.get("default_code", ""),
                    }
                    if prod_odoo.get("image_128"):
                        try:
                            item["img"] = base64.b64decode(prod_odoo["image_128"])
                        except Exception:
                            pass
                    odoo_items.append(item)

                # Añadir imagen del Excel y columnas de ODOO al bloque
                if img_excel:
                    bloques_res.append({"type": "text", "value": "**Tu producto (del Excel):**"})
                    bloques_res.append({"type": "image_bytes", "data": img_excel, "width": 160, "caption": nombre_r})
                if odoo_items:
                    bloques_res.append({"type": "text", "value": "**Productos similares en ODOO:**"})
                    bloques_res.append({"type": "columns", "items": odoo_items})

            st.session_state.chat.append({"role": "assistant", "content": bloques_res})

            with st.spinner("Iniciando el agente..."):
                _iniciar_chat(analisis_cf, productos_finales, advertencias_cf, tipo_cambio, contenedor, n_imgs_cf)
            st.rerun()


# ── PASO: Revisar clasificación padre/variante ────────────────────────────────
if st.session_state.get("clasificacion_activa"):
    with tab_pl:
        _props      = st.session_state.clasificacion_propuestas
        _n_props    = len(_props)
        _analisis_c = st.session_state.get("analisis", {})
        _advertencias_c = st.session_state.get("advertencias_productos", [])
        _productos_c    = st.session_state.get("productos", [])

        st.subheader(f"🗂️ Clasificación de {_n_props} producto(s) — revisa y confirma")
        st.caption(
            "El agente analizó cada producto y propuso cómo clasificarlos en Odoo. "
            "Marca **Requiere revisión** en los que quieras que bodega revise manualmente."
        )

        # ── Estado de Odoo y modo prueba ──────────────────────────────────────
        _cls_odoo_ok = st.session_state.get("odoo_conectado", False)
        _cls_prueba  = st.session_state.get("modo_prueba", False)
        if _cls_odoo_ok:
            if _cls_prueba:
                st.info("🧪 **Modo prueba activado** — al confirmar se crearán productos con sufijo `_test` en categoría `PRUEBAS_AGENTE`")
            else:
                st.success("✅ **Odoo conectado** — al confirmar se crearán los productos en Odoo")
        else:
            st.warning("⚠️ **Odoo no conectado** — al confirmar solo se generará el reporte, no se crearán productos en Odoo")

        # ── Resumen rápido ────────────────────────────────────────────────────
        _acc_counts: dict[str, int] = {}
        for _p in _props:
            _acc_counts[_p["accion"]] = _acc_counts.get(_p["accion"], 0) + 1
        _sum_cols = st.columns(5)
        for _ci, (_acc, _lbl) in enumerate([
            ("crear_padre_y_variante", "🆕 Padre+variante"),
            ("crear_padre_solo",       "🆕 Solo padre"),
            ("crear_variante",         "➕ Variante"),
            ("reutilizar",             "♻️ Reutilizar"),
            ("duplicado",              "🔁 Duplicado"),
        ]):
            with _sum_cols[_ci]:
                st.metric(_lbl, _acc_counts.get(_acc, 0))

        # ── Fase 2: Enriquecer con Vision (títulos, descripciones, imágenes) ────
        _n_pendiente = sum(1 for _p in _props if _p.get("_vision_pendiente"))
        if _n_pendiente > 0:
            _v_offset    = st.session_state.get("vision_bloque_offset", 0)
            _v_blocksize = 50
            _n_total_v   = len(_props)
            _pct_done    = max(0, _v_offset) / _n_total_v if _n_total_v else 1.0
            with st.container(border=True):
                st.markdown(
                    f"**🔍 Análisis de imágenes pendiente** — {_n_pendiente} producto(s) sin título/descripción. "
                    "Pulsa el botón para enriquecer el siguiente bloque de 50 con Vision."
                )
                st.progress(_pct_done, text=f"{_v_offset}/{_n_total_v} productos analizados con Vision")
                _v_col1, _v_col2 = st.columns([2, 1])
                with _v_col1:
                    if st.button(
                        f"🔍 Analizar imágenes (bloque {_v_offset // _v_blocksize + 1})",
                        use_container_width=True,
                        key="btn_vision_bloque",
                    ):
                        with st.spinner(f"Analizando productos {_v_offset + 1}–{min(_v_offset + _v_blocksize, _n_total_v)} con Vision..."):
                            try:
                                enriquecer_vision_bloque(_v_offset, _v_blocksize)
                                st.session_state.vision_bloque_offset = _v_offset + _v_blocksize
                            except Exception as _ve:
                                st.error(f"Error en Vision: {_ve}")
                        st.rerun()
                with _v_col2:
                    if st.button("⏭️ Saltar Vision", use_container_width=True, key="btn_skip_vision"):
                        for _pp in _props:
                            _pp["_vision_pendiente"] = False
                        st.rerun()

        st.divider()

        # ── Encabezados de columnas ───────────────────────────────────────────
        _odoo_urls = st.session_state.get("odoo_imagen_urls", {})
        _hc = st.columns([1, 1, 4, 2, 1, 2])
        for _ht, _hcol in zip(
            ["🖼️ Producto", "🏷️ Padre Odoo", "Nombre / SKU / Acción", "Atributo", "⚠️ Rev.", "Nota bodega"],
            _hc,
        ):
            _hcol.markdown(f"**{_ht}**")
        st.divider()

        # ── Fila por producto ─────────────────────────────────────────────────
        _ACCION_COLOR = {
            "crear_padre_y_variante": "🟦",
            "crear_padre_solo":       "🟦",
            "crear_variante":         "🟩",
            "reutilizar":             "🟨",
            "duplicado":              "🟥",
        }
        for _i, _prop in enumerate(_props):
            _prod_p  = _prop["producto"]
            _fila    = str((_prod_p.get("fila_excel_0idx") or 0) + 1)
            _accion  = _prop["accion"]
            _badge   = _ACCION_COLOR.get(_accion, "⬜")
            _c_img, _c_padre, _c_info, _c_att, _c_rev, _c_nota = st.columns([1, 1, 4, 2, 1, 2])

            # ── Imagen del producto (del Excel) ───────────────────────────────
            with _c_img:
                _img_path = _prop.get("imagen_temp_path")
                if _img_path and Path(_img_path).exists():
                    try:
                        st.image(_img_path, width=80, caption=f"F{_fila}")
                    except Exception:
                        st.caption(f"F{_fila}")
                else:
                    st.caption(f"F{_fila}\n_(sin img)_")

            # ── Imagen del padre en Odoo ──────────────────────────────────────
            with _c_padre:
                _padre_url = None
                if _prop.get("padre_fuente") == "odoo" and _prop.get("padre_sku"):
                    _padre_url = _odoo_urls.get(_prop["padre_sku"])
                    # También intentar con el SKU base (sin atributo) en caso de variante
                    if not _padre_url:
                        _partes_sku = _prop["padre_sku"].split("-")
                        if len(_partes_sku) >= 2:
                            _sku_base = "-".join(_partes_sku[:2])
                            _padre_url = _odoo_urls.get(_sku_base)
                if _padre_url:
                    try:
                        st.image(_padre_url, width=80, caption=_prop.get("padre_odoo_nombre", "")[:20])
                    except Exception:
                        st.caption(_prop.get("padre_odoo_nombre", "—"))
                elif _prop.get("padre_fuente") == "local":
                    st.caption("📋 Local\n(mismo PL)")
                else:
                    st.caption("—")

            # ── Info del producto ─────────────────────────────────────────────
            with _c_info:
                st.text_input(
                    "Nombre",
                    value=_prop["nombre_base"],
                    key=f"_cls_nb_{_i}",
                    label_visibility="collapsed",
                )
                st.caption(f"`{_prop['sku']}` {_badge} {_prop['accion_display']}")
                if _prop.get("padre_sku") and _prop["padre_sku"] != _prop["sku"]:
                    st.caption(f"Padre: `{_prop['padre_sku']}`")

            # ── Atributo ──────────────────────────────────────────────────────
            with _c_att:
                if _prop.get("atributo_tipo") and _prop.get("atributo_valor"):
                    st.caption(f"{_prop['atributo_tipo']}  \n**{_prop['atributo_valor']}**")
                else:
                    st.caption("—")

            # ── Checkbox requiere revisión ────────────────────────────────────
            with _c_rev:
                st.checkbox(
                    "",
                    value=_prop.get("requiere_revision", False),
                    key=f"_cls_rev_{_i}",
                    label_visibility="collapsed",
                )

            # ── Nota ──────────────────────────────────────────────────────────
            with _c_nota:
                st.text_input(
                    "",
                    value=_prop.get("nota_revision", ""),
                    key=f"_cls_nota_{_i}",
                    placeholder="Observación...",
                    label_visibility="collapsed",
                )

            st.divider()

        # ── Botones de acción ─────────────────────────────────────────────────
        _btn_c1, _btn_c2, _btn_c3 = st.columns([3, 2, 1])
        with _btn_c1:
            _confirmar = st.button(
                "✅ Confirmar clasificación y continuar",
                type="primary",
                use_container_width=True,
            )
        with _btn_c2:
            _solo_excels = st.button(
                "📥 Solo generar Excels (sin crear en Odoo)",
                use_container_width=True,
                help="Genera todos los archivos Excel sin modificar Odoo. Útil cuando los productos ya están cargados.",
            )
        with _btn_c3:
            if st.session_state.get("clasificacion_reporte_bytes"):
                st.download_button(
                    "⬇️ Reporte bodega",
                    data=st.session_state.clasificacion_reporte_bytes,
                    file_name=f"{st.session_state.get('filename', 'packing').replace('.xlsx', '')}_CLASIFICACION.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        if _solo_excels:
            # Leer ediciones del usuario en los widgets sin tocar Odoo
            _nombres_editados = {}  # padre_sku -> nombre editado explícitamente por el usuario
            for _i, _prop in enumerate(_props):
                _prop["requiere_revision"] = bool(st.session_state.get(f"_cls_rev_{_i}", False))
                _prop["nota_revision"]     = str(st.session_state.get(f"_cls_nota_{_i}", "") or "")
                _nb_edit = str(st.session_state.get(f"_cls_nb_{_i}", "") or "").strip()
                if _nb_edit:
                    _prop["nombre_base"] = _nb_edit
                    _prop["nombre"]      = _nb_edit
                    if _prop.get("accion") != "duplicado" and _prop.get("padre_sku"):
                        _nombres_editados[_prop["padre_sku"]] = _nb_edit

            # Propagar nombre_base solo cuando el usuario lo editó explícitamente
            for _p in _props:
                if _p["accion"] == "duplicado" and _p.get("padre_sku") in _nombres_editados:
                    _p["nombre_base"] = _nombres_editados[_p["padre_sku"]]
                    _p["nombre"]      = _nombres_editados[_p["padre_sku"]]

            renombrar_imagenes_con_sku(_productos_c)

            # Rellenar descripcion/categoria/atributo — Vision tiene prioridad; texto como fallback
            for _prop in _props:
                _p  = _prop.get("producto")
                _dv = _prop.get("datos_vision", {})
                if _p is not None:
                    if not _p.get("descripcion"):
                        _vision_desc = (_dv.get("descripcion") or "").strip()
                        _nombre_ref  = (_prop.get("nombre_base") or _p.get("nombre", "")).strip()
                        if _vision_desc and _vision_desc.lower() != _nombre_ref.lower():
                            _p["descripcion"] = _vision_desc
                        else:
                            _dp = [_nombre_ref] if _nombre_ref else []
                            _at = _prop.get("atributo_tipo"); _av = _prop.get("atributo_valor")
                            if _at and _av:
                                _dp.append(f"{_at}: {_av}")
                            if _p.get("material"):
                                _dp.append(f"Material: {_p['material']}")
                            if _p.get("uso"):
                                _dp.append(_p["uso"])
                            _p["descripcion"] = " – ".join(_dp)
                    if not _p.get("categoria"):
                        _sub = _prop.get("subcategoria_cod") or _dv.get("subcategoria_cod") or "VAR"
                        _p["categoria"] = (_dv.get("categoria")
                                           or SUBCATEGORIAS.get(_sub, "Varios"))
                    if not _p.get("atributo"):
                        _p["atributo"] = (_dv.get("atributo_desc")
                                          or _prop.get("atributo_valor", "")
                                          or "Estándar")

            _fn_base = st.session_state.get("filename", "packing").replace(".xlsx", "")
            _tc_e    = st.session_state.get("tipo_cambio", 19.0)
            _cc_e    = st.session_state.get("costo_contenedor", 525000.0)

            with st.spinner("Generando archivos Excel..."):
                try:
                    st.session_state.excel_bytes        = generar_excel(_productos_c, _tc_e, _cc_e)
                except Exception as _ex:
                    st.warning(f"Excel FERRAFORME: {_ex}")
                try:
                    st.session_state.excel_master_bytes = generar_excel_master(_productos_c, _tc_e, _cc_e, _fn_base)
                except Exception as _ex:
                    st.warning(f"Excel Master: {_ex}")
                try:
                    st.session_state.clasificacion_reporte_bytes = generar_reporte_clasificacion(_props, _fn_base)
                except Exception as _ex:
                    st.warning(f"Reporte bodega: {_ex}")
                try:
                    st.session_state.purchase_bytes     = generar_excel_purchase(_productos_c, _tc_e, _cc_e)
                except Exception as _ex:
                    st.warning(f"Purchase: {_ex}")

            st.session_state.propuestas_para_reintento = _props
            st.session_state.clasificacion_activa      = False
            st.session_state.chat.append({
                "role": "assistant",
                "content": f"📥 Archivos generados para **{_n_props} productos** (sin crear en Odoo). Descárgalos desde el panel izquierdo."
            })
            st.rerun()

        if _confirmar:
            # Leer valores de los widgets individuales
            for _i, _prop in enumerate(_props):
                _prop["requiere_revision"] = bool(st.session_state.get(f"_cls_rev_{_i}", False))
                _prop["nota_revision"]     = str(st.session_state.get(f"_cls_nota_{_i}", "") or "")
                _nb_edit = str(st.session_state.get(f"_cls_nb_{_i}", "") or "").strip()
                if _nb_edit:
                    _prop["nombre_base"] = _nb_edit
                    _prop["nombre"]      = _nb_edit

            # Propagar nombre_base editado de padres a sus duplicados
            _nb_por_padre = {
                p["padre_sku"]: p["nombre_base"]
                for p in _props
                if p["accion"] != "duplicado" and p.get("padre_sku")
            }
            for _p in _props:
                if _p["accion"] == "duplicado" and _p.get("padre_sku") in _nb_por_padre:
                    _p["nombre_base"] = _nb_por_padre[_p["padre_sku"]]
                    _p["nombre"]      = _nb_por_padre[_p["padre_sku"]]

            # Rellenar descripcion/categoria/atributo — Vision tiene prioridad; texto como fallback
            for _prop in _props:
                _p  = _prop.get("producto")
                _dv = _prop.get("datos_vision", {})
                if _p is not None:
                    if not _p.get("descripcion"):
                        _vision_desc = (_dv.get("descripcion") or "").strip()
                        _nombre_ref  = (_prop.get("nombre_base") or _p.get("nombre", "")).strip()
                        if _vision_desc and _vision_desc.lower() != _nombre_ref.lower():
                            _p["descripcion"] = _vision_desc
                        else:
                            _dp = [_nombre_ref] if _nombre_ref else []
                            _at = _prop.get("atributo_tipo"); _av = _prop.get("atributo_valor")
                            if _at and _av:
                                _dp.append(f"{_at}: {_av}")
                            if _p.get("material"):
                                _dp.append(f"Material: {_p['material']}")
                            if _p.get("uso"):
                                _dp.append(_p["uso"])
                            _p["descripcion"] = " – ".join(_dp)
                    if not _p.get("categoria"):
                        _sub = _prop.get("subcategoria_cod") or _dv.get("subcategoria_cod") or "VAR"
                        _p["categoria"] = (_dv.get("categoria")
                                           or SUBCATEGORIAS.get(_sub, "Varios"))
                    if not _p.get("atributo"):
                        _p["atributo"] = (_dv.get("atributo_desc")
                                          or _prop.get("atributo_valor", "")
                                          or "Estándar")

            # Renombrar imágenes ANTES de subir a Odoo (pueden estar con stem original)
            _, _errs_ren_c = renombrar_imagenes_con_sku(_productos_c)
            _reportar_errores_imagenes(_errs_ren_c, "renombrado de imágenes")

            # ── Crear padres y variantes en Odoo ─────────────────────────────
            _odoo_ok = st.session_state.get("odoo_conectado", False)
            _res_odoo: list[str] = []
            _err_odoo: list[str] = []
            if _odoo_ok:
                # Recargar odoo_skus frescos antes de crear para evitar duplicados
                _skus_fresh2, _err_f2 = cargar_skus_odoo(
                    os.environ.get("ODOO_URL", ""), os.environ.get("ODOO_DB", ""),
                    os.environ.get("ODOO_USER", ""), os.environ.get("ODOO_PASSWORD", ""),
                )
                if not _err_f2 and _skus_fresh2:
                    st.session_state.odoo_skus = _skus_fresh2
                    sincronizar_contadores_con_odoo(_skus_fresh2)
                with st.spinner("Creando productos en Odoo..."):
                    _res_odoo, _err_odoo = crear_clasificacion_en_odoo(
                        _props, tipo_cambio,
                        st.session_state.get("costo_contenedor", 525000.0),
                    )
                # Agregar los SKUs recién creados a odoo_skus para que futuras
                # cargas en la misma sesión no generen números duplicados.
                _nuevos_skus_cls = [
                    p["sku"] for p in _props
                    if p.get("sku") and p.get("accion") != "duplicado"
                ]
                if _nuevos_skus_cls:
                    _skus_set = set(st.session_state.get("odoo_skus", []))
                    _skus_set.update(_nuevos_skus_cls)
                    st.session_state.odoo_skus = sorted(_skus_set)
                    sincronizar_contadores_con_odoo(st.session_state.odoo_skus)

            # ── Generar reporte de clasificación (para bodega) ────────────────
            _nombre_pl = st.session_state.get("filename", "")
            _reporte_bytes = generar_reporte_clasificacion(_props, _nombre_pl)
            st.session_state.clasificacion_reporte_bytes = _reporte_bytes

            # ── Mensaje resumen en el chat ────────────────────────────────────
            _n_rev = sum(1 for p in _props if p["requiere_revision"])
            _res_lines = [f"**Clasificación confirmada — {_n_props} producto(s)**"]
            _acc_display = {
                "crear_padre_y_variante": "🆕 Padre nuevo + variante nueva",
                "crear_padre_solo":       "🆕 Padre nuevo (sin variante)",
                "crear_variante":         "➕ Variante nueva en padre existente",
                "reutilizar":             "♻️ Reutilizado desde Odoo",
                "duplicado":              "🔁 Duplicado (se ignorará)",
            }
            for _acc, _cnt in sorted(_acc_counts.items()):
                if _cnt:
                    _res_lines.append(f"- {_acc_display.get(_acc, _acc)}: **{_cnt}**")
            if _n_rev:
                _res_lines.append(
                    f"\n⚠️ **{_n_rev} producto(s) marcados para revisión** — creados en Odoo con tag 'Requiere Revisión' e incluidos en el reporte.")
            if _res_odoo:
                _res_lines.append(f"\n**Odoo ({len(_res_odoo)} creados):**")
                _res_lines.extend(_res_odoo[:10])  # máx 10 líneas para no saturar el chat
                if len(_res_odoo) > 10:
                    _res_lines.append(f"  … y {len(_res_odoo) - 10} más.")
            if _err_odoo:
                _res_lines.append(f"\n**Errores Odoo ({len(_err_odoo)}):**")
                _res_lines.extend(_err_odoo[:5])
            elif not _odoo_ok:
                _res_lines.append("\n⚠️ Odoo no conectado — productos no creados en el sistema.")

            st.session_state.chat.append({"role": "assistant", "content": "\n".join(_res_lines)})

            # Guardar propuestas para reintento de SKUs fallidos en generar_excel_tool
            st.session_state.propuestas_para_reintento = _props

            # Limpiar estado de clasificación y arrancar chat
            st.session_state.clasificacion_activa = False

            with st.spinner("Iniciando el agente..."):
                _n_imgs_c = st.session_state.get("n_imgs_procesadas", 0)
                _iniciar_chat(_analisis_c, _productos_c, _advertencias_c,
                              tipo_cambio, contenedor, _n_imgs_c)
            st.rerun()


# ── PASO: Agregar productos en lote ──────────────────────────────────────────
if st.session_state.get("agregar_lote_activo"):
    with tab_pl:
        lote_paso = st.session_state.get("agregar_lote_paso", 1)

        # ════════ PASO 1 — Subir imágenes y nombres ═══════════════════════════════
        if lote_paso == 1:
            st.subheader("➕ Agregar productos nuevos")
            st.markdown("Fotografías de productos que no venían en el packing list original.")

            n_lote = st.number_input(
                "¿Cuántos productos deseas agregar?",
                min_value=1, max_value=30, value=st.session_state.get("agregar_lote_n", 1), step=1,
                key="lote_n_input",
            )

            imgs_subidas = st.file_uploader(
                f"Sube hasta {int(n_lote)} imagen(es) de los productos",
                type=["png", "jpg", "jpeg", "webp"],
                accept_multiple_files=True,
                key="lote_imgs_uploader",
            )

            # Leer bytes una sola vez antes de cualquier rerun
            imgs_data_preload = []
            if imgs_subidas:
                for f in imgs_subidas[: int(n_lote)]:
                    ext = Path(f.name).suffix.lstrip(".").lower() or "jpg"
                    imgs_data_preload.append({"data": f.read(), "ext": ext})

            # Previsualizar + nombres opcionales
            nombres_lote: list[str] = []
            if imgs_data_preload:
                st.markdown(f"**{len(imgs_data_preload)} imagen(es) cargadas:**")
                n_cols = min(len(imgs_data_preload), 4)
                preview_cols = st.columns(n_cols)
                for i, img_info in enumerate(imgs_data_preload):
                    with preview_cols[i % n_cols]:
                        st.image(img_info["data"], width=130)
                        nom = st.text_input(
                            f"Nombre {i + 1}",
                            placeholder=f"Producto {i + 1}",
                            key=f"lote_nombre_{i}",
                        )
                        nombres_lote.append(nom)

            col_proc, col_cancel = st.columns(2)
            with col_proc:
                if st.button(
                    "▶️ Procesar imágenes",
                    type="primary",
                    width="stretch",
                    disabled=not imgs_data_preload,
                ):
                    imgs_dict: dict = {}
                    for i, img_info in enumerate(imgs_data_preload):
                        nom = nombres_lote[i].strip() if i < len(nombres_lote) else ""
                        imgs_dict[i] = {"data": img_info["data"], "ext": img_info["ext"], "nombre": nom}
                    st.session_state.agregar_lote_imgs = imgs_dict
                    st.session_state.agregar_lote_n    = len(imgs_dict)
                    st.session_state.agregar_lote_paso = 2
                    st.rerun()
            with col_cancel:
                if st.button("✖️ Cancelar", width="stretch"):
                    st.session_state.agregar_lote_activo = False
                    st.rerun()

        # ════════ PASO 2 — Procesar con Vision + ODOO (automático) ════════════════
        elif lote_paso == 2:
            imgs_dict    = st.session_state.get("agregar_lote_imgs", {})
            skus_odoo    = st.session_state.get("odoo_skus", [])
            prods_odoo   = st.session_state.get("odoo_productos", [])
            phashes_odoo = st.session_state.get("odoo_phashes", {})

            with st.spinner(f"Analizando {len(imgs_dict)} imagen(es) con Vision y comparando con ODOO..."):
                prods_nuevos: list[dict]  = []
                conflictos_nuevos: list[dict] = []
                n_manuales_base = sum(1 for p in st.session_state.get("productos", []) if p.get("_manual"))

                for i, img_info in imgs_dict.items():
                    fila_0idx     = -(n_manuales_base + i + 1)
                    nombre_manual = img_info.get("nombre", "").strip() or f"Producto {i + 1}"

                    prod: dict = {
                        "nombre":          nombre_manual,
                        "fila_excel_0idx": fila_0idx,
                        "_manual":         True,
                    }

                    datos = analizar_imagen_claude(img_info["data"], img_info["ext"], contexto=prod)
                    sub_cod = datos.get("subcategoria_cod", "VAR")
                    att_cod = datos.get("atributo_cod", "GEN")
                    sku_inicial = generar_sku(sub_cod, att_cod)

                    prod.update({
                        "categoria":     datos.get("categoria", ""),
                        "subcategoria":  datos.get("subcategoria", ""),
                        "atributo_desc": datos.get("atributo_desc", ""),
                        "titulo":        datos.get("titulo", nombre_manual),
                        "descripcion":   datos.get("descripcion", ""),
                    })

                    conflicto_entry = None

                    if skus_odoo:
                        validacion = validar_sku_vs_odoo(sku_inicial, skus_odoo)
                        sku_final  = validacion["sku_ajustado"]
                        if validacion["conflicto"]:
                            conflicto_entry = {
                                "idx":           i,
                                "nombre":        nombre_manual,
                                "sku_propuesto": sku_inicial,
                                "sku_ajustado":  sku_final,
                                "datos_gemini":  datos,
                                "productos_odoo": [],
                                "razon":         "sku",
                                "similares_odoo": [],
                            }
                    else:
                        sku_final = sku_inicial

                    if prods_odoo:
                        similares = buscar_similares_odoo(
                            img_info["data"], nombre_manual, prods_odoo, phashes_odoo
                        )
                        if similares and conflicto_entry is None:
                            razones = []
                            if any(s.get("por_imagen") for s in similares):
                                razones.append("imagen similar")
                            if any(s.get("por_nombre") for s in similares):
                                razones.append("nombre similar")
                            conflicto_entry = {
                                "idx":           i,
                                "nombre":        nombre_manual,
                                "sku_propuesto": sku_inicial,
                                "sku_ajustado":  sku_inicial,
                                "datos_gemini":  datos,
                                "razon":         ", ".join(razones) or "imagen similar",
                                "similares_odoo": similares,
                                "productos_odoo": [s["producto_odoo"] for s in similares if s.get("producto_odoo")],
                            }
                        elif similares and conflicto_entry is not None:
                            # Enriquecer conflicto por SKU con similares encontrados
                            conflicto_entry["similares_odoo"] = similares
                            conflicto_entry["productos_odoo"] = [s["producto_odoo"] for s in similares if s.get("producto_odoo")]

                    prod["sku"] = sku_final
                    if st.session_state.get("modo_prueba"):
                        prod["sku"] += "_test"
                    prod["_img_data"] = img_info["data"]
                    prod["_img_ext"]  = img_info["ext"]

                    prods_nuevos.append(prod)
                    if conflicto_entry:
                        conflictos_nuevos.append(conflicto_entry)

            st.session_state.agregar_lote_prods = prods_nuevos
            st.session_state.agregar_lote_conf  = conflictos_nuevos
            st.session_state.agregar_lote_paso  = 3
            st.rerun()

        # ════════ PASO 3 — Revisar resultados y confirmar ═════════════════════════
        elif lote_paso == 3:
            prods_nuevos    = st.session_state.get("agregar_lote_prods", [])
            conflictos_lote = st.session_state.get("agregar_lote_conf", [])

            st.subheader(f"✅ {len(prods_nuevos)} producto(s) listos para agregar")

            # Preview de productos
            n_cols_prev = min(len(prods_nuevos), 4) or 1
            prev_cols   = st.columns(n_cols_prev)
            for i, prod in enumerate(prods_nuevos):
                with prev_cols[i % n_cols_prev]:
                    img_d = prod.get("_img_data")
                    if img_d:
                        st.image(img_d, width=130)
                    sku_badge = f"`{prod.get('sku', '—')}`"
                    st.markdown(f"**{prod.get('nombre', '—')}**  \nSKU: {sku_badge}")
                    if prod.get("categoria"):
                        st.caption(prod["categoria"])

            # Resolución de conflictos ODOO (si los hay)
            resoluciones_lote: dict = {}
            if conflictos_lote:
                st.divider()
                st.warning(f"⚠️ {len(conflictos_lote)} posible(s) coincidencia(s) con productos en ODOO — revisa antes de confirmar.")
                for conf in conflictos_lote:
                    c_idx      = conf["idx"]
                    c_nom      = conf["nombre"]
                    c_razon    = conf.get("razon", "")
                    c_similares = conf.get("similares_odoo", [])
                    c_sku_aj   = conf.get("sku_ajustado", conf.get("sku_propuesto", ""))

                    razon_badge = {
                        "sku":               "🔑 Mismo prefijo SKU",
                        "imagen similar":    "🖼️ Imagen similar",
                        "nombre similar":    "🔤 Nombre similar",
                        "semántica similar": "🧠 Semántica similar",
                    }.get(c_razon, f"⚠️ {c_razon}")

                    st.markdown(f"**{c_nom}** — {razon_badge}")

                    # Mostrar productos similares de ODOO
                    if c_similares:
                        sim_cols = st.columns(min(len(c_similares), 3))
                        for si, sim in enumerate(c_similares[:3]):
                            with sim_cols[si]:
                                _img_url = st.session_state.get("odoo_imagen_urls", {}).get(sim["sku"])
                                if _img_url:
                                    st.image(_img_url, width=100)
                                st.caption(
                                    f"SKU: `{sim['sku']}`  \n"
                                    f"{sim.get('nombre', '—')}"
                                )

                    dec_conf = st.radio(
                        "¿Qué hacemos con este producto?",
                        [
                            f"Asignar SKU nuevo `{c_sku_aj}` (es un producto diferente)",
                            "Reutilizar SKU de ODOO (es el mismo producto)",
                        ],
                        key=f"lote_conf_dec_{c_idx}",
                    )
                    if "Reutilizar" in dec_conf and c_similares:
                        sim_options = [s["sku"] for s in c_similares[:5]]
                        sku_sel = st.selectbox(
                            "SKU de ODOO a reutilizar:",
                            sim_options,
                            key=f"lote_conf_sku_{c_idx}",
                        )
                        resoluciones_lote[c_idx] = {"tipo": "mismo", "sku": sku_sel}
                    else:
                        resoluciones_lote[c_idx] = {"tipo": "diferente", "sku": c_sku_aj}

                    st.divider()

            col_conf, col_back = st.columns(2)
            with col_conf:
                if st.button("✅ Confirmar y agregar al Excel", type="primary", width="stretch"):
                    # Aplicar resoluciones de conflicto
                    for conf in conflictos_lote:
                        c_idx = conf["idx"]
                        res   = resoluciones_lote.get(c_idx, {"tipo": "diferente", "sku": conf.get("sku_ajustado", "")})
                        if res["tipo"] == "mismo":
                            prods_nuevos[c_idx]["sku"] = res["sku"]

                    # Agregar productos a la sesión + guardar imágenes temp
                    prods_agregados = []
                    for prod in prods_nuevos:
                        img_d = prod.pop("_img_data", None)
                        img_e = prod.pop("_img_ext", None)
                        if img_d and img_e:
                            try:
                                stem = f"manual_{abs(prod['fila_excel_0idx'])}"
                                IMAGENES_TEMP_PATH.mkdir(parents=True, exist_ok=True)
                                (IMAGENES_TEMP_PATH / f"{stem}.{img_e}").write_bytes(img_d)
                                prod["imagen_temp_stem"] = stem
                            except Exception:
                                pass
                        st.session_state.productos.append(prod)
                        prods_agregados.append(prod)

                    # Limpiar estado del lote
                    st.session_state.agregar_lote_activo = False
                    st.session_state.agregar_lote_paso   = 1
                    st.session_state.agregar_lote_imgs   = {}
                    st.session_state.agregar_lote_prods  = []
                    st.session_state.agregar_lote_conf   = []

                    # Notificar al agente
                    _msg_lote = (
                        f"Agregué {len(prods_agregados)} producto(s) que no venían en el packing list: "
                        + "; ".join(
                            f"'{p.get('nombre', '?')}' (SKU: `{p.get('sku', '?')}`)"
                            for p in prods_agregados
                        )
                        + ". Inclúyelos en el conteo y en el Excel cuando lo generes."
                    )
                    _sys_lote = build_system_prompt(
                        st.session_state.analisis,
                        st.session_state.productos,
                        st.session_state.get("tipo_cambio", 17.5),
                        st.session_state.get("contenedor_val", ""),
                        respuestas_dudas=st.session_state.respuestas_dudas or None,
                    )
                    st.session_state.chat.append({"role": "user", "content": _msg_lote})
                    with st.spinner("Notificando al agente..."):
                        _resp_lote = llamar_agente(_msg_lote, _sys_lote, st.session_state.lc_messages)
                    st.session_state.lc_messages.append(HumanMessage(content=_msg_lote))
                    st.session_state.lc_messages.append(AIMessage(content=_resp_lote))
                    st.session_state.chat.append({"role": "assistant", "content": _resp_lote})
                    st.rerun()

            with col_back:
                if st.button("↩️ Volver y cambiar imágenes", width="stretch"):
                    st.session_state.agregar_lote_paso  = 1
                    st.session_state.agregar_lote_prods = []
                    st.session_state.agregar_lote_conf  = []
                    st.rerun()



with tab_pl:
    # ── Chat input — siempre accesible cuando hay un archivo cargado ───────────────
    if st.session_state.analisis is not None:
            if st.session_state.esperando_dudas:
                placeholder = "Pregúntale al agente sobre estas dudas..."
            elif st.session_state.esperando_duplicados:
                placeholder = "Haz una aclaración o escribe tus cambios..."
            elif st.session_state.esperando_conflictos:
                placeholder = "Haz una aclaración o escribe tus cambios..."
            elif st.session_state.get("clasificacion_activa"):
                placeholder = "Haz una aclaración sobre la clasificación..."
            else:
                placeholder = "Escríbele al agente..."

            # ── Botones de respuesta rápida (solo en chat principal, no en pasos intermedios) ──
            _en_chat_principal = (
                not st.session_state.esperando_dudas
                and not st.session_state.esperando_duplicados
                and not st.session_state.esperando_conflictos
                and not st.session_state.get("clasificacion_activa")
                and not st.session_state.get("agregar_lote_activo")
            )
            if _en_chat_principal and st.session_state.chat:
                _ultimo_asist = next(
                    (m for m in reversed(st.session_state.chat) if m["role"] == "assistant"),
                    None,
                )
                _txt_ultimo = _ultimo_asist.get("content", "") if _ultimo_asist else ""
                _pide_confirm = isinstance(_txt_ultimo, str) and any(
                    s in _txt_ultimo.lower()
                    for s in ("proceder", "generar el excel", "generar los excel",
                              "¿puedo", "puedo proceder", "deseas que", "quieres que",
                              "procedo", "confirma", "podemos continuar")
                )
                if _pide_confirm:
                    _c1, _c2 = st.columns(2)
                    with _c1:
                        if st.button("✅ Sí, generar Excel", type="primary", width="stretch"):
                            st.session_state._quick_reply = "Sí, procede a generar el Excel."
                            st.rerun()
                    with _c2:
                        if st.button("✏️ Tengo cambios / preguntas", width="stretch"):
                            pass  # el usuario escribe en el chat

            # Consumir respuesta rápida de botón o esperar input del usuario
            _quick = st.session_state.get("_quick_reply")
            if _quick:
                st.session_state._quick_reply = None
            _chat_input = st.chat_input(placeholder)
            user_input = _quick or _chat_input

            if user_input:

                # ── Aclaración durante fase de dudas ──────────────────────────────────
                if st.session_state.esperando_dudas:
                    with st.chat_message("user"):
                        st.markdown(user_input)
                    st.session_state.chat_fase.append({"role": "user", "content": user_input})

                    with st.chat_message("assistant"):
                        with st.spinner("..."):
                            try:
                                _extra_dudas = f"DUDAS MOSTRADAS:\n{json.dumps(st.session_state.dudas_relevantes, ensure_ascii=False, indent=2)}"
                                system_aclaracion = _build_system_fase(
                                    "revisando las dudas encontradas en el packing list",
                                    st.session_state.productos,
                                    st.session_state.analisis,
                                    extra=_extra_dudas,
                                )
                                lc_aclaracion = [
                                    HumanMessage(content=m["content"]) if m["role"] == "user"
                                    else AIMessage(content=m["content"])
                                    for m in st.session_state.chat_fase
                                ]
                                llm_ac = ChatAnthropic(model="claude-sonnet-4-6", max_tokens=1000)
                                resp   = llm_ac.invoke([SystemMessage(content=system_aclaracion)] + lc_aclaracion)
                                respuesta = resp.content.strip()
                                st.markdown(respuesta)
                                st.session_state.chat_fase.append({"role": "assistant", "content": respuesta})
                            except Exception as e:
                                st.error(f"Error: {e}")
                    st.rerun()

                # ── Chat durante clasificación padre/variante ─────────────────────────
                elif st.session_state.get("clasificacion_activa"):
                    with st.chat_message("user"):
                        st.markdown(user_input)
                    st.session_state.chat_fase.append({"role": "user", "content": user_input})
                    with st.chat_message("assistant"):
                        with st.spinner("..."):
                            try:
                                system_clas = _build_system_fase(
                                    "revisando la clasificación de productos (padres, variantes, duplicados)",
                                    st.session_state.productos,
                                    st.session_state.analisis,
                                )
                                lc_clas = [
                                    HumanMessage(content=m["content"]) if m["role"] == "user"
                                    else AIMessage(content=m["content"])
                                    for m in st.session_state.chat_fase
                                    if isinstance(m.get("content"), str)
                                ]
                                llm_clas = ChatAnthropic(model="claude-sonnet-4-6", max_tokens=800)
                                resp_clas = llm_clas.invoke([SystemMessage(content=system_clas)] + lc_clas)
                                respuesta_clas = resp_clas.content.strip()
                                st.markdown(respuesta_clas)
                                st.session_state.chat_fase.append({"role": "assistant", "content": respuesta_clas})
                            except Exception as e:
                                st.error(f"Error: {e}")
                    st.rerun()

                # ── Chat durante duplicados o conflictos: aclaración contextual ────────
                elif st.session_state.esperando_duplicados or st.session_state.esperando_conflictos:
                    with st.chat_message("user"):
                        st.markdown(user_input)
                    st.session_state.chat_fase.append({"role": "user", "content": user_input})

                    with st.chat_message("assistant"):
                        with st.spinner("..."):
                            try:
                                _nombre_fase = (
                                    "revisando posibles productos duplicados o variantes"
                                    if st.session_state.esperando_duplicados
                                    else "revisando conflictos de SKU con ODOO"
                                )
                                system_fase = _build_system_fase(
                                    _nombre_fase,
                                    st.session_state.productos,
                                    st.session_state.analisis,
                                )
                                lc_fase = [
                                    HumanMessage(content=m["content"]) if m["role"] == "user"
                                    else AIMessage(content=m["content"])
                                    for m in st.session_state.chat_fase
                                    if isinstance(m.get("content"), str)
                                ]
                                llm_fase = ChatAnthropic(model="claude-sonnet-4-6", max_tokens=800)
                                resp_fase = llm_fase.invoke([SystemMessage(content=system_fase)] + lc_fase)
                                respuesta_fase = resp_fase.content.strip()
                                st.markdown(respuesta_fase)
                                st.session_state.chat_fase.append({"role": "assistant", "content": respuesta_fase})
                            except Exception as e:
                                st.error(f"Error: {e}")
                    st.rerun()

                # ── Chat principal ─────────────────────────────────────────────────────
                else:
                    if not st.session_state.chat:
                        st.info("Sube un packing list en el panel izquierdo para comenzar.")

                    with st.chat_message("user"):
                        st.markdown(user_input)
                    st.session_state.chat.append({"role": "user", "content": user_input})

                    with st.chat_message("assistant"):
                        with st.spinner("El agente está procesando..."):
                            try:
                                system = build_system_prompt(
                                    st.session_state.analisis,
                                    st.session_state.productos,
                                    tipo_cambio,
                                    contenedor,
                                    respuestas_dudas=st.session_state.respuestas_dudas or None,
                                )
                                excel_antes = st.session_state.excel_bytes
                                texto = llamar_agente(
                                    user_input,
                                    system,
                                    st.session_state.lc_messages,
                                )
                                # Guardar en historial LangChain
                                st.session_state.lc_messages.append(HumanMessage(content=user_input))
                                st.session_state.lc_messages.append(AIMessage(content=texto))
                                st.session_state.chat.append({"role": "assistant", "content": texto})

                                st.markdown(texto)
                                if st.session_state.excel_bytes != excel_antes:
                                    st.success("Excel actualizado — usa los botones de descarga de abajo.")

                            except Exception as e:
                                err = f"Error: {e}"
                                st.session_state.chat.append({"role": "assistant", "content": err})
                                st.error(err)

                    st.rerun()

    # ── Pantalla inicial ───────────────────────────────────────────────────────────
    elif not st.session_state.chat:
            st.info("Sube un packing list en el panel izquierdo para comenzar.")




# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — AGREGAR PRODUCTOS A EXCEL FERRAFORME
# ═══════════════════════════════════════════════════════════════════════════════

def _agregar_productos_a_excel_ferraforme(excel_bytes: bytes,
                                           prods_nuevos: list[dict]) -> bytes:
    """
    Carga el Excel FERRAFORME existente desde un archivo temporal (para preservar
    las imágenes originales exactamente igual) y agrega los nuevos productos al
    final — sin tocar las filas ni imágenes existentes.

    Columnas FERRAFORME (después del delete_cols que hace generar_excel):
      1:nombre  2:sku  3:imagen  4:cbm_por_pieza  5:precio_usd  6:None
      7:piezas_x_caja  8:cbm_caja  9:tipo_producto  10:EMPRESA  11:None
      12:contenedor  13:costo_landed_usd  14:descripcion  15:categoria
      16:atributo  17-20:None
    """
    IMG_W_PX = 120
    IMG_H_PX = 120
    ROW_H_PT = 90

    # ── 1. Abrir el Excel existente ───────────────────────────────────────────
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    # ── 2. Encontrar la última fila con datos (scan de arriba hacia abajo) ──────
    #    max_row no es confiable cuando el Excel viene del template (reporta 2075
    #    por celdas fantasma aunque los datos reales terminen mucho antes).
    #    Scan top-down: avanzar fila a fila y parar en la primera fila donde
    #    tanto col 1 (nombre) como col 2 (sku) estén vacías.
    _last_data_row = 2  # mínimo: fila de encabezados
    for _r in range(3, ws.max_row + 2):
        if ws.cell(_r, 1).value or ws.cell(_r, 2).value:
            _last_data_row = _r
        else:
            break  # primera fila vacía → aquí terminan los datos

    # Inferir contenedor de los existentes
    _contenedor_ref = ""
    for _r in range(3, _last_data_row + 1):
        _cv = ws.cell(_r, 12).value
        if _cv:
            _contenedor_ref = str(_cv)
            break

    # ── 3. Agregar filas de los productos nuevos al final ─────────────────────
    for _i, _np in enumerate(prods_nuevos):
        _r   = _last_data_row + 1 + _i
        _sku = _np.get("sku", "")
        _nom = _np.get("titulo") or _np.get("nombre", "")

        ws.cell(_r, 1).value = _nom
        ws.cell(_r, 2).value = _sku

        # Imagen del producto nuevo
        _img_incrustada = False
        _img_d = _np.get("_img_data")
        if _img_d and _PILLOW_OK:
            try:
                _pil = PILImage.open(io.BytesIO(_img_d)).convert("RGB")
                _pil.thumbnail((IMG_W_PX, IMG_H_PX), PILImage.LANCZOS)
                _buf_i = io.BytesIO()
                _pil.save(_buf_i, format="PNG")
                _buf_i.seek(0)
                _xl_img = XLImage(_buf_i)
                _xl_img.width  = IMG_W_PX
                _xl_img.height = IMG_H_PX
                _xl_img.anchor = f"C{_r}"
                ws.add_image(_xl_img)
                ws.row_dimensions[_r].height = ROW_H_PT
                _img_incrustada = True
            except Exception:
                pass
        if not _img_incrustada:
            ws.cell(_r, 3).value = ""

        _cu = _safe_float(_np.get("cbm_por_pieza"))
        _cbm_cell = ws.cell(_r, 4)
        _cbm_cell.value = _cu if _cu > 0 else None
        if _cbm_cell.value is not None:
            _cbm_cell.number_format = "0.000000"

        ws.cell(_r, 5).value = _np.get("precio_usd")
        ws.cell(_r, 6).value = _np.get("piezas_total")    # col "Unidades"

        _px = _np.get("piezas_x_caja")
        ws.cell(_r, 7).value = _px
        _cbm_caja = ws.cell(_r, 8)
        _cbm_caja.value = round(float(_px) * _cu, 6) if _px and _cu > 0 else None
        if _cbm_caja.value is not None:
            _cbm_caja.number_format = "0.000000"

        ws.cell(_r, 9).value  = "Producto almacenable"
        ws.cell(_r, 10).value = EMPRESA
        ws.cell(_r, 11).value = None
        ws.cell(_r, 12).value = _contenedor_ref
        ws.cell(_r, 13).value = float(_np["precio_usd"]) if _np.get("precio_usd") else None
        ws.cell(_r, 14).value = _np.get("descripcion")
        ws.cell(_r, 15).value = _np.get("categoria")
        ws.cell(_r, 16).value = _np.get("atributo_desc") or _np.get("atributo")
        for _col in range(17, 21):
            ws.cell(_r, _col).value = None

    # ── 4. Guardar a BytesIO y devolver ───────────────────────────────────────
    _buf_out = io.BytesIO()
    wb.save(_buf_out)
    _buf_out.seek(0)
    return _buf_out.read(), len(prods_nuevos)


_solo_defaults = {
    "agregar_solo_paso":         1,
    "agregar_solo_imgs":         {},
    "agregar_solo_prods":        [],
    "agregar_solo_conf":         [],
    "agregar_solo_n":            1,
    "agregar_solo_excel_bytes":  None,
    "agregar_solo_excel_nombre": "",
    "agregar_solo_excel_result": None,   # bytes del Excel ya actualizado
    "agregar_solo_ultimo_error": None,   # error persistente para mostrar tras rerun
}
for _k, _v in _solo_defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

with tab_agregar:
    solo_paso = st.session_state.get("agregar_solo_paso", 1)

    # ════════ PASO 1 — Subir Excel base + imágenes ════════════════════════════
    if solo_paso == 1:
        st.subheader("Agregar productos a un Excel FERRAFORME")
        st.markdown(
            "Carga el Excel que ya tienes y las fotografías de los productos nuevos. "
            "El sistema generará el SKU, nombre, descripción y atributos a partir de "
            "la imagen, los comparará con Odoo y los agregará al Excel."
        )
        st.divider()

        # ── Excel FERRAFORME ──────────────────────────────────────────────────
        st.markdown("**1. Excel FERRAFORME al que quieres agregar productos**")
        solo_excel_subido = st.file_uploader(
            "Sube el Excel (.xlsx)",
            type=["xlsx", "xls"],
            key="solo_excel_uploader",
        )
        if solo_excel_subido:
            st.success(f"✅ **{solo_excel_subido.name}** cargado")

        st.divider()

        # ── Imágenes de los nuevos productos ─────────────────────────────────
        st.markdown("**2. Fotografías de los productos a agregar**")
        solo_n = st.number_input(
            "Cantidad de productos",
            min_value=1, max_value=30,
            value=st.session_state.get("agregar_solo_n", 1), step=1,
            key="solo_n_input",
        )
        solo_imgs_subidas = st.file_uploader(
            f"Sube hasta {int(solo_n)} imagen(es)",
            type=["png", "jpg", "jpeg", "webp"],
            accept_multiple_files=True,
            key="solo_imgs_uploader",
        )

        solo_imgs_data: list[dict] = []
        if solo_imgs_subidas:
            for _f in solo_imgs_subidas[: int(solo_n)]:
                _ext = Path(_f.name).suffix.lstrip(".").lower() or "jpg"
                solo_imgs_data.append({"data": _f.read(), "ext": _ext})

        solo_nombres: list[str] = []
        if solo_imgs_data:
            st.markdown(f"**{len(solo_imgs_data)} imagen(es) cargadas:**")
            _nc = min(len(solo_imgs_data), 4)
            _pc = st.columns(_nc)
            for _i, _img in enumerate(solo_imgs_data):
                with _pc[_i % _nc]:
                    st.image(_img["data"], width=130)
                    _nom = st.text_input(
                        f"Nombre {_i + 1} (opcional)",
                        placeholder=f"Producto {_i + 1}",
                        key=f"solo_nombre_{_i}",
                    )
                    solo_nombres.append(_nom)

        _puede_procesar = solo_imgs_data and solo_excel_subido
        if not solo_excel_subido and solo_imgs_data:
            st.warning("Sube el Excel FERRAFORME para continuar.")

        if st.button(
            "▶️ Analizar y generar SKUs",
            type="primary",
            width="stretch",
            disabled=not _puede_procesar,
            key="solo_procesar_btn",
        ):
            _imgs_dict: dict = {}
            for _i, _img in enumerate(solo_imgs_data):
                _nom = solo_nombres[_i].strip() if _i < len(solo_nombres) else ""
                _imgs_dict[_i] = {"data": _img["data"], "ext": _img["ext"], "nombre": _nom}
            st.session_state.agregar_solo_imgs         = _imgs_dict
            st.session_state.agregar_solo_n            = len(_imgs_dict)
            st.session_state.agregar_solo_excel_bytes  = solo_excel_subido.read()
            st.session_state.agregar_solo_excel_nombre = solo_excel_subido.name
            st.session_state.agregar_solo_excel_result = None
            st.session_state.agregar_solo_paso         = 2
            st.rerun()

    # ════════ PASO 2 — Vision + ODOO (automático) ════════════════════════════
    elif solo_paso == 2:
        _solo_imgs_dict    = st.session_state.get("agregar_solo_imgs", {})
        _solo_skus_odoo    = st.session_state.get("odoo_skus", [])
        _solo_prods_odoo   = st.session_state.get("odoo_productos", [])
        _solo_phashes_odoo = st.session_state.get("odoo_phashes", {})

        with st.spinner(f"Analizando {len(_solo_imgs_dict)} imagen(es) con Vision..."):
            _solo_prods_nuevos: list[dict] = []
            _solo_conflictos:   list[dict] = []

            for _i, _img_info in _solo_imgs_dict.items():
                _nombre_manual = _img_info.get("nombre", "").strip()
                # Solo pasar nombre como contexto si el usuario lo escribió;
                # si no, dejar que Vision lo genere libre desde la imagen.
                _ctx = {"fila_excel_0idx": -(_i + 1), "_manual": True}
                if _nombre_manual:
                    _ctx["nombre"] = _nombre_manual

                _datos   = analizar_imagen_claude(_img_info["data"], _img_info["ext"], contexto=_ctx)
                _sub_cod = _datos.get("subcategoria_cod", "VAR")
                _att_cod = _datos.get("atributo_cod", "GEN")
                _sku_ini = generar_sku(_sub_cod, _att_cod)

                # El nombre final es: lo que Vision generó (titulo), o lo que escribió el usuario
                _titulo_final = _datos.get("titulo") or _nombre_manual or f"Producto {_i + 1}"
                _prod: dict = {
                    "nombre":        _titulo_final,
                    "titulo":        _titulo_final,
                    "categoria":     _datos.get("categoria", ""),
                    "subcategoria":  _datos.get("subcategoria", ""),
                    "atributo_desc": _datos.get("atributo_desc", ""),
                    "atributo":      _datos.get("atributo_desc", ""),
                    "descripcion":   _datos.get("descripcion", ""),
                    "fila_excel_0idx": -(_i + 1),
                    "_manual": True,
                }

                _conf_entry = None
                if _solo_skus_odoo:
                    _val = validar_sku_vs_odoo(_sku_ini, _solo_skus_odoo)
                    _sku_final = _val["sku_ajustado"]
                    if _val["conflicto"]:
                        _conf_entry = {
                            "idx": _i, "nombre": _nombre_manual,
                            "sku_propuesto": _sku_ini, "sku_ajustado": _sku_final,
                            "datos_gemini": _datos, "productos_odoo": [],
                            "razon": "sku", "similares_odoo": [],
                        }
                else:
                    _sku_final = _sku_ini

                if _solo_prods_odoo:
                    _sims = buscar_similares_odoo(
                        _img_info["data"], _nombre_manual, _solo_prods_odoo, _solo_phashes_odoo
                    )
                else:
                    _sims = []

                if _sims and _conf_entry is None:
                    _razones = []
                    if any(s.get("por_imagen") for s in _sims): _razones.append("imagen similar")
                    if any(s.get("por_nombre") for s in _sims): _razones.append("nombre similar")
                    _conf_entry = {
                        "idx": _i, "nombre": _nombre_manual,
                        "sku_propuesto": _sku_ini, "sku_ajustado": _sku_ini,
                        "datos_gemini": _datos,
                        "razon": ", ".join(_razones) or "similitud detectada",
                        "similares_odoo": _sims,
                        "productos_odoo": [s["producto_odoo"] for s in _sims if s.get("producto_odoo")],
                    }
                elif _sims and _conf_entry is not None:
                    _conf_entry["similares_odoo"] = _sims
                    _conf_entry["productos_odoo"] = [s["producto_odoo"] for s in _sims if s.get("producto_odoo")]

                _prod["sku"]       = _sku_final
                _prod["_img_data"] = _img_info["data"]
                _prod["_img_ext"]  = _img_info["ext"]

                _solo_prods_nuevos.append(_prod)
                if _conf_entry:
                    _solo_conflictos.append(_conf_entry)

        st.session_state.agregar_solo_prods = _solo_prods_nuevos
        st.session_state.agregar_solo_conf  = _solo_conflictos
        st.session_state.agregar_solo_paso  = 3
        st.rerun()

    # ════════ PASO 3 — Revisar, resolver conflictos y descargar Excel ════════
    elif solo_paso == 3:
        _solo_prods_nuevos    = st.session_state.get("agregar_solo_prods", [])
        _solo_conflictos_lote = st.session_state.get("agregar_solo_conf", [])
        _solo_excel_nombre    = st.session_state.get("agregar_solo_excel_nombre", "archivo.xlsx")
        _solo_excel_result    = st.session_state.get("agregar_solo_excel_result")
        _solo_excel_bytes_ok  = bool(st.session_state.get("agregar_solo_excel_bytes"))

        # Mostrar error persistente si lo hubo en el intento anterior
        if st.session_state.get("agregar_solo_ultimo_error"):
            st.error(st.session_state.agregar_solo_ultimo_error)

        _n_esperados = st.session_state.get("agregar_solo_n", len(_solo_prods_nuevos))
        _n_reales    = len(_solo_prods_nuevos)

        st.subheader(f"✅ {_n_reales} producto(s) listos para agregar")
        if not _solo_excel_bytes_ok:
            st.error("⚠️ No se encontraron los bytes del Excel. Vuelve al Paso 1 y sube el archivo de nuevo.")
        else:
            st.caption(f"Se agregarán al Excel: **{_solo_excel_nombre}**")

        # Aviso si la cantidad analizada no cuadra con lo que el usuario subió
        if _n_reales != _n_esperados:
            st.warning(
                f"⚠️ Indicaste **{_n_esperados}** producto(s) pero solo se analizaron **{_n_reales}**. "
                "Verifica que todas las imágenes se subieron correctamente o vuelve al Paso 1."
            )

        # ── Vista previa de productos ─────────────────────────────────────────
        _nc2 = min(len(_solo_prods_nuevos), 4) or 1
        _pc2 = st.columns(_nc2)
        for _i, _prod in enumerate(_solo_prods_nuevos):
            with _pc2[_i % _nc2]:
                if _prod.get("_img_data"):
                    st.image(_prod["_img_data"], width=130)
                st.markdown(
                    f"**{_prod.get('titulo') or _prod.get('nombre', '—')}**  \n"
                    f"SKU: `{_prod.get('sku', '—')}`"
                )
                if _prod.get("categoria"):
                    st.caption(f"📂 {_prod['categoria']}")
                if _prod.get("atributo_desc"):
                    st.caption(f"🏷️ {_prod['atributo_desc']}")
                if _prod.get("descripcion"):
                    st.caption(f"📝 {_prod['descripcion']}")

        # ── Resolución de conflictos ODOO ─────────────────────────────────────
        _solo_resoluciones: dict = {}
        if _solo_conflictos_lote:
            st.divider()
            st.warning(
                f"⚠️ {len(_solo_conflictos_lote)} posible(s) coincidencia(s) con productos en Odoo — "
                "revisa antes de continuar."
            )
            for _conf in _solo_conflictos_lote:
                _c_idx      = _conf["idx"]
                _c_nom      = _conf["nombre"]
                _c_razon    = _conf.get("razon", "")
                _c_similares = _conf.get("similares_odoo", [])
                _c_sku_aj   = _conf.get("sku_ajustado", _conf.get("sku_propuesto", ""))

                _razon_badge = {
                    "sku":            "🔑 Mismo prefijo SKU",
                    "imagen similar": "🖼️ Imagen similar",
                    "nombre similar": "🔤 Nombre similar",
                }.get(_c_razon, f"⚠️ {_c_razon}")

                st.markdown(f"**{_c_nom}** — {_razon_badge}")
                if _c_similares:
                    _sc = st.columns(min(len(_c_similares), 3))
                    for _si, _sim in enumerate(_c_similares[:3]):
                        with _sc[_si]:
                            _img_url = st.session_state.get("odoo_imagen_urls", {}).get(_sim["sku"])
                            if _img_url:
                                st.image(_img_url, width=100)
                            st.caption(f"SKU: `{_sim['sku']}`  \n{_sim.get('nombre', '—')}")

                _dec_conf = st.radio(
                    "¿Qué hacemos?",
                    [
                        f"SKU nuevo `{_c_sku_aj}` (es un producto diferente)",
                        "Reutilizar SKU de Odoo (ya existe)",
                    ],
                    key=f"solo_conf_dec_{_c_idx}",
                )
                if "Reutilizar" in _dec_conf and _c_similares:
                    _sim_opts = [s["sku"] for s in _c_similares[:5]]
                    _sku_sel = st.selectbox(
                        "SKU de Odoo a reutilizar:",
                        _sim_opts,
                        key=f"solo_conf_sku_{_c_idx}",
                    )
                    _solo_resoluciones[_c_idx] = {"tipo": "mismo", "sku": _sku_sel}
                else:
                    _solo_resoluciones[_c_idx] = {"tipo": "diferente", "sku": _c_sku_aj}
                st.divider()

        # ── Botones ───────────────────────────────────────────────────────────
        _col_gen, _col_back2 = st.columns(2)

        with _col_gen:
            if st.button("📥 Generar Excel actualizado", type="primary",
                         width="stretch", key="solo_generar_btn",
                         disabled=not _solo_excel_bytes_ok):
                # Aplicar resoluciones de SKU
                for _conf in _solo_conflictos_lote:
                    _c_idx = _conf["idx"]
                    _res   = _solo_resoluciones.get(
                        _c_idx, {"tipo": "diferente", "sku": _conf.get("sku_ajustado", "")}
                    )
                    if _res["tipo"] == "mismo":
                        _solo_prods_nuevos[_c_idx]["sku"] = _res["sku"]

                st.session_state.agregar_solo_ultimo_error = None
                with st.spinner("Agregando productos al Excel..."):
                    try:
                        _result_bytes, _n_agregados = _agregar_productos_a_excel_ferraforme(
                            st.session_state.agregar_solo_excel_bytes,
                            _solo_prods_nuevos,
                        )
                        # Avisar si se agregaron menos filas de las esperadas
                        if _n_agregados != _n_esperados:
                            st.session_state.agregar_solo_ultimo_error = (
                                f"⚠️ Se agregaron {_n_agregados} fila(s) pero se esperaban "
                                f"{_n_esperados}. Verifica el resultado antes de usar el Excel."
                            )
                        st.session_state.agregar_solo_excel_result  = _result_bytes
                        st.session_state.agregar_solo_prods         = _solo_prods_nuevos
                        st.session_state.agregar_solo_n_agregados   = _n_agregados
                        st.rerun()
                    except Exception as _ex:
                        import traceback
                        _tb = traceback.format_exc()
                        st.session_state.agregar_solo_ultimo_error = (
                            f"Error al actualizar el Excel: {_ex}\n\nDetalle:\n{_tb}"
                        )
                        st.rerun()

        with _col_back2:
            if st.button("↩️ Volver y cambiar imágenes", width="stretch", key="solo_back_btn"):
                st.session_state.agregar_solo_paso  = 1
                st.session_state.agregar_solo_prods = []
                st.session_state.agregar_solo_conf  = []
                st.session_state.agregar_solo_excel_result = None
                st.rerun()

        # ── Botón de descarga (aparece tras generar) ──────────────────────────
        if _solo_excel_result:
            _base_nom, _ = (lambda p: (p.rsplit(".", 1)[0], p.rsplit(".", 1)[1]) if "." in p else (p, ""))(_solo_excel_nombre)
            _nombre_out = _base_nom + "_actualizado.xlsx"
            _n_agr = st.session_state.get("agregar_solo_n_agregados", len(_solo_prods_nuevos))
            if _n_agr == _n_esperados:
                st.success(f"✅ Excel listo — se agregaron **{_n_agr}** fila(s) correctamente.")
            else:
                st.warning(f"⚠️ Excel generado, pero se agregaron **{_n_agr}** de **{_n_esperados}** fila(s) esperadas.")
            st.download_button(
                label="⬇️ Descargar Excel actualizado",
                data=_solo_excel_result,
                file_name=_nombre_out,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="solo_download_btn",
                use_container_width=True,
            )
            if st.button("➕ Agregar más productos", width="stretch", key="solo_reset_btn"):
                # Mantener el Excel resultado como base para la siguiente ronda
                st.session_state.agregar_solo_excel_bytes  = _solo_excel_result
                st.session_state.agregar_solo_paso         = 1
                st.session_state.agregar_solo_imgs         = {}
                st.session_state.agregar_solo_prods        = []
                st.session_state.agregar_solo_conf         = []
                st.session_state.agregar_solo_excel_result = None
                st.rerun()

# ── Historial de conversación (visible desde ambas tabs) ──────────────────────
with tab_agregar:
    if st.session_state.chat:
        with st.expander(
            f"🕓 Historial del agente ({len(st.session_state.chat)} mensajes)",
            expanded=False,
        ):
            for _hmsg in st.session_state.chat:
                with st.chat_message(_hmsg["role"]):
                    _hcontent = _hmsg["content"]
                    if isinstance(_hcontent, str):
                        st.markdown(_hcontent)
                    else:
                        for _hbloque in _hcontent:
                            if not isinstance(_hbloque, dict):
                                st.markdown(str(_hbloque))
                            elif _hbloque.get("type") == "text":
                                st.markdown(_hbloque["value"])
                            elif _hbloque.get("type") == "image_bytes":
                                try:
                                    st.image(_hbloque["data"], width=_hbloque.get("width", 160),
                                             caption=_hbloque.get("caption", ""))
                                except Exception:
                                    pass
