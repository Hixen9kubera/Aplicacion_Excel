"""
agregar_variantes_directo.py
Para cada SKU del ferraforme que no exista en Odoo como product.product,
crea el product.product usando el product.template.attribute.value correcto.
"""
import sys, re
import xmlrpc.client
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ODOO_URL  = "https://ifull.odoo.com"
ODOO_DB   = "ifullmx-brea-main-6396587"
ODOO_USER = "valeria.rivero@kubera.mx"
ODOO_PASS = "Vale:kubera28"

common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)
print(f"Conectado uid={uid}")

BASE_DIR = Path(r"c:\Users\valer\OneDrive\Documentos\Desarrollo_Agentes_IA\skus listos")

CONFIGS = {
    "71": dict(file="OOCU8248653 PL contenedor 71.xlsx",      container="OOCU8248653 cont 71",  col_sku=3,  col_atributo=6,  col_ingles=7),
    "72": dict(file="CAAU5061672 PL contenedor 72.xlsx",      container="CAAU5061672 cont 72",  col_sku=3,  col_atributo=6,  col_ingles=7),
    "73": dict(file="CSNU6409280 PL contenedor 73.xlsx",      container="CSNU6409280 cont 73",  col_sku=2,  col_atributo=5,  col_ingles=6),
    "74": dict(file="LEX25-510_WHSU6230286_PL contenedor 74.xlsx", container="WHSU6230286 cont 74", col_sku=2, col_atributo=6, col_ingles=7),
    "75": dict(file="PRY25-543_GAOU7102197_PL contenedor 75.xlsx", container="GAOU7102197 cont 75", col_sku=2, col_atributo=5, col_ingles=7),
    "76": dict(file="LEX25-429_CSNU6006475_PL contenedor 76.xlsx", container="CSNU6006475 cont 76", col_sku=3, col_atributo=6, col_ingles=7),
    "77": dict(file="TONY25-275 TRHU6472913 Lista de empaque contenedor 77.xlsx", container="TRHU6472913 cont 77", col_sku=3, col_atributo=6, col_ingles=7),
    "78": dict(file="TONY25-284 TLLU5922910 Lista de empaque contenedor 78.xlsx", container="TLLU5922910 cont 78", col_sku=3, col_atributo=7, col_ingles=6),
}

only = sys.argv[1] if len(sys.argv) > 1 else None


def _parse_parent(sku):
    parts = str(sku).strip().split("-")
    if len(parts) >= 3:
        return f"{parts[0]}-{parts[1]}", "-".join(parts[2:])
    return sku, None


_sku_conflict_log = []

def _check_sku_libre(sku, mi_container):
    """Devuelve True si el SKU está libre para asignar.
    Si ya existe en otro contenedor activo, lo registra y devuelve False.
    """
    existentes = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.product", "search_read",
        [[["default_code", "=", sku], ["active", "=", True]]],
        {"fields": ["id", "product_tmpl_id"]})
    if not existentes:
        return True
    for pp in existentes:
        tmpl = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.template", "read",
            [[pp["product_tmpl_id"][0]]], {"fields": ["container_numbers", "name"]})[0]
        otro_cont = tmpl.get("container_numbers") or "sin contenedor"
        if otro_cont != mi_container:
            msg = f"  ⚠ CONFLICTO SKU {sku}: ya existe en '{otro_cont}' ({tmpl['name']}) pp_id={pp['id']} — omitido"
            print(msg)
            _sku_conflict_log.append(msg)
            return False
    return False  # ya existe en este mismo contenedor


total_creados = total_ya_existe = total_errores = 0

for num, cfg in CONFIGS.items():
    if only and num != only:
        continue

    print(f"\n{'='*60}")
    print(f"Cont {num}: {cfg['container']}")

    wb = openpyxl.load_workbook(str(BASE_DIR / cfg["file"]), data_only=True)
    ws = wb.active

    # Leer SKUs del ferraforme agrupados por padre
    padres = {}  # parent_sku → [{"sku": ..., "att_cod": ...}]
    seen = set()
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, cfg["col_sku"]).value
        if not cell:
            continue
        sku = str(cell).strip()
        if not re.match(r'^[A-Z]{2,6}-\d{3,4}', sku) or len(sku) > 25 or " " in sku:
            continue
        if sku in seen:
            continue
        seen.add(sku)
        parent, att_cod = _parse_parent(sku)
        padres.setdefault(parent, []).append({"sku": sku, "att_cod": (att_cod or "EST").upper()})

    total_skus = sum(len(v) for v in padres.values())
    print(f"  SKUs en ferraforme: {total_skus}")

    # Obtener templates del contenedor
    tmpl_ids = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.template", "search",
        [[["container_numbers", "=", cfg["container"]]]])
    tmpl_recs = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.template", "read",
        [tmpl_ids], {"fields": ["default_code"]}) if tmpl_ids else []

    # Mapa: parent_sku → tmpl_id
    tmpl_by_parent = {}
    for t in tmpl_recs:
        code = (t["default_code"] or "").strip()
        if code:
            parent, _ = _parse_parent(code)
            tmpl_by_parent[parent] = t["id"]

    # Para cada template, leer los product.template.attribute.value (ptav)
    # ptav tiene: product_attribute_value_id (el valor) → nos da el nombre del att_cod
    # y se usa para crear product.product con la combinacion correcta
    tmpl_ptav = {}  # tmpl_id → {att_cod_name: ptav_id}
    for tmpl_id in tmpl_ids:
        ptavs = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
            "product.template.attribute.value", "search_read",
            [[["product_tmpl_id", "=", tmpl_id]]],
            {"fields": ["product_attribute_value_id", "id"]})
        name_to_ptav = {}
        for ptav in ptavs:
            pav = ptav["product_attribute_value_id"]
            pav_id = pav[0] if isinstance(pav, (list, tuple)) else pav
            pav_name = pav[1] if isinstance(pav, (list, tuple)) and len(pav) > 1 else None
            if not pav_name:
                pav_rec = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                    "product.attribute.value", "read", [[pav_id]], {"fields": ["name"]})[0]
                pav_name = pav_rec["name"]
            name_to_ptav[pav_name.upper().strip()] = ptav["id"]
        tmpl_ptav[tmpl_id] = name_to_ptav

    # Obtener SKUs existentes en product.product
    existing_pp = {}  # sku → pp_id
    for tmpl_id in tmpl_ids:
        pp_recs = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.product", "search_read",
            [[["product_tmpl_id", "=", tmpl_id]]], {"fields": ["default_code"]})
        for pp in pp_recs:
            if pp["default_code"]:
                existing_pp[pp["default_code"].strip()] = pp["id"]

    print(f"  product.product existentes: {len(existing_pp)}")

    creados = ya_existe = errores = 0

    for parent_sku, variantes in padres.items():
        tmpl_id = tmpl_by_parent.get(parent_sku)
        if not tmpl_id:
            # Buscar por prefijo
            for code, tid in tmpl_by_parent.items():
                p2, _ = _parse_parent(code)
                if p2 == parent_sku:
                    tmpl_id = tid
                    break

        if not tmpl_id:
            for v in variantes:
                if v["sku"] not in existing_pp:
                    print(f"  ! Sin template para {parent_sku}")
                    errores += 1
                    break
            continue

        ptav_map = tmpl_ptav.get(tmpl_id, {})

        for v in variantes:
            sku = v["sku"]
            if sku in existing_pp:
                ya_existe += 1
                continue

            att_cod = v["att_cod"]
            ptav_id = ptav_map.get(att_cod)

            if not ptav_id:
                print(f"  ! Sin ptav para {sku} (att_cod={att_cod}, disponibles={list(ptav_map.keys())[:5]})")
                errores += 1
                continue

            if not _check_sku_libre(sku, cfg["container"]):
                errores += 1
                continue

            try:
                models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.product", "create",
                    [{"product_tmpl_id": tmpl_id,
                      "default_code": sku,
                      "product_template_attribute_value_ids": [(4, ptav_id)]}])
                print(f"  + {sku}")
                creados += 1
            except Exception as e:
                print(f"  ! Error {sku}: {e}")
                errores += 1

    print(f"  Creados: {creados} | Ya existian: {ya_existe} | Errores: {errores}")
    total_creados += creados
    total_ya_existe += ya_existe
    total_errores += errores

print(f"\nTOTAL: {total_creados} creados | {total_ya_existe} ya existian | {total_errores} errores")

if _sku_conflict_log:
    print(f"\n{'='*60}")
    print(f"CONFLICTOS DE SKU ({len(_sku_conflict_log)}) — SKUs NO creados por duplicado en otro contenedor:")
    for msg in _sku_conflict_log:
        print(msg)
