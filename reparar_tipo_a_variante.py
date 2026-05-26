"""
reparar_tipo_a_variante.py
Para cada template de los contenedores 71-78 cuya linea de atributo
usa un atributo con create_variant != 'always':
  1. Lee los valores (att_cods) de esa linea
  2. Borra la linea (archiva el product.product existente)
  3. Crea valores para el atributo "Variante" (create_variant=always)
  4. Crea nueva linea "Variante" con todos los valores
  5. Asigna SKUs a los nuevos product.product via PTAV matching

Requiere la info del ferraforme para saber que SKU va con cada att_cod.
"""
import sys, re
import xmlrpc.client
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ODOO_URL  = "https://ifull.odoo.com"
ODOO_DB   = "ifullmx-brea-main-6396587"
ODOO_USER = "valeria.rivero@kubera.mx"
ODOO_PASS = "Vale:kubera28"

common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
uid    = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASS, {})
models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)
print(f"Conectado uid={uid}")

BASE_DIR = Path(r"c:\Users\valer\OneDrive\Documentos\Desarrollo_Agentes_IA\skus listos")

CONFIGS = {
    "71": dict(file="OOCU8248653 PL contenedor 71.xlsx",      container="OOCU8248653 cont 71",  col_sku=3,  col_atributo=6,  col_ingles=7),
    "72": dict(file="CAAU5061672 PL contenedor 72.xlsx",      container="CAAU5061672 cont 72",  col_sku=3,  col_atributo=6,  col_ingles=7),
    "73": dict(file="CSNU6409280 PL contenedor 73.xlsx",      container="CSNU6409280 cont 73",  col_sku=2,  col_atributo=5,  col_ingles=6),
    "74": dict(file="LEX25-510_WHSU6230286_PL contenedor 74.xlsx", container="WHSU6230286 cont 74", col_sku=2, col_atributo=6, col_ingles=7),
    "75": dict(file="PRY25-543_GAOU7102197_PL contenedor 75.xlsx", container="GAOU7102197 cont 75", col_sku=2, col_atributo=5, col_ingles=7),
    "76": dict(file="LEX25-429_CSNU6006475_PL contenedor 76.xlsx", container="CSNU6006475 cont 76", col_sku=3, col_atributo=6, col_ingles=7),
    "77": dict(file="TONY25-275 TRHU6472913 Lista de empaque contenedor 77.xlsx", container="TRHU6472913 cont 77", col_sku=3, col_atributo=6, col_ingles=7),
    "78": dict(file="TONY25-284 TLLU5922910 Lista de empaque contenedor 78.xlsx", container="TLLU5922910 cont 78", col_sku=3, col_atributo=7, col_ingles=6),
}

only = sys.argv[1] if len(sys.argv) > 1 else None

# Cache del atributo "Variante"
def _get_variante_attr():
    ids = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute", "search",
                            [[["name", "=", "Variante"]]])
    if ids:
        rec = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute", "read",
                                [[ids[0]]], {"fields": ["create_variant"]})[0]
        if rec["create_variant"] != "always":
            models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute", "write",
                              [[ids[0]], {"create_variant": "always"}])
        return ids[0]
    return models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute", "create",
                             [{"name": "Variante", "create_variant": "always"}])

variante_attr_id = _get_variante_attr()
print(f"Atributo Variante id={variante_attr_id}")

_val_cache = {}
def _get_or_create_val(attr_id, valor):
    key = (attr_id, valor)
    if key in _val_cache:
        return _val_cache[key]
    ids = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute.value", "search",
                            [[["name", "=", valor], ["attribute_id", "=", attr_id]]])
    if ids:
        _val_cache[key] = ids[0]
        return ids[0]
    try:
        new_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute.value", "create",
                                   [{"name": valor, "attribute_id": attr_id}])
    except Exception:
        # Race: value was created between search and create — re-search
        ids2 = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute.value", "search",
                                 [[["name", "=", valor], ["attribute_id", "=", attr_id]]])
        new_id = ids2[0]
    _val_cache[key] = new_id
    return new_id


def _parse_parent(sku):
    parts = str(sku).strip().split("-")
    if len(parts) >= 3:
        return f"{parts[0]}-{parts[1]}", "-".join(parts[2:])
    return sku, None


_sku_conflict_log = []

def _check_sku_libre(sku, mi_container):
    """Devuelve True si el SKU está libre para asignar.
    Si ya existe en otro contenedor, lo registra en _sku_conflict_log y devuelve False.
    Si ya existe en el MISMO contenedor, también devuelve False (ya asignado, ok).
    """
    existentes = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.product", "search_read",
        [[["default_code", "=", sku], ["active", "=", True]]],
        {"fields": ["id", "product_tmpl_id"]})
    if not existentes:
        return True
    for pp in existentes:
        tmpl = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.template", "read",
            [[pp["product_tmpl_id"][0]]], {"fields": ["container_numbers", "name"]})[0]
        otro_cont = tmpl.get("container_numbers") or "sin contenedor"
        if otro_cont != mi_container:
            msg = f"  ⚠ CONFLICTO SKU {sku}: ya existe en '{otro_cont}' ({tmpl['name']}) pp_id={pp['id']} — omitido"
            print(msg)
            _sku_conflict_log.append(msg)
            return False
    return False  # ya existe en este mismo contenedor


total_reparados = total_ok = total_errores = 0

for num, cfg in CONFIGS.items():
    if only and num != only:
        continue

    print(f"\n{'='*60}")
    print(f"Cont {num}: {cfg['container']}")

    # Leer ferraforme: parent_sku → [{"sku", "att_cod"}]
    wb = openpyxl.load_workbook(str(BASE_DIR / cfg["file"]), data_only=True)
    ws = wb.active
    padres = {}
    seen = set()
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, cfg["col_sku"]).value
        if not cell:
            continue
        sku = str(cell).strip()
        if not re.match(r'^[A-Z]{2,6}-\d{3,4}', sku) or len(sku) > 25 or " " in sku:
            continue
        if sku in seen:
            continue
        seen.add(sku)
        parent, att_cod = _parse_parent(sku)
        padres.setdefault(parent, []).append({"sku": sku, "att_cod": (att_cod or "EST").upper()})

    # Obtener templates del contenedor
    tmpl_ids = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.template", "search",
        [[["container_numbers", "=", cfg["container"]]]])
    tmpl_recs = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.template", "read",
        [tmpl_ids], {"fields": ["default_code"]}) if tmpl_ids else []

    tmpl_by_parent = {}
    for t in tmpl_recs:
        code = (t["default_code"] or "").strip()
        if code:
            parent, _ = _parse_parent(code)
            tmpl_by_parent[parent] = t["id"]

    reparados = ok = errores = 0

    for tmpl_id in tmpl_ids:
        # Leer lineas de atributo del template
        lines = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
            "product.template.attribute.line", "search_read",
            [[["product_tmpl_id", "=", tmpl_id]]],
            {"fields": ["id", "attribute_id", "value_ids"]})

        if not lines:
            continue

        line = lines[0]
        attr_id_line = line["attribute_id"][0] if isinstance(line["attribute_id"], (list, tuple)) else line["attribute_id"]

        # Verificar si ya usa "Variante" con create_variant=always
        attr_rec = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute", "read",
            [[attr_id_line]], {"fields": ["create_variant", "name"]})[0]

        if attr_rec["create_variant"] == "always":
            ok += 1
            continue  # Ya esta bien

        # Necesita reparacion: cambiar a atributo "Variante"
        val_ids_old = line["value_ids"]
        line_id = line["id"]

        # Leer nombres de los valores actuales
        val_recs = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.attribute.value", "read",
            [val_ids_old], {"fields": ["name"]}) if val_ids_old else []
        att_cods_en_odoo = [r["name"].upper().strip() for r in val_recs]

        # Obtener template info para buscar en ferraforme
        tmpl = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.template", "read",
            [[tmpl_id]], {"fields": ["default_code"]})[0]
        tmpl_code = (tmpl["default_code"] or "").strip()
        parent, first_att = _parse_parent(tmpl_code)

        # Buscar variantes en ferraforme para este padre
        variantes_ferraforme = padres.get(parent, [])
        if not variantes_ferraforme and first_att:
            # Buscar por el att_cod del primer variant
            for p, vs in padres.items():
                if p == parent:
                    variantes_ferraforme = vs
                    break

        # Construir mapeo att_cod → sku desde ferraforme
        att_to_sku = {v["att_cod"]: v["sku"] for v in variantes_ferraforme}

        # Crear valores para "Variante"
        val_id_to_sku = {}
        for att_cod in att_cods_en_odoo:
            val_id = _get_or_create_val(variante_attr_id, att_cod)
            sku = att_to_sku.get(att_cod, f"{parent}-{att_cod}")
            val_id_to_sku[val_id] = sku

        if not val_id_to_sku:
            errores += 1
            continue

        try:
            # Borrar la linea con el atributo no-variant
            models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                              "product.template.attribute.line", "unlink", [[line_id]])

            # Crear nueva linea "Variante" con todos los valores a la vez
            new_line_id = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                "product.template.attribute.line", "create",
                [{"product_tmpl_id": tmpl_id,
                  "attribute_id": variante_attr_id,
                  "value_ids": [(4, v_id) for v_id in val_id_to_sku.keys()]}])

            # Asignar SKUs a los nuevos product.product via PTAV
            pp_records = models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.product", "search_read",
                [[["product_tmpl_id", "=", tmpl_id]]],
                {"fields": ["default_code", "product_template_attribute_value_ids"]})

            # Leer los nuevos PTAVs
            new_ptavs = models.execute_kw(ODOO_DB, uid, ODOO_PASS,
                "product.template.attribute.value", "search_read",
                [[["product_tmpl_id", "=", tmpl_id], ["attribute_line_id", "=", new_line_id]]],
                {"fields": ["id", "product_attribute_value_id"]})

            # Mapa ptav_id → sku
            ptav_to_sku = {}
            for ptav in new_ptavs:
                pav = ptav["product_attribute_value_id"]
                pav_id = pav[0] if isinstance(pav, (list, tuple)) else pav
                sku = val_id_to_sku.get(pav_id)
                if sku:
                    ptav_to_sku[ptav["id"]] = sku

            assigned = 0
            for pp in pp_records:
                if pp.get("default_code"):
                    continue
                ptav_ids = pp.get("product_template_attribute_value_ids", [])
                for ptav_id in ptav_ids:
                    sku = ptav_to_sku.get(ptav_id)
                    if sku:
                        if not _check_sku_libre(sku, cfg["container"]):
                            break
                        models.execute_kw(ODOO_DB, uid, ODOO_PASS, "product.product", "write",
                            [[pp["id"]], {"default_code": sku}])
                        assigned += 1
                        break

            n_pp = len(pp_records)
            n_exp = len(val_id_to_sku)
            status = "OK" if n_pp >= n_exp else f"WARN {n_pp}/{n_exp}"
            print(f"  [{status}] {parent} ({n_pp} pp, {assigned} SKUs asignados)")
            reparados += 1

        except Exception as e:
            print(f"  ! ERROR {parent}: {e}")
            errores += 1

    print(f"  Reparados: {reparados} | Ya OK: {ok} | Errores: {errores}")
    total_reparados += reparados
    total_ok += ok
    total_errores += errores

print(f"\nTOTAL: {total_reparados} reparados | {total_ok} ya OK | {total_errores} errores")

if _sku_conflict_log:
    print(f"\n{'='*60}")
    print(f"CONFLICTOS DE SKU ({len(_sku_conflict_log)}) — SKUs NO asignados por duplicado en otro contenedor:")
    for msg in _sku_conflict_log:
        print(msg)
