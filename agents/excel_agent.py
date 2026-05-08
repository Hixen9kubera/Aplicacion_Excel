"""
Agente Excel — Fase E de la arquitectura multi-agente FERRAFORME.

Responsabilidades:
- Generación de Excel de productos (FORMULA FERRAFORME)
- Excel maestro de costos con fórmulas
- Excel de Purchase Order
- Reporte de clasificación padre/variante

Sin dependencias de st.session_state — funciones puramente de datos + archivo.
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
try:
    from PIL import Image as PILImage
    _PILLOW_OK = True
except ImportError:
    PILImage   = None
    _PILLOW_OK = False

from agents.utils import _safe_float, IMAGENES_TEMP_PATH

# ── Constantes ────────────────────────────────────────────────────────────────

TEMPLATE_PATH          = Path(__file__).parent.parent / "FORMULA FERRAFORME PRODUCTOS .xlsx"
PURCHASE_TEMPLATE_PATH = Path(__file__).parent.parent / "Purchase.xlsx"
EMPRESA                = "FERRAFORME MS"


# ── Helpers CBM ───────────────────────────────────────────────────────────────

def _cbm_por_pieza(prod: dict) -> float:
    """Deriva CBM por pieza individual desde lo que traiga el packing list."""
    cbm_pz  = _safe_float(prod.get("cbm_por_pieza"))
    cbm_mc  = _safe_float(prod.get("cbm_master_carton"))
    cbm_tot = _safe_float(prod.get("cbm_total_sku"))
    pzs_caja = _safe_float(prod.get("piezas_x_caja"))
    pzs_tot  = _safe_float(prod.get("piezas_total")) or pzs_caja

    if cbm_pz > 0:
        return cbm_pz
    if cbm_mc > 0 and pzs_caja > 0:
        return cbm_mc / pzs_caja
    if cbm_tot > 0 and pzs_tot > 0:
        return cbm_tot / pzs_tot
    return 0.0


def _cbm_total_fila(prod: dict) -> float:
    """
    Deriva el CBM total de una fila según lo que traiga el packing list.
    Prioridad: cbm_total_sku > cbm_master_carton × cajas > cbm_por_pieza × piezas_total.
    """
    cbm_pz  = _safe_float(prod.get("cbm_por_pieza"))
    cbm_mc  = _safe_float(prod.get("cbm_master_carton"))
    cbm_tot = _safe_float(prod.get("cbm_total_sku"))
    cajas   = _safe_float(prod.get("cajas_master"))
    pzs_caja = _safe_float(prod.get("piezas_x_caja"))
    pzs_tot  = _safe_float(prod.get("piezas_total")) or pzs_caja

    if cbm_tot > 0:
        return cbm_tot
    if cbm_mc > 0:
        return cbm_mc * cajas if cajas > 0 else cbm_mc * (pzs_tot / pzs_caja if pzs_caja > 0 else 1)
    if cbm_pz > 0:
        return cbm_pz * pzs_tot
    return 0.0


# ── Generadores Excel ─────────────────────────────────────────────────────────

def generar_excel(productos: list[dict], tipo_cambio: float, contenedor: str) -> bytes:
    _no_fill   = PatternFill(fill_type=None)
    _no_border = Border(
        left=Side(style=None), right=Side(style=None),
        top=Side(style=None),  bottom=Side(style=None),
    )
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ultima = 3 + len(productos) - 1

    for r in range(3, ultima + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
    for r in range(ultima + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = _no_fill
            cell.border = _no_border
    if "Sheet2" in wb.sheetnames:
        del wb["Sheet2"]

    # Eliminar columnas largo, ancho, alto (originalmente cols 4, 5, 6)
    ws.delete_cols(4, 3)

    # Actualizar encabezados tras el delete
    ws.cell(2, 3).value  = "Imagen"
    ws.cell(2, 13).value = "Costo Landed USD"

    # Tamaño fijo para la columna de imagen (col 3) y filas de datos
    IMG_W_PX  = 120
    IMG_H_PX  = 120
    IMG_COL_W = 18
    ROW_H_PT  = 90
    ws.column_dimensions["C"].width = IMG_COL_W

    # Después del delete_cols las columnas quedan:
    # 1:nombre  2:sku  3:imagen  4:cbm_por_pieza  5:precio_usd  6:None
    # 7:piezas_x_caja  8:cbm_caja  9:tipo_producto  10:EMPRESA  11:None
    # 12:contenedor  13:costo_landed_usd  14:descripcion  15:categoria
    # 16:atributo  17-20:None
    for i, prod in enumerate(productos):
        r   = i + 3
        sku = prod.get("sku", "")
        ws.cell(r, 1).value = prod.get("nombre")
        ws.cell(r, 2).value = sku

        _img_incrustada = False
        if sku and IMAGENES_TEMP_PATH.exists():
            _archivo_img = next(
                (f for f in IMAGENES_TEMP_PATH.iterdir() if f.stem == sku),
                None,
            )
            if _archivo_img and _PILLOW_OK:
                try:
                    pil_img = PILImage.open(_archivo_img).convert("RGB")
                    pil_img.thumbnail((IMG_W_PX, IMG_H_PX), PILImage.LANCZOS)
                    _buf_img = io.BytesIO()
                    pil_img.save(_buf_img, format="PNG")
                    _buf_img.seek(0)
                    xl_img = XLImage(_buf_img)
                    xl_img.width  = IMG_W_PX
                    xl_img.height = IMG_H_PX
                    xl_img.anchor = f"C{r}"
                    ws.add_image(xl_img)
                    ws.row_dimensions[r].height = ROW_H_PT
                    _img_incrustada = True
                except Exception:
                    pass
        if not _img_incrustada:
            ws.cell(r, 3).value = ""
        cbm = ws.cell(r, 4)
        _cu = _cbm_por_pieza(prod)
        cbm.value = _cu if _cu > 0 else None
        if cbm.value is not None:
            cbm.number_format = "0.000000"
        ws.cell(r, 5).value  = prod.get("precio_usd")
        ws.cell(r, 6).value  = prod.get("piezas_total")
        ws.cell(r, 7).value  = prod.get("piezas_x_caja")
        cbm_caja = ws.cell(r, 8)
        px = prod.get("piezas_x_caja")
        cbm_caja.value = round(float(px) * _cu, 6) if px and _cu > 0 else None
        if cbm_caja.value is not None:
            cbm_caja.number_format = "0.000000"
        ws.cell(r, 9).value  = prod.get("tipo_producto", "Producto almacenable")
        ws.cell(r, 10).value = EMPRESA
        ws.cell(r, 11).value = None
        ws.cell(r, 12).value = contenedor
        ws.cell(r, 13).value = float(prod["precio_usd"]) if prod.get("precio_usd") else None
        ws.cell(r, 14).value = prod.get("descripcion")
        ws.cell(r, 15).value = prod.get("categoria")
        ws.cell(r, 16).value = prod.get("atributo")
        ws.cell(r, 17).value = None
        ws.cell(r, 18).value = None
        ws.cell(r, 19).value = None
        ws.cell(r, 20).value = None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_excel_master(productos: list[dict], tipo_cambio: float,
                         costo_contenedor: float, nombre_packing: str) -> bytes:
    """
    Genera el Excel maestro de costos con fórmulas visibles.
    Columnas A-R: identificación, CBM, parámetros de precio, costos calculados.
    Parámetros editables en cols T-U (tipo cambio, costo contenedor).
    """
    n = len(productos)
    DATA_START = 3
    DATA_END   = DATA_START + n - 1
    TOTAL_ROW  = DATA_END + 1

    T, U = 20, 21
    Tc = get_column_letter(T)
    Uc = get_column_letter(U)

    REF_TC  = f"${Uc}$2"
    REF_CC  = f"${Uc}$3"
    REF_CBM = f"${Uc}$4"
    REF_M3  = f"${Uc}$5"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Costos"

    _col_color = {}
    for color, cols in [
        ("1F3864", [1, 2, 3, 4, 5, 6]),
        ("1B6B3A", [7, 8, 9, 10]),
        ("2E4057", [11, 12]),
        ("7B2D00", [13, 14, 15, 16, 17, 18]),
    ]:
        for c in cols:
            _col_color[c] = color

    HEADERS = [
        "Packing List", "SKU", "Producto", "Cajas", "Piezas Total",
        "Piezas x Caja", "CBM Total", "CBM Master Carton", "CBM Inner Carton",
        "CBM por Pieza", "Tipo de Cambio", "Precio USD", "Precio MXN",
        "Costo Contenedor", "Costo por M³", "Costo CBM/Pieza",
        "Costo Unitario", "Costo Total",
    ]
    COL_WIDTHS = [26, 16, 34, 8, 13, 13, 14, 18, 16, 14, 13, 12, 14, 16, 14, 15, 14, 14]

    FMT_MXN  = '"$"#,##0.00'
    FMT_USD  = '#,##0.00'
    FMT_CBM  = '0.000000'
    FMT_CBM4 = '0.0000'
    FMT_INT  = '#,##0'
    FMT_NUM  = '#,##0.00'

    COL_FMT = {
        4: FMT_INT,  5: FMT_INT,  6: FMT_INT,
        7: FMT_CBM,  8: FMT_CBM,  9: FMT_CBM, 10: FMT_CBM,
        11: FMT_NUM, 12: FMT_USD, 13: FMT_MXN, 14: FMT_MXN,
        15: FMT_MXN, 16: FMT_MXN, 17: FMT_MXN, 18: FMT_MXN,
    }

    FILL_ODD        = PatternFill("solid", fgColor="F0F4FA")
    FILL_EVEN       = PatternFill("solid", fgColor="FFFFFF")
    FILL_HDR        = PatternFill("solid", fgColor="0D1F3C")
    FILL_TOTAL      = PatternFill("solid", fgColor="0D1F3C")
    FILL_PARAM_LBL  = PatternFill("solid", fgColor="2E4057")
    FILL_PARAM_VAL  = PatternFill("solid", fgColor="F0F4FA")
    FILL_PARAM_CALC = PatternFill("solid", fgColor="E8F5E9")

    FONT_TITLE = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    FONT_HDR   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    FONT_DATA  = Font(name="Calibri", size=10)
    FONT_TOTAL = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    FONT_PLBL  = Font(name="Calibri", bold=True, size=9,  color="FFFFFF")
    FONT_PVAL  = Font(name="Calibri", size=10)
    FONT_PCALC = Font(name="Calibri", italic=True, size=10, color="1B6B3A")

    ALIGN_CTR = Alignment(horizontal="center", vertical="center")
    ALIGN_LFT = Alignment(horizontal="left",   vertical="center")
    ALIGN_RGT = Alignment(horizontal="right",  vertical="center")

    border_hdr  = Border(left=Side(style="thin",   color="FFFFFF"),
                         right=Side(style="thin",  color="FFFFFF"),
                         bottom=Side(style="medium", color="FFFFFF"))
    border_data = Border(bottom=Side(style="thin", color="D9D9D9"),
                         right=Side(style="thin",  color="D9D9D9"))

    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    c = ws["A1"]
    c.value     = f"MASTER COSTOS — {nombre_packing}"
    c.font      = FONT_TITLE
    c.fill      = FILL_HDR
    c.alignment = ALIGN_CTR
    ws.row_dimensions[1].height = 22

    for ci, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=ci, value=header)
        c.fill      = PatternFill("solid", fgColor=_col_color.get(ci, "2E4057"))
        c.font      = FONT_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border_hdr
    ws.row_dimensions[2].height = 30

    for ci, ancho in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    ws.column_dimensions[Tc].width = 26
    ws.column_dimensions[Uc].width = 18

    ws.merge_cells(f"{Tc}1:{Uc}1")
    c = ws.cell(row=1, column=T, value="⚙ PARÁMETROS")
    c.font = FONT_TITLE; c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = ALIGN_CTR

    params = [
        ("Tipo de cambio (MXN/USD)",  tipo_cambio,       FMT_NUM,  FILL_PARAM_VAL,  FONT_PVAL,  False),
        ("Costo contenedor (MXN)",    costo_contenedor,  FMT_MXN,  FILL_PARAM_VAL,  FONT_PVAL,  False),
        ("CBM Contenedor (M³)",
         f"=SUM(G{DATA_START}:G{DATA_END})",
         FMT_CBM4,  FILL_PARAM_CALC, FONT_PCALC, True),
        ("Costo por M³ (MXN)",
         f"={Uc}3/{Uc}4",
         FMT_MXN,   FILL_PARAM_CALC, FONT_PCALC, True),
    ]
    for pi, (label, value, fmt, fill_v, font_v, _) in enumerate(params, start=2):
        lc = ws.cell(row=pi, column=T, value=label)
        lc.font = FONT_PLBL; lc.fill = FILL_PARAM_LBL
        lc.alignment = ALIGN_LFT
        lc.border = Border(bottom=Side(style="thin", color="FFFFFF"),
                           right=Side(style="thin",  color="FFFFFF"))
        vc = ws.cell(row=pi, column=U, value=value)
        vc.font = font_v; vc.fill = fill_v
        vc.alignment = ALIGN_RGT
        vc.number_format = fmt
        vc.border = Border(bottom=Side(style="thin", color="BBBBBB"))

    nota_row = 6
    ws.merge_cells(f"{Tc}{nota_row}:{Uc}{nota_row}")
    nc = ws.cell(row=nota_row, column=T,
                 value="U4 y U5 se calculan con fórmula — editables U2 y U3")
    nc.font      = Font(name="Calibri", italic=True, size=8, color="666666")
    nc.alignment = ALIGN_CTR

    for row_num, prod in enumerate(productos, start=DATA_START):
        r = row_num

        pzs_caja  = _safe_float(prod.get("piezas_x_caja"))
        pzs_tot   = _safe_float(prod.get("piezas_total")) or pzs_caja
        cajas     = _safe_float(prod.get("cajas_master"))
        precio    = _safe_float(prod.get("precio_usd"))
        cbm_pz    = _safe_float(prod.get("cbm_por_pieza"))
        cbm_mc    = _safe_float(prod.get("cbm_master_carton"))
        cbm_ic    = _safe_float(prod.get("cbm_inner_carton"))
        cbm_tot   = _safe_float(prod.get("cbm_total_sku"))

        if cbm_pz > 0:
            cbm_por_pieza_val  = cbm_pz
            cbm_total_fila_val = cbm_tot if cbm_tot > 0 else round(cbm_pz * pzs_tot, 6)
        elif cbm_mc > 0 and pzs_caja > 0:
            cbm_por_pieza_val  = round(cbm_mc / pzs_caja, 6)
            cbm_total_fila_val = cbm_tot if cbm_tot > 0 else (
                round(cbm_mc * cajas, 6) if cajas > 0 else round(cbm_por_pieza_val * pzs_tot, 6)
            )
        elif cbm_tot > 0 and pzs_tot > 0:
            cbm_por_pieza_val  = round(cbm_tot / pzs_tot, 6)
            cbm_total_fila_val = cbm_tot
        else:
            cbm_por_pieza_val  = 0
            cbm_total_fila_val = 0

        valores_directos = {
            1:  nombre_packing,
            2:  prod.get("sku", ""),
            3:  prod.get("nombre", ""),
            4:  int(cajas)    if cajas    == int(cajas)    else cajas,
            5:  int(pzs_tot)  if pzs_tot  == int(pzs_tot)  else pzs_tot,
            6:  int(pzs_caja) if pzs_caja == int(pzs_caja) else pzs_caja,
            7:  round(cbm_total_fila_val, 6) or None,
            8:  round(cbm_mc, 6) or None,
            9:  round(cbm_ic, 6) or None,
            10: round(cbm_por_pieza_val, 6) or None,
            12: precio if precio > 0 else None,
        }

        formulas = {
            11: f"={REF_TC}",
            13: f"=IF(L{r}=\"\",0,L{r}*{REF_TC})",
            14: f"={REF_CC}",
            15: f"={REF_M3}",
            16: f"=J{r}*{REF_M3}",
            17: f"=M{r}+P{r}",
            18: f"=Q{r}*E{r}",
        }

        fill = FILL_ODD if row_num % 2 == 1 else FILL_EVEN
        ws.row_dimensions[r].height = 18

        for ci in range(1, 19):
            valor = formulas.get(ci) if ci in formulas else valores_directos.get(ci)
            cell  = ws.cell(row=r, column=ci, value=valor)
            cell.fill      = fill
            cell.font      = FONT_DATA
            cell.border    = border_data
            cell.alignment = ALIGN_LFT if ci == 3 else ALIGN_CTR
            if ci in COL_FMT:
                cell.number_format = COL_FMT[ci]

    ws.row_dimensions[TOTAL_ROW].height = 20
    totales_vals = {
        1:  "TOTALES",
        5:  f"=SUM(E{DATA_START}:E{DATA_END})",
        7:  f"={REF_CBM}",
        18: f"=SUM(R{DATA_START}:R{DATA_END})",
    }
    totales_fmt = {5: FMT_INT, 7: FMT_CBM4, 18: FMT_MXN}

    for ci in range(1, 19):
        cell = ws.cell(row=TOTAL_ROW, column=ci, value=totales_vals.get(ci, ""))
        cell.fill      = FILL_TOTAL
        cell.font      = FONT_TOTAL
        cell.alignment = ALIGN_LFT if ci == 1 else ALIGN_CTR
        if ci in totales_fmt:
            cell.number_format = totales_fmt[ci]

    ws.freeze_panes = f"A{DATA_START}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_excel_purchase(productos: list[dict], tipo_cambio: float,
                           costo_contenedor: float) -> bytes:
    """
    Genera el Excel de Purchase Order a partir del template Purchase.xlsx.
    Llena: PRODUCT (SKU), QUANTITY (piezas_total), DESCRIPTION, PRICE (costo_unitario).
    Las columnas fijas (PURCHASE ID, SUPPLIER, CURRENCY, UOM, DATE) se toman del template.
    """
    cbm_total    = sum(_cbm_total_fila(p) for p in productos)
    costo_por_m3 = costo_contenedor / cbm_total if cbm_total > 0 else 0.0

    wb = openpyxl.load_workbook(PURCHASE_TEMPLATE_PATH)
    ws = wb.active

    _purchase_id = ws.cell(2, 1).value or ""
    _supplier    = ws.cell(2, 2).value or EMPRESA
    _currency    = ws.cell(2, 3).value or "MXN"
    _uom         = ws.cell(2, 6).value or "Unidades"
    _date        = ws.cell(2, 10).value

    for c in range(1, 11):
        ws.cell(2, c).value = None

    for i, prod in enumerate(productos):
        r = 2 + i

        precio_usd     = _safe_float(prod.get("precio_usd"))
        precio_mxn     = round(precio_usd * tipo_cambio, 2)
        cbm_pz         = _cbm_por_pieza(prod)
        costo_cbm_pz   = round(cbm_pz * costo_por_m3, 2)
        costo_unitario = round(precio_mxn + costo_cbm_pz, 2)

        piezas_total = prod.get("piezas_total")
        try:
            piezas_total = int(float(piezas_total)) if piezas_total else None
        except (ValueError, TypeError):
            piezas_total = None

        ws.cell(r, 1).value  = _purchase_id
        ws.cell(r, 2).value  = _supplier
        ws.cell(r, 3).value  = _currency
        ws.cell(r, 10).value = _date
        ws.cell(r, 4).value  = prod.get("sku", "")
        ws.cell(r, 5).value  = piezas_total
        ws.cell(r, 6).value  = _uom
        ws.cell(r, 7).value  = prod.get("descripcion", "")
        ws.cell(r, 8).value  = costo_unitario if costo_unitario > 0 else None
        ws.cell(r, 9).value  = None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generar_reporte_clasificacion(propuestas: list[dict], nombre_packing: str = "") -> bytes:
    """
    Genera Excel de reporte de clasificación para revisión de bodega.
    Columnas: Fila PL, Nombre, Nombre Base, SKU, Padre SKU, Atributo, Acción,
              Padre en Odoo, Requiere Revisión, Nota.
    """
    from openpyxl.styles import PatternFill as _PF, Font as _Ft, Alignment as _Al

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clasificación Productos"

    HEADERS = [
        "Fila PL", "Nombre producto", "Nombre base", "SKU propuesto",
        "SKU padre", "Tipo atributo", "Valor atributo",
        "Acción", "Padre en Odoo", "Requiere revisión", "Nota",
    ]
    hdr_fill = _PF("solid", fgColor="1F3864")
    hdr_font = _Ft(color="FFFFFF", bold=True, size=10)
    aln_ctr  = _Al(horizontal="center", vertical="center", wrap_text=True)
    aln_lft  = _Al(horizontal="left",   vertical="center", wrap_text=True)

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = aln_ctr

    FILL_REV  = _PF("solid", fgColor="FFF2CC")
    FILL_DUP  = _PF("solid", fgColor="E2EFDA")
    FILL_ODOO = _PF("solid", fgColor="DAE8FC")
    FILL_NONE = _PF(fill_type=None)

    for r, prop in enumerate(propuestas, 2):
        prod   = prop["producto"]
        fila   = str((prod.get("fila_excel_0idx") or 0) + 1)
        accion = prop["accion"]
        if prop["requiere_revision"]:
            fill = FILL_REV
        elif accion == "duplicado":
            fill = FILL_DUP
        elif prop["padre_fuente"] == "odoo":
            fill = FILL_ODOO
        else:
            fill = FILL_NONE

        if prop.get("atributo_tipo") and prop.get("atributo_valor"):
            att_tipo  = prop["atributo_tipo"]
            att_valor = prop["atributo_valor"]
        else:
            att_tipo  = prop.get("atributo_tipo") or "—"
            att_valor = prop.get("atributo_valor") or "—"

        valores = [
            fila,
            prop["nombre"],
            prop["nombre_base"],
            prop["sku"],
            prop["padre_sku"] or "—",
            att_tipo,
            att_valor,
            prop["accion_display"],
            prop["padre_odoo_nombre"] or "—",
            "⚠️ SÍ" if prop["requiere_revision"] else "",
            prop["nota_revision"] or "",
        ]
        for c, v in enumerate(valores, 1):
            cell = ws.cell(r, c, v)
            cell.fill      = fill
            cell.alignment = aln_lft if c in (2, 3, 11) else aln_ctr

    widths = [8, 30, 25, 16, 14, 14, 16, 26, 25, 16, 30]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
